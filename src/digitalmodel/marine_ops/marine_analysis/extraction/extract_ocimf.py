"""
OCIMF Database Extraction from Excel File.

This script extracts the complete OCIMF coefficient database from the Excel file
and creates a structured CSV database with all wind and current coefficients.

Excel Structure (ACTUAL):
- Wind CXw: Rows 5-17 (13 angles)
- Wind CYw: Rows 23-35 (13 angles)
- Wind CMw: Rows 42-54 (13 angles, labeled "Cxyw")
- Current CXc: Rows 60-161 (102 angles)
- Current CYc: Rows 110-145 (36 angles)
- Current CMc: Rows 149-184 (36 angles, labeled "CxYw")
"""

import openpyxl
import pandas as pd
import numpy as np
from pathlib import Path
from typing import List, Dict

# Excel file path
EXCEL_FILE = r"D:\workspace-hub\_temp\marine_analysis_data.xlsm"
OUTPUT_CSV = r"D:\workspace-hub\digitalmodel\data\ocimf_database.csv"


def extract_wind_coefficients(ws) -> pd.DataFrame:
    """Extract wind coefficients (CXw, CYw, CMw) from rows 5-54."""
    data = []

    # Section 1: CXw (rows 5-17)
    # Header row 4 has displacement ratios: 1.02, 1.05, 1.1, 3.0, etc.
    displacements_row4 = [ws.cell(4, j).value for j in range(2, 12)]

    for row in range(5, 18):  # Rows 5-17
        angle = ws.cell(row, 1).value
        if angle is None:
            continue

        for col_idx, disp in enumerate(displacements_row4):
            if disp is None:
                continue

            value = ws.cell(row, col_idx + 2).value
            if value is not None:
                # Determine vessel type from column position
                if col_idx < 4:
                    vessel_type = "Loaded Tanker"
                elif col_idx < 7:
                    vessel_type = "Ballasted Tanker"
                else:
                    vessel_type = "LNGC Carrier"

                data.append({
                    'vessel_type': vessel_type,
                    'displacement_ratio': float(disp),
                    'heading': float(angle),
                    'CXw': float(value)
                })

    df_cxw = pd.DataFrame(data)

    # Section 2: CYw (rows 23-35)
    data_cyw = []
    displacements_row22 = [ws.cell(22, j).value for j in range(2, 12)]

    for row in range(23, 36):  # Rows 23-35
        angle = ws.cell(row, 1).value
        if angle is None:
            continue

        for col_idx, disp in enumerate(displacements_row22):
            if disp is None:
                continue

            value = ws.cell(row, col_idx + 2).value
            if value is not None:
                if col_idx < 4:
                    vessel_type = "Loaded Tanker"
                elif col_idx == 4:
                    vessel_type = "Unknown"  # Column 6 has value 6
                elif col_idx < 8:
                    vessel_type = "Ballasted Tanker"
                else:
                    vessel_type = "LNGC Carrier"

                data_cyw.append({
                    'vessel_type': vessel_type,
                    'displacement_ratio': float(disp),
                    'heading': float(angle),
                    'CYw': float(value)
                })

    df_cyw = pd.DataFrame(data_cyw)

    # Section 3: CMw (rows 42-54)
    data_cmw = []
    displacements_row41 = [ws.cell(41, j).value for j in range(2, 12)]

    for row in range(42, 55):  # Rows 42-54
        angle = ws.cell(row, 1).value
        if angle is None:
            continue

        for col_idx, disp in enumerate(displacements_row41):
            if disp is None:
                continue

            value = ws.cell(row, col_idx + 2).value
            if value is not None:
                if col_idx < 4:
                    vessel_type = "Loaded Tanker"
                elif col_idx < 7:
                    vessel_type = "Ballasted Tanker"
                else:
                    vessel_type = "LNGC Carrier"

                data_cmw.append({
                    'vessel_type': vessel_type,
                    'displacement_ratio': float(disp),
                    'heading': float(angle),
                    'CMw': float(value)
                })

    df_cmw = pd.DataFrame(data_cmw)

    # Merge wind coefficients
    df_wind = df_cxw.merge(df_cyw, on=['vessel_type', 'displacement_ratio', 'heading'], how='outer')
    df_wind = df_wind.merge(df_cmw, on=['vessel_type', 'displacement_ratio', 'heading'], how='outer')

    return df_wind


def extract_current_coefficients(ws) -> pd.DataFrame:
    """Extract current coefficients (CXc, CYc, CMc)."""
    data = []

    # Section 1: CXc (rows 60-161)
    # Header row 59 has vessel types
    vessel_types_row59 = [ws.cell(59, j).value for j in range(2, 7)]

    for row in range(60, 162):  # Rows 60-161
        angle = ws.cell(row, 1).value
        if angle is None or not isinstance(angle, (int, float)):
            continue

        for col_idx, vessel in enumerate(vessel_types_row59):
            if vessel is None:
                continue

            value = ws.cell(row, col_idx + 2).value
            if value is not None and isinstance(value, (int, float)):
                data.append({
                    'vessel_type': str(vessel),
                    'heading': float(angle),
                    'CXc': float(value)
                })

    df_cxc = pd.DataFrame(data)

    # Section 2: CYc (rows 110-145)
    data_cyc = []
    vessel_types_row109 = [ws.cell(109, j).value for j in range(2, 7)]

    for row in range(110, 146):  # Rows 110-145
        angle = ws.cell(row, 1).value
        if angle is None or not isinstance(angle, (int, float)):
            continue

        for col_idx, vessel in enumerate(vessel_types_row109):
            if vessel is None:
                continue

            value = ws.cell(row, col_idx + 2).value
            if value is not None and isinstance(value, (int, float)):
                data_cyc.append({
                    'vessel_type': str(vessel),
                    'heading': float(angle),
                    'CYc': float(value)
                })

    df_cyc = pd.DataFrame(data_cyc)

    # Section 3: CMc (rows 149-184)
    data_cmc = []
    vessel_types_row148 = [ws.cell(148, j).value for j in range(2, 7)]

    for row in range(149, 185):  # Rows 149-184
        angle = ws.cell(row, 1).value
        if angle is None or not isinstance(angle, (int, float)):
            continue

        for col_idx, vessel in enumerate(vessel_types_row148):
            if vessel is None:
                continue

            value = ws.cell(row, col_idx + 2).value
            if value is not None and isinstance(value, (int, float)):
                data_cmc.append({
                    'vessel_type': str(vessel),
                    'heading': float(angle),
                    'CMc': float(value)
                })

    df_cmc = pd.DataFrame(data_cmc)

    # Merge current coefficients
    df_current = df_cxc.copy()
    if not df_cyc.empty:
        df_current = df_current.merge(df_cyc, on=['vessel_type', 'heading'], how='outer')
    if not df_cmc.empty:
        df_current = df_current.merge(df_cmc, on=['vessel_type', 'heading'], how='outer')

    return df_current


def extract_ocimf_database(excel_path: str, output_csv: str) -> pd.DataFrame:
    """Extract complete OCIMF database from Excel file."""
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['OCIMF (raw)']

    print(f"Sheet dimensions: {ws.max_row} rows × {ws.max_column} columns")

    # Extract wind coefficients
    print("\nExtracting wind coefficients...")
    df_wind = extract_wind_coefficients(ws)
    print(f"  Wind coefficients: {len(df_wind)} entries")
    print(f"  Unique vessels: {df_wind['vessel_type'].nunique()}")
    print(f"  Displacement ratios: {sorted(df_wind['displacement_ratio'].unique())}")

    # Extract current coefficients
    print("\nExtracting current coefficients...")
    df_current = extract_current_coefficients(ws)
    print(f"  Current coefficients: {len(df_current)} entries")
    print(f"  Unique vessels: {df_current['vessel_type'].nunique()}")

    # Combine wind and current data
    # Strategy: Create unified database by averaging current coefficients per heading
    print("\nMerging wind and current coefficients...")

    # Average current coefficients by heading
    df_current_avg = df_current.groupby('heading')[['CXc', 'CYc', 'CMc']].mean().reset_index()

    # Merge with wind data
    df_merged = df_wind.merge(df_current_avg, on='heading', how='left')

    # Add displacement column (estimate based on displacement ratio)
    # Assume typical VLCC displacement of ~300,000 tonnes for ratio = 1.0
    df_merged['displacement'] = df_merged['displacement_ratio'] * 300000

    # Sort by vessel type, displacement, and heading
    df_merged = df_merged.sort_values(['vessel_type', 'displacement', 'heading'])

    # Save to CSV
    Path(output_csv).parent.mkdir(parents=True, exist_ok=True)
    df_merged.to_csv(output_csv, index=False)

    print(f"\n[SUCCESS] Database saved to: {output_csv}")
    print(f"  Total entries: {len(df_merged)}")
    print(f"  Columns: {list(df_merged.columns)}")

    # Data quality check
    print("\nData Quality Check:")
    for col in ['CXw', 'CYw', 'CMw', 'CXc', 'CYc', 'CMc']:
        if col in df_merged.columns:
            missing = df_merged[col].isna().sum()
            total = len(df_merged)
            print(f"  {col}: {total - missing}/{total} values ({100*(total-missing)/total:.1f}% complete)")

    # Statistics
    print("\nCoefficient Statistics:")
    for col in ['CXw', 'CYw', 'CMw', 'CXc', 'CYc', 'CMc']:
        if col in df_merged.columns:
            values = df_merged[col].dropna()
            if len(values) > 0:
                print(f"  {col}: min={values.min():.3f}, max={values.max():.3f}, " +
                      f"mean={values.mean():.3f}, std={values.std():.3f}")

    return df_merged


if __name__ == "__main__":
    # Extract database
    df = extract_ocimf_database(EXCEL_FILE, OUTPUT_CSV)

    print("\n[SUCCESS] Extraction complete!")
    print(f"\nFirst 5 entries:\n{df.head()}")
    print(f"\nLast 5 entries:\n{df.tail()}")
