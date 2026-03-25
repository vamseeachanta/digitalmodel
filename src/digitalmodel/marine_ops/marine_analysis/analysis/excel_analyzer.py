"""
Excel File Analyzer for Marine Engineering Data
Analyzes XLSM files to extract worksheets, formulas, VBA code, and marine engineering features
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys
import json
from collections import defaultdict
import re

def analyze_excel_file(filepath):
    """Comprehensive analysis of Excel file structure and content"""

    print("=" * 80)
    print("MARINE ENGINEERING EXCEL FILE ANALYSIS REPORT")
    print("=" * 80)
    print(f"\nFile: {filepath}\n")

    try:
        # Load workbook
        wb = openpyxl.load_workbook(filepath, data_only=False, keep_vba=True)

        analysis = {
            'worksheets': [],
            'formulas_by_sheet': {},
            'data_structures': {},
            'marine_features': [],
            'vba_present': False,
            'named_ranges': [],
            'charts': []
        }

        # 1. WORKSHEET ANALYSIS
        print("\n" + "=" * 80)
        print("1. WORKSHEETS PRESENT")
        print("=" * 80)

        for idx, sheet_name in enumerate(wb.sheetnames, 1):
            sheet = wb[sheet_name]

            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column

            worksheet_info = {
                'name': sheet_name,
                'index': idx,
                'dimensions': f"{max_row} rows x {max_col} columns",
                'max_row': max_row,
                'max_col': max_col
            }

            analysis['worksheets'].append(worksheet_info)

            print(f"\n  [{idx}] {sheet_name}")
            print(f"      Dimensions: {max_row} rows x {max_col} columns")

            # Identify potential data types based on headers
            headers = []
            for col in range(1, min(max_col + 1, 26)):
                cell_value = sheet.cell(1, col).value
                if cell_value:
                    headers.append(str(cell_value))

            if headers:
                print(f"      Headers: {', '.join(headers[:10])}{'...' if len(headers) > 10 else ''}")

            worksheet_info['headers'] = headers

        # 2. FORMULA ANALYSIS
        print("\n" + "=" * 80)
        print("2. FORMULAS AND CALCULATIONS")
        print("=" * 80)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            formulas = []

            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 1000)):
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        # Handle both regular formulas and array formulas
                        formula_text = str(cell.value) if cell.value else ""
                        formulas.append({
                            'cell': cell.coordinate,
                            'formula': formula_text
                        })

            if formulas:
                analysis['formulas_by_sheet'][sheet_name] = formulas
                print(f"\n  Sheet: {sheet_name}")
                print(f"  Total formulas found: {len(formulas)}")

                # Show sample formulas
                for i, formula_info in enumerate(formulas[:5], 1):
                    print(f"    [{i}] {formula_info['cell']}: {formula_info['formula'][:100]}")

                if len(formulas) > 5:
                    print(f"    ... and {len(formulas) - 5} more formulas")

        # 3. MARINE ENGINEERING FEATURE DETECTION
        print("\n" + "=" * 80)
        print("3. MARINE ENGINEERING FEATURES DETECTED")
        print("=" * 80)

        marine_keywords = {
            'RAO': ['RAO', 'Response Amplitude Operator', 'Transfer Function'],
            'Motion Analysis': ['Surge', 'Sway', 'Heave', 'Roll', 'Pitch', 'Yaw', '6DOF', 'DOF'],
            'Hydrodynamics': ['Added Mass', 'Damping', 'Radiation', 'Diffraction', 'Hydrostatic'],
            'Wave Spectra': ['JONSWAP', 'Pierson-Moskowitz', 'Wave Spectrum', 'Spectral', 'Hs', 'Tp', 'Tz'],
            'Ship Dynamics': ['Vessel', 'Hull', 'Draft', 'Displacement', 'CoG', 'Center of Gravity'],
            'AQWA': ['AQWA', 'Hydrodynamic DB', 'Frequency Domain', 'Time Domain'],
            'OrcaFlex': ['OrcaFlex', 'Line', 'Mooring', 'Riser'],
            'Environmental': ['Wind', 'Current', 'Wave Height', 'Wave Period', 'Direction'],
            'Structural': ['Stress', 'Strain', 'Fatigue', 'S-N Curve', 'Damage']
        }

        feature_matches = defaultdict(list)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Search in cell values
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 100),
                                       max_col=min(sheet.max_column, 50)):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_value_upper = cell.value.upper()

                        for category, keywords in marine_keywords.items():
                            for keyword in keywords:
                                if keyword.upper() in cell_value_upper:
                                    feature_matches[category].append({
                                        'sheet': sheet_name,
                                        'cell': cell.coordinate,
                                        'keyword': keyword,
                                        'value': cell.value[:100]
                                    })

        for category, matches in sorted(feature_matches.items()):
            print(f"\n  {category}:")
            print(f"    Found {len(matches)} references")
            for match in matches[:3]:
                print(f"      - Sheet '{match['sheet']}', Cell {match['cell']}: {match['value']}")
            if len(matches) > 3:
                print(f"      ... and {len(matches) - 3} more")

        analysis['marine_features'] = dict(feature_matches)

        # 4. NAMED RANGES
        print("\n" + "=" * 80)
        print("4. NAMED RANGES")
        print("=" * 80)

        if wb.defined_names:
            named_ranges_list = list(wb.defined_names)
            print(f"\n  Total named ranges: {len(named_ranges_list)}")
            for idx, named_range_name in enumerate(named_ranges_list[:20], 1):
                named_range = wb.defined_names[named_range_name]
                print(f"    [{idx}] {named_range_name}: {named_range.value if hasattr(named_range, 'value') else 'N/A'}")
                analysis['named_ranges'].append({
                    'name': named_range_name,
                    'value': str(named_range.value) if hasattr(named_range, 'value') else 'N/A'
                })

            if len(named_ranges_list) > 20:
                print(f"    ... and {len(named_ranges_list) - 20} more")
        else:
            print("\n  No named ranges found")

        # 5. VBA MACROS
        print("\n" + "=" * 80)
        print("5. VBA MACROS")
        print("=" * 80)

        if wb.vba_archive:
            analysis['vba_present'] = True
            print("\n  [+] VBA macros are present in this file")
            print("  Note: VBA code extraction requires additional libraries (e.g., oletools)")
            print("  Recommend using: pip install oletools")
            print("  Then run: olevba marine_analysis_data.xlsm")
        else:
            print("\n  No VBA macros detected")

        # 6. DATA STRUCTURE ANALYSIS
        print("\n" + "=" * 80)
        print("6. DATA STRUCTURE SUMMARY")
        print("=" * 80)

        for sheet_info in analysis['worksheets']:
            sheet_name = sheet_info['name']
            sheet = wb[sheet_name]

            # Sample data types
            data_types = defaultdict(int)
            numeric_cols = []
            text_cols = []

            for col_idx in range(1, min(sheet.max_column + 1, 50)):
                col_letter = get_column_letter(col_idx)
                sample_values = []

                for row_idx in range(2, min(sheet.max_row + 1, 102)):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.value is not None:
                        sample_values.append(cell.value)

                if sample_values:
                    if all(isinstance(v, (int, float)) for v in sample_values[:10]):
                        numeric_cols.append(col_letter)
                    elif all(isinstance(v, str) for v in sample_values[:10]):
                        text_cols.append(col_letter)

            print(f"\n  Sheet: {sheet_name}")
            if numeric_cols:
                print(f"    Numeric columns: {', '.join(numeric_cols[:10])}")
            if text_cols:
                print(f"    Text columns: {', '.join(text_cols[:10])}")

            analysis['data_structures'][sheet_name] = {
                'numeric_columns': numeric_cols,
                'text_columns': text_cols
            }

        # 7. ENGINEERING MODEL IDENTIFICATION
        print("\n" + "=" * 80)
        print("7. ENGINEERING MODELS & CALCULATIONS IDENTIFIED")
        print("=" * 80)

        engineering_patterns = {
            'RAO Processing': [r'RAO', r'Transfer.*Function', r'Amplitude'],
            'Motion Calculations': [r'SQRT', r'SUM.*SQUARE', r'RMS'],
            'Statistical Analysis': [r'AVERAGE', r'STDEV', r'PERCENTILE', r'MAX', r'MIN'],
            'Integration/Differentiation': [r'SUMPRODUCT', r'TRAPZ'],
            'Matrix Operations': [r'MMULT', r'MINVERSE', r'TRANSPOSE'],
            'Interpolation': [r'FORECAST', r'TREND', r'LINEST', r'VLOOKUP', r'INTERPOLATE'],
            'Spectral Analysis': [r'FFT', r'POWER.*SPECTRUM', r'PSD'],
            'Time Series': [r'INDEX.*MATCH', r'OFFSET'],
        }

        model_findings = defaultdict(list)

        for sheet_name, formulas in analysis['formulas_by_sheet'].items():
            for formula_info in formulas:
                formula = formula_info['formula'].upper()

                for model_type, patterns in engineering_patterns.items():
                    for pattern in patterns:
                        if re.search(pattern, formula):
                            model_findings[model_type].append({
                                'sheet': sheet_name,
                                'cell': formula_info['cell']
                            })
                            break

        for model_type, findings in sorted(model_findings.items()):
            print(f"\n  {model_type}:")
            print(f"    Found in {len(findings)} cells")
            sheets = set(f['sheet'] for f in findings)
            print(f"    Sheets: {', '.join(sorted(sheets))}")

        # 8. PYTHON IMPLEMENTATION RECOMMENDATIONS
        print("\n" + "=" * 80)
        print("8. PYTHON IMPLEMENTATION RECOMMENDATIONS")
        print("=" * 80)

        recommendations = []

        if 'RAO' in feature_matches:
            recommendations.append({
                'module': 'RAO Processing Module',
                'description': 'Extract and process Response Amplitude Operators',
                'libraries': ['numpy', 'pandas', 'scipy.interpolate'],
                'features': ['RAO data loading', 'Interpolation', 'Transfer function calculations']
            })

        if 'Motion Analysis' in feature_matches:
            recommendations.append({
                'module': '6DOF Motion Analysis',
                'description': 'Ship motion calculations and analysis',
                'libraries': ['numpy', 'scipy', 'matplotlib'],
                'features': ['Surge/Sway/Heave calculations', 'Roll/Pitch/Yaw analysis', 'Motion statistics']
            })

        if 'Wave Spectra' in feature_matches:
            recommendations.append({
                'module': 'Wave Spectra Module',
                'description': 'Wave spectrum generation and analysis',
                'libraries': ['numpy', 'scipy.stats', 'matplotlib'],
                'features': ['JONSWAP spectrum', 'Pierson-Moskowitz spectrum', 'Spectral analysis']
            })

        if 'Hydrodynamics' in feature_matches:
            recommendations.append({
                'module': 'Hydrodynamic Coefficients',
                'description': 'Added mass, damping, and hydrostatic calculations',
                'libraries': ['numpy', 'pandas'],
                'features': ['Coefficient interpolation', 'Frequency-dependent analysis']
            })

        if model_findings.get('Statistical Analysis'):
            recommendations.append({
                'module': 'Statistical Analysis Tools',
                'description': 'Marine engineering statistical calculations',
                'libraries': ['numpy', 'scipy.stats', 'pandas'],
                'features': ['Extreme value analysis', 'Distribution fitting', 'Statistical summaries']
            })

        for idx, rec in enumerate(recommendations, 1):
            print(f"\n  [{idx}] {rec['module']}")
            print(f"      Description: {rec['description']}")
            print(f"      Required libraries: {', '.join(rec['libraries'])}")
            print(f"      Key features: {', '.join(rec['features'])}")

        print("\n" + "=" * 80)
        print("ANALYSIS COMPLETE")
        print("=" * 80)

        # Save analysis to JSON
        output_file = filepath.replace('.xlsm', '_analysis.json')

        # Convert defaultdict to regular dict for JSON serialization
        analysis_for_json = {
            'worksheets': analysis['worksheets'],
            'formulas_count_by_sheet': {k: len(v) for k, v in analysis['formulas_by_sheet'].items()},
            'marine_features_count': {k: len(v) for k, v in analysis['marine_features'].items()},
            'named_ranges_count': len(analysis['named_ranges']),
            'vba_present': analysis['vba_present'],
            'python_recommendations': recommendations
        }

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(analysis_for_json, f, indent=2)

        print(f"\nDetailed analysis saved to: {output_file}")

        return analysis

    except Exception as e:
        print(f"\nError analyzing file: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        filepath = r"D:\workspace-hub\_temp\marine_analysis_data.xlsm"

    analyze_excel_file(filepath)
