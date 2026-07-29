"""I/O primitives shared by the one-shot reference-catalog extractor."""

import csv
from contextlib import contextmanager
from decimal import Decimal, InvalidOperation
import hashlib
from pathlib import Path
import os
import unicodedata

from openpyxl import load_workbook

from ._reference_catalog_schema import PROHIBITED_PATTERNS


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def source_path(root: Path, relative: Path) -> Path:
    root = root.resolve()
    candidate = (root / relative).resolve()
    if not candidate.is_relative_to(root):
        raise ValueError(f"source path escapes source root: {relative}")
    if not candidate.is_file():
        raise FileNotFoundError(candidate)
    return candidate


def text(value) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def decimal_text(value) -> str:
    if value is None or text(value).upper() in {"NULL", "N/A"}:
        return ""
    try:
        number = Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError(f"invalid numeric catalog value: {value!r}") from exc
    if not number.is_finite():
        raise ValueError(f"non-finite catalog value: {value!r}")
    return format(number.normalize(), "f")


def is_numeric(value) -> bool:
    try:
        parsed = decimal_text(value)
    except ValueError:
        return False
    return bool(parsed)


def read_rows(
    path: Path, sheet_name: str, first_row=2, header_row=1, stop_at_blank=False
):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = []
    for _ in range(header_row):
        headers = [text(value) for value in next(rows)]
    records = []
    for source_row, values in enumerate(rows, header_row + 1):
        has_value = any(value is not None for value in values)
        if stop_at_blank and source_row >= first_row and not has_value:
            break
        if source_row < first_row or not has_value:
            continue
        records.append((source_row, dict(zip(headers, values))))
    dimensions = (worksheet.max_row, worksheet.max_column)
    workbook.close()
    return records, dimensions


def scan_safe(rows: list[dict], catalog_name: str) -> None:
    for row in rows:
        for value in row.values():
            if not isinstance(value, str):
                continue
            if contains_prohibited(value):
                raise ValueError(f"prohibited content in {catalog_name}")


def contains_prohibited(value: str) -> bool:
    return any(pattern.search(value) for pattern in PROHIBITED_PATTERNS)


def write_csv(path: Path, fields: list[str], rows: list[dict], metadata: dict):
    with path.open("w", newline="", encoding="utf-8") as stream:
        for key, value in metadata.items():
            stream.write(f"# {key}: {value}\n")
        writer = csv.DictWriter(stream, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


@contextmanager
def output_lock(output: Path):
    lock = output.with_name(f".{output.name}.lock")
    descriptor = os.open(lock, os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o600)
    try:
        yield
    finally:
        os.close(descriptor)
        lock.unlink(missing_ok=True)
