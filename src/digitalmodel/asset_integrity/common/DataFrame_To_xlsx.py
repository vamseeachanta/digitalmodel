import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.borders import BORDER_THIN, Border, Side

"""
https://xlsxwriter.readthedocs.io/working_with_pandas.html
"""
def DataFrame_To_xlsx_xlsxwriter(df, data):
    import xlrd
    writer = pd.ExcelWriter(data['FileName'], engine='xlsxwriter')

    try:
        # WorkSheet = wb.get_sheet_by_name(data['SheetName'])
        WorkSheet = wb[data['SheetName']]
    except:
        wb.create_sheet(data['SheetName'])
        WorkSheet = wb[data['SheetName']]


    df.to_excel(writer, data['SheetName'])
    writer._save()

"""
References:
https://stackoverflow.com/questions/42370977/how-to-save-a-new-sheet-in-an-existing-excel-file-using-pandas/42371251
https://stackoverflow.com/questions/36814050/openpyxl-get-sheet-by-name
"""
def DataFrame_To_xlsx_openpyxl(df, data):
    try:
        wb = load_workbook(data['FileName'])
        writer = pd.ExcelWriter(data['FileName'], engine = 'openpyxl')
        writer.wb = wb
    except:
        writer = pd.ExcelWriter(data['FileName'])

    try:
        # WorkSheet = wb.get_sheet_by_name(data['SheetName'])
        WorkSheet = wb[data['SheetName']]
    except:
        wb.create_sheet(data['SheetName'])
        WorkSheet = wb[data['SheetName']]

        #  For xlsxwriter
        # WorkSheet = wb.add_worksheet(data['SheetName'])

    df.to_excel(writer, data['SheetName'])
    writer._save()
"""
References:
https://stackoverflow.com/questions/24917201/applying-borders-to-a-cell-in-openpyxl
"""
def DataFrameArray_To_xlsx_openpyxl(dfArray, data):
    # try:
    #     #  Load existing workbook
    #     wb = load_workbook(filename = data['FileName'])
    #     writer = pd.ExcelWriter
    #     writer.wb = wb
    # except:
    print("Opening new workbook")
    writer = pd.ExcelWriter(data['FileName'], engine = 'openpyxl')

    for dfIndex in range(0, len(dfArray)):
        dfArray[dfIndex].to_excel(writer, data['SheetNames'][dfIndex])
        # property cell.border should be used instead of cell.style.border
        # if data['thin_border']:
        #     ws = wb.active
        #     thin_border = Border(
        #         left=Side(border_style=BORDER_THIN, color='00000000'),
        #         right=Side(border_style=BORDER_THIN, color='00000000'),
        #         top=Side(border_style=BORDER_THIN, color='00000000'),
        #         bottom=Side(border_style=BORDER_THIN, color='00000000')
        #     )
        #     rows = len(dfArray[dfIndex])+1
        #     columns = len(dfArray[dfIndex].columns)+1
        #     ws.cell(row=rows, column=columns).border = thin_border

    writer._save()

if __name__ == '__main__':
    import numpy as np

    df = pd.DataFrame(np.random.random([3, 3]), 
        columns=['A', 'B', 'C'], index=['first', 'second', 'third'])
    data = {"FileName" : "example.xlsx",
            "SheetName": 'Sheet1'}
    DataFrame_To_xlsx(df, data)

