#
# ---------------------------------------------------------------------------
# OUTPUT LAT LONG PADS.py
# Created on: 2015-12-25 11:46:00.00000
# Craeted by: Jose Marrero
# Description: Program from the series of excel tools for surface planning.
#              OUTPUT THE LAT LONG IN EXCEL.
# Updates:
#       Updated to use inidividual coordinate system from each well 10/24/2017
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Reference Websites
#   Insert Cursor (arcpy.da) at 10.1                http://anothergisblog.blogspot.com/2012/07/insert-cursor-arcpyda-at-101.html
#   Extending the example for multiple polylines    https://www.e-education.psu.edu/geog485/book/export/html/139
#   List of spatial reference factory code here     http://resources.arcgis.com/en/help/main/10.1/018z/pdf/projected_coordinate_systems.pdf
# ---------------------------------------------------------------------------
# Imports from spreadsheet
#   -Spreadsheet Location
#   -SPatial Reference ( arreglado) * No puede tener espacios en el nombre del excel
#   - Falta Wells Per Pad
# ---------------------------------------------------------------------------
import logging
import os
import time

start_time = time.time()
StartDay = time.strftime("%y%m%d", time.localtime( start_time ))

#-----SET UP LOGGING-----
logDirectory = r'C:\Data\log'
logger = logging.getLogger(os.path.basename(logDirectory))
logger.setLevel(logging.DEBUG)

logName = "log_{}.log".format(StartDay)
#-create the logging file handler
fh = logging.FileHandler(filename=os.path.join(logDirectory, logName),mode='a+')
#-Create formatter settings
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
#-Set Log Level for this handler
fh.setLevel(logging.DEBUG)
#-add handler to logger object
logger.addHandler(fh)


import xlrd

logger.info("xlrd")
from openpyxl import Workbook

logger.info("openpyxl Workbook")
from openpyxl import load_workbook

logger.info("openpyxlload_workbook")
from xlrd import open_workbook

logger.info("xlrd open_workbook")
import arcpy

logger.info("arcpy")
import os

logger.info("os")
import csv
import datetime
import sys
import traceback

logger.info("csv")
import struct

import openpyxl

arcpy.env.overwriteOutput = True
# ---------------------------------------------------------------------------


arcpy.env.overwriteOutput = True






# ---------------------------------------------------------------------------
# Set up inputs to tool
SPRDLOC = sys.argv[1]
TEMPLOC = sys.argv[2]
SRCODE = int(sys.argv[3])


##SPRDLOC = r'S:\WCPO\FDP Dev Team\00. Team\Clean Templates\EXPORT_TOOL_NEW_SDE_Excel_2016_NM.xlsm'
##TEMPLOC = r'S:\WCPO\FDP Dev Team\00. Team\Clean Templates'
##SRCODE = int(32039)


logger.info("Start")


#print('Location: ' + str(SPRDLOC))
#print('Code: ' + str(SRCODE))
#time.sleep(5.5) #pause for 5.5 seconds
pointsworkplace = r"C:\Users\Public\Data\SURFACE_PLANNING_TOOLS\SCRATCH_WORKPLACE\targetslatlongconvertion.shp" #r"in_memory\\points1.shp" #r"C:\Users\marreroj\Desktop\delet\Points.shp"
#F:\Jose\Template\V4\Shapefiles\Points.shp

spatialRef = arcpy.SpatialReference(SRCODE)
desc = arcpy.Describe(pointsworkplace)
shapefieldname = desc.ShapeFieldName
cline = 1

Coor_dic={'GCS_WGS_1984': 4326, 'GCS_North_American_1927': 4267, 'GCS_North_American_1983': 4269, 'NAD_1927_StatePlane_Texas_North_FIPS_4201': 32037, 'NAD_1927_StatePlane_Texas_North_Central_FIPS_4202': 32038, 'NAD_1927_StatePlane_Texas_Central_FIPS_4203': 32039, 'NAD_1927_StatePlane_Texas_South_Central_FIPS_4204': 32040, 'NAD_1927_StatePlane_Texas_South_FIPS_4205': 32041, 'NAD_1983_StatePlane_Texas_North_FIPS_4201_Feet1': 2275, 'NAD_1983_StatePlane_Texas_North_Central_FIPS_4202_Feet': 2276, 'NAD_1983_StatePlane_Texas_Central_FIPS_4203_Feet': 2277, 'NAD_1983_StatePlane_Texas_South_Central_FIPS_4204_Feet': 2278, 'NAD_1983_StatePlane_Texas_South_FIPS_4205_Feet': 2279, 'NAD_1983_StatePlane_New_Mexico_East_FIPS_3001_Feet': 2257, 'NAD_1927_StatePlane_New_Mexico_East_FIPS_3001': 32012,'NAD_1927_StatePlane_New_Mexico_Central_FIPS_3002': 32013}


#var1 = sys.argv[1]
# ---------------------------------------------------------------------------


# Create an empty array object
vertexArray = arcpy.Array()

# ---------------------------------------------------------------------------
# Function to add a polyline
def addPolygon(cursor, array, sr, dte,cline):
    #print(str(array) +"  "+ str(Padname))
    polyline = arcpy.Polygon(array, sr, True)
    cursor.insertRow([Padname,paddesign,padrotation,dte,polyline])
    #arcpy.AddMessage('line number: '+ str(cline))
    #cline = cline + 1
    array.removeAll()
# ---------------------------------------------------------------------------

c =1
padname = ""

workbook = xlrd.open_workbook(SPRDLOC, on_demand = True)
worksheet = workbook.sheet_by_name('WELLS SUMMARY')
#wtworkbook = xlwt.Workbook(SPRDLOC) #xlwt.open_workbook(SPRDLOC, on_demand = True)#open_workbook(SPRDLOC,formatting_info=True)
#wtsheet = workbook.get_sheet("PLAT")
#Sorting Excel
#data = [worksheet.row_values(i) for i in xrange(worksheet.nrows)]
#labels = data[0]    # Don't sort our headers
#data = data[1:]     # Data begins on the second row
#data.sort(key=lambda x: x[0])

#delete
with arcpy.da.UpdateCursor(pointsworkplace, ['Name','SHAPE@']) as cursor:

    for row in cursor:
        cursor.deleteRow()

with arcpy.da.InsertCursor(pointsworkplace, ['Name','TARGET','SHAPE@'] ) as cursor:
# Loop through the lines in the file and get each coordinate
    for row in range(1,worksheet.nrows):
        #padname = worksheet.cell(row,1).value
        #dte = datetime.datetime.now()
        #paddesign = worksheet.cell(row,7).value
        #print row, " ", 18
        WELLNAME = worksheet.cell(row,1).value
        #print WELLNAME
        SHL_X = worksheet.cell(row,8).value
        SHL_Y = worksheet.cell(row,9).value

        PP_X = worksheet.cell(row,10).value
        PP_Y = worksheet.cell(row,11).value

        FTP_X = worksheet.cell(row,12).value
        FTP_Y = worksheet.cell(row,13).value

        LTP_X = worksheet.cell(row,14).value
        LTP_Y = worksheet.cell(row,15).value

        EOT_X = worksheet.cell(row,16).value
        EOT_Y = worksheet.cell(row,17).value

        COORDINATE_SYSTEM = worksheet.cell(row,79).value
        SRCODE=Coor_dic[COORDINATE_SYSTEM]
        spatialRef = arcpy.SpatialReference(SRCODE)

        isnew = padname != worksheet.cell(row,19).value
        #print(isnew)
        if isnew != 0:
            #print("Yes")
            #print row
            if vertexArray.count > 0:
                #padname = worksheet.cell(row,1).value
                #addPolygon(cursor, vertexArray, spatialRef, padname, paddesign, padrotation, dte, cline)
                #padname = worksheet.cell(row,1).value
                #print "yes2"
                vertex = arcpy.Point(float(worksheet.cell(row,19).value),float(worksheet.cell(row,20).value)) #xyz vertex = arcpy.Point(float(fields[1]),float(fields[2]),-1*float(fields[0]))
                #print(str(worksheet.cell(row,19).value) + "  "+ str(worksheet.cell(row,20).value))

            #SHL
            vertex = arcpy.Point(float(SHL_X),float(SHL_Y))
            #print( str(SHL_X) + "  "+ str(SHL_Y))
            cursor.insertRow([WELLNAME,"SHL",arcpy.PointGeometry(vertex,spatialRef)])

            #KOP/PP
            vertex = arcpy.Point(float(PP_X),float(PP_Y))
            #print( str(KOP_X) + "  "+ str(KOP_Y))
            cursor.insertRow([WELLNAME,"PP",arcpy.PointGeometry(vertex,spatialRef)])

            #FTP
            vertex = arcpy.Point(float(FTP_X),float(FTP_Y))
            #print( str(HEEL_X) + "  "+ str(HEEL_Y))
            cursor.insertRow([WELLNAME,"FTP",arcpy.PointGeometry(vertex,spatialRef)])

            #LTP
            vertex = arcpy.Point(float(LTP_X),float(LTP_Y))
            #print( str(HEEL_X) + "  "+ str(HEEL_Y))
            cursor.insertRow([WELLNAME,"LTP",arcpy.PointGeometry(vertex,spatialRef)])

            #TOE/EOT
            vertex = arcpy.Point(float(EOT_X),float(EOT_Y))
            #print( str(TOE_X) + "  "+ str(TOE_Y))
            cursor.insertRow([WELLNAME,"EOT",arcpy.PointGeometry(vertex,spatialRef)])


workbook2 = openpyxl.Workbook()
worksheet2 = workbook2.active
#worksheet['A2'] = 'Hello world2!'

spatialRef2 = arcpy.SpatialReference(4267) #GCS_North_American_1927 WKID =  4267
rows = arcpy.da.SearchCursor(pointsworkplace, ['Name','TARGET','SHAPE@'])
shlrow = 1
pprow = 1
ftprow = 1
ltprow = 1
eotrow = 1

logger.info("Write To excel....")

for val, row in enumerate (rows):
    #print row, val
    #print row[1]
    pnt_lat_long = row[2].projectAs(spatialRef2) #row[2].projectAs(spatialRef2)
    #print pnt_lat_long.centroid.X, pnt_lat_long.centroid.Y
    newrow = val + 1

    if row[1] == "SHL":
        logger.info(str(pnt_lat_long.centroid.X))
        worksheet2.cell(row = shlrow , column = 2).value = str(pnt_lat_long.centroid.X)
        worksheet2.cell(row = shlrow , column = 1).value = str(pnt_lat_long.centroid.Y)
        shlrow = shlrow + 1
    if row[1] == "PP":
        worksheet2.cell(row = pprow , column = 4).value = str(pnt_lat_long.centroid.X)
        worksheet2.cell(row = pprow , column = 3).value = str(pnt_lat_long.centroid.Y)
        pprow = pprow + 1
    if row[1] == "FTP":
        worksheet2.cell(row = ftprow , column = 6).value = str(pnt_lat_long.centroid.X)
        worksheet2.cell(row = ftprow , column = 5).value = str(pnt_lat_long.centroid.Y)
        ftprow = ftprow + 1
    if row[1] == "LTP":
        worksheet2.cell(row = ltprow , column = 8).value = str(pnt_lat_long.centroid.X)
        worksheet2.cell(row = ltprow , column = 7).value = str(pnt_lat_long.centroid.Y)
        ltprow = ltprow + 1
    if row[1] == "EOT":
        worksheet2.cell(row = eotrow , column = 10).value = str(pnt_lat_long.centroid.X)
        worksheet2.cell(row = eotrow , column = 9).value = str(pnt_lat_long.centroid.Y)
        eotrow = eotrow + 1

TEMPLOC = os.path.join(TEMPLOC,"", "tempfile.xlsx")
print TEMPLOC
#workbook2.save(filename = TEMPLOC)
workbook2.save(filename = r'C:\Data\log\tempfile.xlsx')



#delete
with arcpy.da.UpdateCursor(pointsworkplace, ['Name','SHAPE@']) as cursor:

    for row in cursor:
        cursor.deleteRow()
print "Complete"
    #print
    #expression3 = "!SHAPE.truecentroid@DECIMALDEGREES!.split()[1]"
##    expression3 = "!SHAPE.firstpoint@DECIMALDEGREES!.split()[0]"
##    arcpy.CalculateField_management(pointsworkplace,'LAT', expression3, "PYTHON")
    #arcpy.CalculateField_management(pointsworkplace,'LAT', "!SHAPE.truecentroid@DECIMALDEGREES!.split()[0]", "PYTHON")

##for row in rows:
##    shape=row.getValue(dsc.shapeFieldName)
##    geom = shape.getPart(3)
##    x = geom.X
##    y = geom.Y
##    row.setValue('LONG_DD', x)
##    row.setValue('LAT_DD', y)
##    cursor.updateRow(row)

    #pnt= row[1].trueCentroid.X

    #print pnt


        #vertexArray.add(vertex)
##        cursor.insertRow([target,vertex])
        #feat = row.getValue(desc.OIDFieldName)
##        pnt = feat.getPart()
##        print pnt.X, pnt.Y
##        print(str(padname) + "  "+ str(worksheet.cell(row,19).value) + "  "+ str(worksheet.cell(row,20).value))

##    cline = cline +1
##    addPolygon(cursor, vertexArray, spatialRef, padname, paddesign, padrotation, dte, cline)

#for v in worksheet.col_slice(colx=0):
    #print (list(v))
    #for row in range(worksheet.nrows):
        #print( str(worksheet.cell(row,0).value) + "    "+ str(worksheet.cell(row,1).value) + "    "+ str(worksheet.cell(row,3).value) + "    "+ str(worksheet.cell(row,4).value))
#user_input = raw_input("Place input here: ")

#print(var1)
    #print('FINISHED')
#time.sleep(15.5) #pause for 5.5 seconds
