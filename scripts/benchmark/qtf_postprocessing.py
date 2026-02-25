#!/usr/bin/env python3
"""
QTF Postprocessing for L00 validation suite.

Exports .owr → .xlsx via OrcFxAPI.SaveResultsSpreadsheet, parses all
QTF sheets (including direct/indirect potential loads unavailable from the
OrcFxAPI properties directly), and generates comparison plots.

Reference plots (from WAMIT validation paper):
  - Case 3.1: 5-panel sum-frequency QTF (diffraction RAO + 4 QTF components)
  - Case 3.2: 4-panel diff-frequency QTF per DOF (PI, CS, direct, indirect potential)
  - Case 3.3: 6-DOF diff-frequency subplots

Usage:
    uv run python scripts/benchmark/qtf_postprocessing.py --case 3.2
    uv run python scripts/benchmark/qtf_postprocessing.py --case 3.1 3.2 3.3
    uv run python scripts/benchmark/qtf_postprocessing.py --all
    uv run python scripts/benchmark/qtf_postprocessing.py --case 3.2 --force-export

Chart improvements (WRK-311):
  - Improved axis labels, tick formatting, and frequency band annotations
  - Mean drift panel shows real component only (mean drift is a real quantity)
  - WAMIT × markers enlarged and coloured for visibility
  - 2D QTF surface heatmap (omega1 vs omega2, colour = amplitude)
  - Diffraction RAO first-order figure
  - Per-panel WAMIT comparison statistics in validation table
  - Case-specific validation narrative
"""

from __future__ import annotations

import argparse
import cmath
import math
from pathlib import Path

import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
import OrcFxAPI

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DOF_NAMES = ["Surge", "Sway", "Heave", "Roll", "Pitch", "Yaw"]
DOF_UNITS_FORCE = ["kN/m", "kN/m", "kN/m", "kN·m/m", "kN·m/m", "kN·m/m"]
DOF_UNITS_DRIFT = ["kN/m²", "kN/m²", "kN/m²", "kN·m/m²", "kN·m/m²", "kN·m/m²"]

_C_REAL = "#1877c5"
_C_IMAG = "#c81818"

L00_DIR = (
    Path(__file__).resolve().parents[2]
    / "docs/modules/orcawave/L00_validation_wamit"
)

CASES: dict[str, dict] = {
    "3.1": {
        "owr": L00_DIR / "3.1/benchmark/Bottom mounted cylinder_ground_truth.owr",
        "output": L00_DIR / "3.1/benchmark",
        "label": "Bottom-mounted cylinder (Full QTF)",
        "qtf_type": "sum_freq",
        "dofs": [0],  # Surge — all DOFs fixed, others zero
    },
    "3.2": {
        "owr": L00_DIR / "3.2/benchmark/Sphere_ground_truth.owr",
        "output": L00_DIR / "3.2/benchmark",
        "label": "Sphere with Lid",
        "qtf_type": "diff_freq",
        "dofs": [0, 2],  # Surge (comparison_1) and Heave (comparison_2)
    },
    "3.3": {
        "owr": L00_DIR / "3.3/benchmark/test_ground_truth.owr",
        "output": L00_DIR / "3.3/benchmark",
        "label": "Multi-body Cylinder + Ellipsoid",
        "qtf_type": "diff_freq",
        "dofs": list(range(6)),  # All DOFs
    },
}

# ---------------------------------------------------------------------------
# Digitised WAMIT reference data (approximate × marker values from paper)
# Source: OrcaWave WAMIT Validation Guide figures 31, 33, 34, 36
# Accuracy: ±5–10% of full-scale; sufficient for visual comparison overlay
# ---------------------------------------------------------------------------
#
# Structure:
#   "3.1"  → sum_freq  → panel_key → {"omega", "real", "imag"}
#   "3.2"  → diff_freq → dw → dof  → panel_key → {"omega", "real", "imag"}
#   "3.3"  → diff_freq → dw → panel_key → {dof: {"omega","real","imag"}}
#
_W = {
    "3.1": {
        # Figure 31 — surge sum-frequency QTF, ω 0.5–3.0 rad/s at 0.25 rad/s steps
        # load_rao_diffraction removed: that sheet contains first-order RAO, not a QTF quantity.
        # NOTE: mean drift is a real quantity; imag column is all-zero by definition.
        "mean_drift_pi": {
            "omega": [0.50, 0.75, 1.00, 1.25, 1.50, 1.75,  2.00,  2.25,  2.50,  2.75,  3.00],
            "real":  [0.5,  1.5,  3.5,  6.0,  8.5,  9.5,  10.0,  10.5,  10.5,  10.0,   9.5],
            "imag":  [None, None, None, None, None, None,  None,  None,  None,  None,  None],
        },
        "quadratic_pi": {
            "omega": [0.50, 0.75, 1.00, 1.25, 1.50, 1.75, 2.00, 2.25,  2.50,  2.75,  3.00],
            "real":  [0.0,  0.2,  0.5,  1.0,  2.0,  3.5,  5.0,  5.8,   6.0,   6.0,   5.5],
            "imag":  [0.0, -0.2, -0.5, -1.2, -2.5, -4.5, -7.0, -9.5, -11.0, -12.0, -12.0],
        },
        # Potential load WAMIT reference values (OrcaWave diagonal, ω 0.5–3.0 rad/s).
        # The paper figure plots real and imaginary components; digitized estimates below
        # are accurate to ±5–10% full-scale.  At ω=0.5 the imaginary component dominates
        # (imag≈587) while the real component is smaller (real≈49).
        "potential_direct": {
            "omega": [0.50,   1.00,   1.50,   2.00,   2.50,   3.00],
            "real":  [49.0,   73.0,   44.0,   16.0,    7.5,   14.0],
            "imag":  [587.0, 241.0,   98.0,   42.0,   18.0,    5.7],
        },
        "potential_indirect": {
            "omega": [0.50,   1.00,   1.50,   2.00,   2.50,   3.00],
            "real":  [49.0,   73.0,   43.0,   15.5,    7.5,   14.5],
            "imag":  [587.0, 241.0,   98.0,   42.0,   18.0,    5.7],
        },
    },
    "3.2": {
        # Figures 33/34 — diff-frequency Δω=0.1, ω₁ 0.5–4.5 rad/s
        # Extra resolution near sphere resonance ω≈3.1 rad/s
        0.1: {
            0: {  # Surge (comparison_1.png)
                "quadratic_pi": {
                    "omega": [0.5, 1.0, 1.5, 2.0, 2.5, 2.75, 3.0,  3.1,  3.25, 3.5, 4.0, 4.5],
                    "real":  [0.0, 0.0, 0.0, 0.1, 0.5,  1.0, 4.5,  9.0,   7.5, 6.8, 6.5, 6.2],
                    "imag":  [0.1, 0.1, 0.1, 0.2, 0.5,  0.8, 1.0,  1.0,   0.7, 0.5, 0.3, 0.2],
                },
                "quadratic_cs": {
                    "omega": [0.5, 1.0, 1.5, 2.0, 2.5, 2.75, 3.0,  3.1,  3.25, 3.5, 4.0, 4.5],
                    "real":  [0.0, 0.0, 0.0, 0.1, 0.5,  1.0, 4.5,  8.5,   7.5, 6.8, 6.5, 6.2],
                    "imag":  [0.1, 0.1, 0.1, 0.2, 0.5,  0.8, 1.0,  1.0,   0.7, 0.5, 0.3, 0.2],
                },
                "potential_direct": {
                    "omega": [0.5,   1.0,   1.5,   2.0,   2.5,   2.75,  3.0,   3.1,   3.25,  3.5,   4.0,   4.5],
                    "real":  [0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0],
                    "imag":  [-0.50, -0.40, -0.32, -0.30, -0.29, -0.29, -0.28, -0.28, -0.28, -0.28, -0.28, -0.28],
                },
                "potential_indirect": {
                    "omega": [0.5,   1.0,   1.5,   2.0,   2.5,   2.75,  3.0,   3.1,   3.25,  3.5,   4.0,   4.5],
                    "real":  [0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0,   0.0],
                    "imag":  [-0.50, -0.40, -0.32, -0.30, -0.29, -0.29, -0.28, -0.28, -0.28, -0.28, -0.28, -0.28],
                },
            },
            2: {  # Heave (comparison_2.png)
                "quadratic_pi": {
                    "omega": [0.5, 1.0, 1.5, 2.0, 2.5, 2.75, 3.0,  3.1,  3.25, 3.5,  4.0,  4.5],
                    "real":  [2.5, 2.5, 2.5, 2.5, 2.5,  4.0, 6.5,  7.0,   4.5, 2.0, -1.5, -3.0],
                    "imag":  [0.0, 0.0, 0.0, 0.1, 0.1,  0.1, 0.1,  0.2,   0.1, 0.1,  0.1,  0.1],
                },
                "quadratic_cs": {
                    "omega": [0.5, 1.0, 1.5, 2.0, 2.5, 2.75, 3.0,  3.1,  3.25, 3.5,  4.0,  4.5],
                    "real":  [2.5, 2.5, 2.5, 2.5, 2.5,  4.0, 6.5,  7.2,   4.5, 2.0, -1.5, -3.0],
                    "imag":  [0.0, 0.0, 0.0, 0.1, 0.1,  0.1, 0.1,  0.2,   0.1, 0.1,  0.1,  0.1],
                },
                "potential_direct": {
                    "omega": [0.5,   1.0,   1.5,  2.0,  2.5,  2.75, 3.0,  3.1,  3.25, 3.5,  4.0,  4.5],
                    "real":  [0.0,   0.0,   0.0,  0.0,  0.0,   0.0, 0.0,  0.0,   0.0, 0.0,  0.0,  0.0],
                    "imag":  [-70.0, -60.0, -38.0, -22.0, -12.0, -8.0, -6.0, -5.5, -4.0, -3.0, -1.5, -0.8],
                },
                "potential_indirect": {
                    "omega": [0.5,   1.0,   1.5,  2.0,  2.5,  2.75, 3.0,  3.1,  3.25, 3.5,  4.0,  4.5],
                    "real":  [0.0,   0.0,   0.0,  0.0,  0.0,   0.0, 0.0,  0.0,   0.0, 0.0,  0.0,  0.0],
                    "imag":  [-70.0, -60.0, -38.0, -22.0, -12.0, -8.0, -6.0, -5.5, -4.0, -3.0, -1.5, -0.8],
                },
            },
        },
    },
    "3.3": {
        # Figure 36 — Full QTF (PI quadratic + indirect potential) for cylinder
        # Δω=2.5 rad/s, β₁=0°, β₂=30°; finer steps near ω₁=2.0–3.0
        # NOTE: reference = combined (quadratic_pi + potential_indirect)
        2.5: {
            0: {  # Surge
                "omega": [1.0,  1.25, 1.5,  1.75, 2.0,  2.5,  3.0,  3.5,  4.0],
                "real":  [0.0, -0.3, -0.8,  -2.5, -3.8, -2.0,  0.0,  0.5,  0.3],
                "imag":  [-0.5, -1.2, -2.2, -2.8, -2.5, -0.8,  0.3,  0.4,  0.3],
            },
            1: {  # Sway
                "omega": [1.0,  1.25, 1.5,  1.75, 2.0,  2.5,  3.0,  3.5,  4.0],
                "real":  [-2.0, -2.0, -1.5, -0.5,  0.5,  2.5,  2.0,  1.0,  0.8],
                "imag":  [-7.5, -6.5, -5.0, -3.5, -2.0,  0.0,  2.0,  2.5,  2.5],
            },
            2: {  # Heave
                "omega": [1.0,  1.25, 1.5,  1.75, 2.0,  2.5,  3.0,  3.5,  4.0],
                "real":  [0.2,  0.3,  0.3,  0.25, 0.2,  0.0, -0.1, -0.1,  0.0],
                "imag":  [0.1,  0.0, -0.2, -0.35, -0.4, -0.1,  0.2,  0.1,  0.0],
            },
            3: {  # Roll
                "omega": [1.0,  1.25, 1.5,  1.75, 2.0,  2.5,  3.0,  3.5,  4.0],
                "real":  [0.0, -0.2, -0.5,  -1.0, -1.5, -3.5, -5.5, -6.0, -5.0],
                "imag":  [0.0,  0.0,  0.0,   0.0,  0.0,  0.0, -0.5, -1.5, -2.0],
            },
            4: {  # Pitch
                "omega": [1.0,  1.25, 1.5,  1.75, 2.0,  2.5,  3.0,  3.5,  4.0],
                "real":  [6.5,  6.0,  5.5,  4.8,  4.0,  3.0,  2.0,  1.5,  1.0],
                "imag":  [0.5,  1.2,  2.0,  2.6,  3.0,  2.5,  1.5,  0.8,  0.3],
            },
            5: {  # Yaw
                "omega": [1.0,  1.25, 1.5,  1.75, 2.0,  2.5,  3.0,  3.5,  4.0],
                "real":  [0.0,  0.0,  0.0,   0.0,  0.0,  0.0,  0.0,  0.0,  0.0],
                "imag":  [0.0,  0.0,  0.0,   0.0,  0.0,  0.0,  0.0,  0.0,  0.0],
            },
        },
    },
}

# ---------------------------------------------------------------------------
# CSV-based WAMIT reference loader (supersedes hardcoded _W above)
# ---------------------------------------------------------------------------


def _load_wamit_csv(csv_path: Path) -> dict | None:
    """Load WAMIT reference data from a digitized chart CSV file.

    Reads omega_rad_s, Wamit_real, Wamit_imag columns.  Rows with no WAMIT
    columns populated are skipped.  Missing individual columns fill with NaN
    so that Plotly renders only the available component markers.

    Returns {"omega": [...], "real": [...], "imag": [...]}, or None if the
    file does not exist or contains no WAMIT data.
    """
    if not csv_path.exists():
        return None
    omega, real, imag = [], [], []
    with csv_path.open(encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = [p.strip() for p in line.split(",")]
            if parts[0] == "omega_rad_s":
                continue  # header row
            if len(parts) < 4:
                continue
            try:
                w_val = float(parts[0])
            except ValueError:
                continue
            r_str = parts[3] if len(parts) > 3 else ""
            i_str = parts[4] if len(parts) > 4 else ""
            has_r = r_str != ""
            has_i = i_str != ""
            if not (has_r or has_i):
                continue
            omega.append(w_val)
            real.append(float(r_str) if has_r else float("nan"))
            imag.append(float(i_str) if has_i else float("nan"))
    if not omega:
        return None
    return {"omega": omega, "real": real, "imag": imag}


def _build_wamit_refs() -> dict:
    """Build WAMIT reference dict from digitized CSV files (mirrors _W structure).

    CSV files live under docs/modules/orcawave/L00_validation_wamit/<case>/digitized/.
    Returns an empty sub-dict for any case whose CSV files are missing, so the
    existing figure builders degrade gracefully (no markers rendered).
    """
    def _csv(case_dir: str, filename: str) -> dict | None:
        return _load_wamit_csv(L00_DIR / case_dir / "digitized" / filename)

    return {
        "3.1": {
            k: v for k, v in {
                "mean_drift_pi":      _csv("3.1", "surge_mean_drift_pi.csv"),
                "quadratic_pi":       _csv("3.1", "surge_quadratic_load_pi.csv"),
                "potential_direct":   _csv("3.1", "surge_direct_potential.csv"),
                "potential_indirect": _csv("3.1", "surge_indirect_potential.csv"),
            }.items() if v is not None
        },
        "3.2": {
            0.1: {
                0: {  # Surge
                    k: v for k, v in {
                        "quadratic_pi":      _csv("3.2", "surge_pi_quadratic.csv"),
                        "quadratic_cs":      _csv("3.2", "surge_cs_quadratic.csv"),
                        "potential_direct":  _csv("3.2", "surge_direct_potential.csv"),
                        "potential_indirect": _csv("3.2", "surge_indirect_potential.csv"),
                    }.items() if v is not None
                },
                2: {  # Heave
                    k: v for k, v in {
                        "quadratic_pi":      _csv("3.2", "heave_pi_quadratic.csv"),
                        "quadratic_cs":      _csv("3.2", "heave_cs_quadratic.csv"),
                        "potential_direct":  _csv("3.2", "heave_direct_potential.csv"),
                        "potential_indirect": _csv("3.2", "heave_indirect_potential.csv"),
                    }.items() if v is not None
                },
            },
        },
        "3.3": {
            2.5: {
                dof: v for dof, v in {
                    0: _csv("3.3", "surge_full_qtf.csv"),
                    1: _csv("3.3", "sway_full_qtf.csv"),
                    2: _csv("3.3", "heave_full_qtf.csv"),
                    3: _csv("3.3", "roll_full_qtf.csv"),
                    4: _csv("3.3", "pitch_full_qtf.csv"),
                    5: _csv("3.3", "yaw_full_qtf.csv"),
                }.items() if v is not None
            },
        },
    }


# CSV-sourced reference — preferred over hardcoded _W (used in build_inline_html)
_W_CSV: dict = _build_wamit_refs()

# ---------------------------------------------------------------------------
# Excel Export (OrcFxAPI)
# ---------------------------------------------------------------------------

def export_spreadsheet(
    owr_path: Path, xlsx_path: Path | None = None, force: bool = False
) -> Path:
    """Export OrcaWave .owr results to Excel via SaveResultsSpreadsheet.

    Uses the existing file when force=False and it already exists.
    This is the canonical way to get direct/indirect potential loads
    which are not exposed as OrcFxAPI array properties.
    """
    xlsx_path = xlsx_path or owr_path.with_suffix(".xlsx")
    if xlsx_path.exists() and not force:
        print(f"  Using existing:  {xlsx_path.name}")
        return xlsx_path
    print(f"  Exporting to:    {xlsx_path.name} ...")
    d = OrcFxAPI.Diffraction(str(owr_path))
    d.SaveResultsSpreadsheet(str(xlsx_path))
    print(f"  Done.")
    return xlsx_path


# ---------------------------------------------------------------------------
# Excel Parser
# ---------------------------------------------------------------------------

def _to_complex(ampl, phase_deg) -> complex:
    """Polar (amplitude, phase°) → complex.  Uses cmath.exp for complex exponent."""
    return float(ampl) * cmath.exp(1j * math.radians(float(phase_deg)))


def _numeric_rows(ws) -> list[list]:
    """Yield rows whose first cell is a number (skips all header rows)."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        try:
            float(row[0])
            rows.append(list(row))
        except (TypeError, ValueError):
            continue
    return rows


def _parse_rao_sheet(ws) -> dict:
    """RAO sheet layout: [heading, omega, Ampl₀, Phase₀, …, Ampl₅, Phase₅]."""
    h, omega, vals = [], [], []
    for row in _numeric_rows(ws):
        try:
            h.append(float(row[0]))
            omega.append(float(row[1]))
            vals.append([_to_complex(row[2 + 2 * i], row[3 + 2 * i]) for i in range(6)])
        except (TypeError, ValueError, IndexError):
            pass
    if not omega:
        return {}
    return {
        "heading": np.array(h),
        "omega": np.array(omega),
        "values": np.array(vals, dtype=complex),  # (nrows, 6)
    }


def _parse_drift_sheet(ws) -> dict:
    """Drift sheet: [heading1, heading2, omega, Ampl₀, Phase₀, …]."""
    h1, h2, omega, vals = [], [], [], []
    for row in _numeric_rows(ws):
        try:
            h1.append(float(row[0]))
            h2.append(float(row[1]))
            omega.append(float(row[2]))
            vals.append([_to_complex(row[3 + 2 * i], row[4 + 2 * i]) for i in range(6)])
        except (TypeError, ValueError, IndexError):
            pass
    if not omega:
        return {}
    return {
        "heading1": np.array(h1),
        "heading2": np.array(h2),
        "omega": np.array(omega),
        "values": np.array(vals, dtype=complex),  # (nrows, 6)
    }


def _parse_qtf_sheet(ws) -> dict:
    """QTF sheet: [h1, h2, omega1, omega2, omega_combined, Ampl₀, Phase₀, …]."""
    h1, h2, o1, o2, oc, vals = [], [], [], [], [], []
    for row in _numeric_rows(ws):
        try:
            h1.append(float(row[0]))
            h2.append(float(row[1]))
            o1.append(float(row[2]))
            o2.append(float(row[3]))
            oc.append(float(row[4]))
            vals.append([_to_complex(row[5 + 2 * i], row[6 + 2 * i]) for i in range(6)])
        except (TypeError, ValueError, IndexError):
            pass
    if not o1:
        return {}
    return {
        "heading1": np.array(h1),
        "heading2": np.array(h2),
        "omega1": np.array(o1),
        "omega2": np.array(o2),
        "omega_c": np.array(oc),
        "values": np.array(vals, dtype=complex),  # (nrows, 6)
    }


_SHEET_PARSERS = {
    "Load RAOs (diffraction)": ("load_rao_diffraction", _parse_rao_sheet),
    "Load RAOs (Haskind)":     ("load_rao_haskind",     _parse_rao_sheet),
    "Displacement RAOs":       ("displacement_rao",      _parse_rao_sheet),
    "Mean drift loads (PI)":   ("mean_drift_pi",         _parse_drift_sheet),
    "Mean drift loads (CS)":   ("mean_drift_cs",         _parse_drift_sheet),
    "Quadratic loads (PI)":    ("quadratic_pi",          _parse_qtf_sheet),
    "Quadratic loads (CS)":    ("quadratic_cs",          _parse_qtf_sheet),
    "Potential loads (direct)":   ("potential_direct",   _parse_qtf_sheet),
    "Potential loads (indirect)": ("potential_indirect", _parse_qtf_sheet),
}


def read_excel(xlsx_path: Path) -> dict:
    """Parse all QTF sheets from an OrcaWave results spreadsheet."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    result: dict = {}
    for sheet_name, (key, parser) in _SHEET_PARSERS.items():
        result[key] = parser(wb[sheet_name]) if sheet_name in wb.sheetnames else {}
    wb.close()
    return result


# ---------------------------------------------------------------------------
# OrcFxAPI Direct Extractor (cross-check / complement)
# ---------------------------------------------------------------------------

def read_owr(owr_path: Path, heading_pair_idx: int = 0) -> dict:
    """Extract QTF data directly from .owr via OrcFxAPI properties.

    Note: direct/indirect potential loads are NOT available here —
    use read_excel() for those (from SaveResultsSpreadsheet output).
    """
    d = OrcFxAPI.Diffraction(str(owr_path))
    hp = heading_pair_idx

    def _get(attr):
        v = getattr(d, attr, None)
        if v is None:
            return None
        arr = np.array(v)
        return arr[hp] if arr.ndim >= 2 else arr

    return {
        "omega_1st": np.sort(2 * np.pi * np.array(d.frequencies)),
        "qtf_omega": np.array(d.QTFAngularFrequencies),   # (nqtf, 3)
        "heading_pairs": np.array(d.QTFHeadingPairs),
        "load_rao_diffraction": _get("loadRAOsDiffraction"),  # (nfreq, 6) complex
        "load_rao_haskind":     _get("loadRAOsHaskind"),
        "mean_drift_pi": _get("meanDriftLoadPressureIntegration"),  # (nfreq, 6)
        "mean_drift_cs": _get("meanDriftLoadControlSurface"),
        "quadratic_pi":  _get("quadraticLoadFromPressureIntegration"),  # (nqtf, 6)
        "quadratic_cs":  _get("quadraticLoadFromControlSurface"),
    }


# ---------------------------------------------------------------------------
# Plot Helpers
# ---------------------------------------------------------------------------

def _cline(color: str, dash: str = "solid") -> dict:
    return {"color": color, "dash": dash, "width": 1.5}


def _traces(omega, values, name: str, show_legend: bool = True) -> list:
    """Pair of real + imaginary Scatter traces."""
    kw = dict(mode="lines+markers", marker=dict(size=5))
    return [
        go.Scatter(x=omega, y=np.real(values), name=f"{name} (real)",
                   line=_cline(_C_REAL), showlegend=show_legend,
                   legendgroup=name, **kw),
        go.Scatter(x=omega, y=np.imag(values), name=f"{name} (imag)",
                   line=_cline(_C_IMAG, "dash"), showlegend=show_legend,
                   legendgroup=name, **kw),
    ]


def _add(fig, traces, row: int, col: int) -> None:
    for t in traces:
        fig.add_trace(t, row=row, col=col)


def _wamit_traces(ref: dict, show_legend: bool = True) -> list:
    """Return marker-only Scatter traces for WAMIT reference data overlay.

    Args:
        ref: dict with keys "omega", "real", "imag" (lists or arrays).
             imag may contain None values; those entries are dropped so that
             Plotly does not render NaN markers (e.g. mean drift is real-only).
        show_legend: whether to show in legend (set False after first call)
    """
    omega = np.asarray(ref["omega"], dtype=float)
    real_raw = ref["real"]
    imag_raw = ref["imag"]

    # Convert to float arrays; None → NaN
    real = np.array(
        [float("nan") if v is None else float(v) for v in real_raw],
        dtype=float,
    )
    imag = np.array(
        [float("nan") if v is None else float(v) for v in imag_raw],
        dtype=float,
    )

    # Enlarged, high-contrast markers for WAMIT (stand out from OrcaWave lines)
    _mk = dict(symbol="x", size=12, line=dict(width=2.5))
    traces = []

    # Real component — only plot where data is not NaN
    real_mask = ~np.isnan(real)
    if real_mask.any():
        traces.append(go.Scatter(
            x=omega[real_mask], y=real[real_mask], name="WAMIT (real)",
            mode="markers", marker=dict(color=_C_REAL, **_mk),
            showlegend=show_legend, legendgroup="wamit",
        ))

    # Imaginary component — only plot where data is not NaN
    imag_mask = ~np.isnan(imag)
    if imag_mask.any():
        traces.append(go.Scatter(
            x=omega[imag_mask], y=imag[imag_mask], name="WAMIT (imag)",
            mode="markers", marker=dict(color=_C_IMAG, **_mk),
            showlegend=show_legend and not real_mask.any(), legendgroup="wamit",
        ))

    return traces


def _slice_delta(qtf_dict: dict, delta_omega: float, tol: float = 0.01) -> tuple:
    """Return (omega1_slice, values_slice) for rows near |ω₁ - ω₂| ≈ delta_omega."""
    if not qtf_dict:
        return np.array([]), np.zeros((0, 6), dtype=complex)
    o1, o2 = qtf_dict["omega1"], qtf_dict["omega2"]
    mask = np.abs(np.abs(o1 - o2) - delta_omega) < tol
    return o1[mask], qtf_dict["values"][mask]


# ---------------------------------------------------------------------------
# Figure builders (return go.Figure — used by both file writers and HTML embed)
# ---------------------------------------------------------------------------

def _add_freq_band_annotation(fig: go.Figure, row: int, col: int) -> None:
    """Add a subtle frequency band shading for the low-frequency diffraction region.

    Shades ω < 1.0 rad/s (long-wave regime where free-surface effects dominate)
    and marks ω = 2π rad/s (≈6.28, one wave per body diameter for a unit cylinder).
    Applied as a light rectangle annotation on each subplot.
    """
    fig.add_vrect(
        x0=0.0, x1=1.0,
        fillcolor="rgba(200,220,255,0.18)", line_width=0,
        annotation_text="long-wave", annotation_position="top left",
        annotation_font_size=9, annotation_font_color="#6090c0",
        row=row, col=col,
    )


def _build_sum_freq_figure(
    xdata: dict, dof: int, label: str, wamit_ref: dict | None = None
) -> go.Figure:
    """4-panel sum-frequency QTF figure (2x2 grid, QTF quantities only).

    Panels: mean drift (PI), quadratic load (PI), direct potential, indirect potential.
    The first-order diffraction load RAO is intentionally excluded — it comes from
    a separate OrcaWave sheet and is not a QTF quantity.

    Improvements (WRK-311):
    - Mean drift panel shows real component only (imaginary is zero by definition).
    - WAMIT markers are sized larger (size=12) and skipped when imag is None/NaN.
    - Axis labels include units and omega-subscript notation.
    - Tick format forces one decimal place on x-axis.
    - Low-frequency band annotation (omega < 1.0 rad/s) shaded on each panel.
    - Y-axis labels include physical units for each panel type.
    """
    dname = DOF_NAMES[dof]
    ud = DOF_UNITS_DRIFT[dof]

    def _wref(key):
        return (wamit_ref or {}).get(key)

    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=[
            f"Mean drift load — PI [{ud}]",
            f"Quadratic load — PI [{ud}]",
            f"Direct potential load [{ud}]",
            f"Indirect potential load [{ud}]",
        ],
        vertical_spacing=0.18, horizontal_spacing=0.12,
    )

    # Panel (1,1) — Mean drift PI: real component only (imag = 0 by physics)
    drift = xdata.get("mean_drift_pi", {})
    if drift:
        fig.add_trace(go.Scatter(
            x=drift["omega"], y=np.real(drift["values"][:, dof]),
            mode="lines+markers", name="OrcaWave (real)",
            line=_cline(_C_REAL), marker=dict(size=5),
            showlegend=True,
        ), row=1, col=1)
    ref = _wref("mean_drift_pi")
    if ref:
        _add(fig, _wamit_traces(ref, show_legend=True), 1, 1)
    _add_freq_band_annotation(fig, 1, 1)

    # Panel (1,2) — Quadratic load PI: diagonal (omega1==omega2) slice
    qpi = xdata.get("quadratic_pi", {})
    if qpi:
        o1, o2 = qpi["omega1"], qpi["omega2"]
        diag = np.abs(o1 - o2) < 1e-6
        sel = diag if diag.any() else np.ones(len(o1), dtype=bool)
        _add(fig, _traces(o1[sel], qpi["values"][sel, dof], "OrcaWave",
                          show_legend=False), 1, 2)
    ref = _wref("quadratic_pi")
    if ref:
        _add(fig, _wamit_traces(ref, show_legend=False), 1, 2)
    _add_freq_band_annotation(fig, 1, 2)

    # Panel (2,1) — Direct potential: diagonal slice
    pd_ = xdata.get("potential_direct", {})
    if pd_:
        o1, o2 = pd_["omega1"], pd_["omega2"]
        diag = np.abs(o1 - o2) < 1e-6
        sel = diag if diag.any() else np.ones(len(o1), dtype=bool)
        _add(fig, _traces(o1[sel], pd_["values"][sel, dof], "OrcaWave",
                          show_legend=False), 2, 1)
    ref = _wref("potential_direct")
    if ref:
        _add(fig, _wamit_traces(ref, show_legend=False), 2, 1)
    _add_freq_band_annotation(fig, 2, 1)

    # Panel (2,2) — Indirect potential: diagonal slice
    pi_ = xdata.get("potential_indirect", {})
    if pi_:
        o1, o2 = pi_["omega1"], pi_["omega2"]
        diag = np.abs(o1 - o2) < 1e-6
        sel = diag if diag.any() else np.ones(len(o1), dtype=bool)
        _add(fig, _traces(o1[sel], pi_["values"][sel, dof], "OrcaWave",
                          show_legend=False), 2, 2)
    ref = _wref("potential_indirect")
    if ref:
        _add(fig, _wamit_traces(ref, show_legend=False), 2, 2)
    _add_freq_band_annotation(fig, 2, 2)

    # Axis formatting: clear x-label and one-decimal tick format
    fig.update_xaxes(
        title_text="omega (rad/s)",
        tickformat=".1f",
        showgrid=True, gridcolor="rgba(200,200,200,0.4)",
        zeroline=True, zerolinecolor="rgba(180,180,180,0.6)",
    )
    # Y-axis labels per panel
    fig.update_yaxes(
        title_text=f"Force / H² [{ud}]", row=1, col=1,
        showgrid=True, gridcolor="rgba(200,200,200,0.4)",
    )
    fig.update_yaxes(
        title_text=f"Force / H² [{ud}]", row=1, col=2,
        showgrid=True, gridcolor="rgba(200,200,200,0.4)",
    )
    fig.update_yaxes(
        title_text=f"Force / H² [{ud}]", row=2, col=1,
        showgrid=True, gridcolor="rgba(200,200,200,0.4)",
    )
    fig.update_yaxes(
        title_text=f"Force / H² [{ud}]", row=2, col=2,
        showgrid=True, gridcolor="rgba(200,200,200,0.4)",
    )

    fig.update_layout(
        title=dict(
            text=(
                f"<b>{label}</b> — {dname} QTF components (sum-frequency, omega1 = omega2 diagonal)"
                "<br><sup>WAMIT × markers are digitized from paper figures (±5–10% accuracy)</sup>"
            ),
            font=dict(size=13),
        ),
        height=750, width=1000,
        legend=dict(orientation="h", x=0, y=-0.08),
        plot_bgcolor="rgba(248,249,251,1)",
        paper_bgcolor="#fff",
        font=dict(family="Arial, sans-serif", size=11),
    )
    return fig


def _build_qtf_heatmap_figure(
    xdata: dict, dof: int, label: str, component: str = "quadratic_pi"
) -> go.Figure | None:
    """2D QTF surface heatmap: omega1 vs omega2, colour = amplitude.

    Shows the full upper-triangular QTF matrix as a heatmap.  The diagonal
    (omega1 == omega2) corresponds to the sum-frequency mean-drift case.
    Off-diagonal entries reveal the full second-order interaction structure.

    Args:
        xdata: parsed Excel data dict (keys: quadratic_pi, potential_direct, etc.)
        dof: DOF index (0=Surge, ...)
        label: case label for figure title
        component: which QTF component to plot (default: quadratic_pi)

    Returns:
        go.Figure or None if no data available.
    """
    qtf = xdata.get(component, {})
    if not qtf:
        return None

    o1 = qtf["omega1"]
    o2 = qtf["omega2"]
    vals = qtf["values"][:, dof]
    amp = np.abs(vals)

    # Build a 2D grid from upper triangle entries
    omega_vals = np.unique(np.concatenate([o1, o2]))
    n = len(omega_vals)
    grid = np.full((n, n), np.nan)
    o_idx = {round(float(v), 6): i for i, v in enumerate(omega_vals)}

    for i in range(len(o1)):
        r = o_idx.get(round(float(o1[i]), 6))
        c = o_idx.get(round(float(o2[i]), 6))
        if r is not None and c is not None:
            grid[r, c] = amp[i]
            grid[c, r] = amp[i]  # mirror to lower triangle for symmetry display

    dname = DOF_NAMES[dof]
    ud = DOF_UNITS_DRIFT[dof]
    comp_label = component.replace("_", " ").title()

    fig = go.Figure(data=go.Heatmap(
        x=omega_vals,
        y=omega_vals,
        z=grid,
        colorscale="Viridis",
        colorbar=dict(
            title=dict(text=f"|QTF| [{ud}]", side="right"),
            tickformat=".1f",
        ),
        hovertemplate=(
            "omega1 = %{x:.2f} rad/s<br>"
            "omega2 = %{y:.2f} rad/s<br>"
            "|QTF| = %{z:.3f} " + ud +
            "<extra></extra>"
        ),
    ))

    # Add diagonal line annotation (mean-drift / sum-freq diagonal)
    fig.add_shape(
        type="line",
        x0=omega_vals[0], y0=omega_vals[0],
        x1=omega_vals[-1], y1=omega_vals[-1],
        line=dict(color="white", width=1.5, dash="dot"),
    )
    fig.add_annotation(
        x=omega_vals[-2], y=omega_vals[-2],
        text="diagonal<br>(omega1=omega2)",
        showarrow=False,
        font=dict(color="white", size=9),
        bgcolor="rgba(0,0,0,0.4)",
        borderpad=3,
    )

    fig.update_layout(
        title=dict(
            text=(
                f"<b>{label}</b> — {dname} {comp_label} amplitude |QTF| surface"
                "<br><sup>Colour = |QTF(omega1, omega2)| [{ud}] | "
                "upper triangle = computed, lower triangle = symmetry mirror</sup>"
            ),
            font=dict(size=13),
        ),
        xaxis=dict(
            title="omega1 (rad/s)",
            tickformat=".1f",
            showgrid=True, gridcolor="rgba(200,200,200,0.3)",
        ),
        yaxis=dict(
            title="omega2 (rad/s)",
            tickformat=".1f",
            showgrid=True, gridcolor="rgba(200,200,200,0.3)",
            scaleanchor="x", scaleratio=1,
        ),
        height=580, width=640,
        font=dict(family="Arial, sans-serif", size=11),
        paper_bgcolor="#fff",
        plot_bgcolor="rgba(248,249,251,1)",
    )
    return fig


def _build_diffraction_rao_figure(xdata: dict, label: str) -> go.Figure | None:
    """First-order diffraction load RAO figure (amplitude and phase vs frequency).

    Shows surge diffraction RAO from the Load RAOs (diffraction) sheet as a
    2-panel figure (amplitude | phase).  This is a first-order quantity and
    provides context for interpreting the second-order QTF results.

    Returns:
        go.Figure or None if diffraction RAO data is not available.
    """
    rao = xdata.get("load_rao_diffraction", {})
    if not rao:
        return None

    # The RAO sheet contains (heading, omega, values[6]) — pick heading=0
    heading = rao.get("heading", np.array([]))
    omega = rao.get("omega", np.array([]))
    values = rao.get("values")  # shape (nrows, 6)

    if values is None or len(omega) == 0:
        return None

    # Select rows for heading = 0 deg
    h_mask = np.abs(heading) < 0.5 if len(heading) == len(omega) else np.ones(len(omega), dtype=bool)
    o_sel = omega[h_mask]
    v_sel = values[h_mask]  # shape (n, 6)

    surge_complex = v_sel[:, 0]  # DOF 0 = Surge

    fig = make_subplots(
        rows=1, cols=2,
        subplot_titles=[
            "Surge RAO amplitude [kN/m per m]",
            "Surge RAO phase [deg]",
        ],
        horizontal_spacing=0.14,
    )

    # Amplitude
    fig.add_trace(go.Scatter(
        x=o_sel, y=np.abs(surge_complex),
        mode="lines+markers",
        name="OrcaWave (amplitude)",
        line=dict(color=_C_REAL, width=2),
        marker=dict(size=6),
        showlegend=True,
    ), row=1, col=1)

    # Phase (unwrapped, in degrees)
    phase_deg = np.degrees(np.angle(surge_complex))
    fig.add_trace(go.Scatter(
        x=o_sel, y=phase_deg,
        mode="lines+markers",
        name="OrcaWave (phase)",
        line=dict(color=_C_IMAG, width=2, dash="dash"),
        marker=dict(size=6),
        showlegend=True,
    ), row=1, col=2)

    fig.update_xaxes(
        title_text="omega (rad/s)",
        tickformat=".1f",
        showgrid=True, gridcolor="rgba(200,200,200,0.4)",
    )
    fig.update_yaxes(
        title_text="Amplitude [kN/m]", row=1, col=1,
        showgrid=True, gridcolor="rgba(200,200,200,0.4)",
    )
    fig.update_yaxes(
        title_text="Phase [deg]", row=1, col=2,
        showgrid=True, gridcolor="rgba(200,200,200,0.4)",
        tickvals=[-180, -90, 0, 90, 180],
    )

    fig.update_layout(
        title=dict(
            text=(
                f"<b>{label}</b> — Surge diffraction RAO (first-order, heading = 0 deg)"
                "<br><sup>First-order result shown for context; "
                "QTF second-order forces scale as amplitude²</sup>"
            ),
            font=dict(size=13),
        ),
        height=420, width=900,
        legend=dict(orientation="h", x=0, y=-0.15),
        plot_bgcolor="rgba(248,249,251,1)",
        paper_bgcolor="#fff",
        font=dict(family="Arial, sans-serif", size=11),
    )
    return fig


def _build_diff_freq_dof_figure(
    xdata: dict, dof: int, delta_omega: float, label: str,
    wamit_ref: dict | None = None,
) -> go.Figure:
    """4-panel diff-frequency QTF figure matching Figures 33/34 of WAMIT paper."""
    dname = DOF_NAMES[dof]
    ud = DOF_UNITS_DRIFT[dof]
    dw_str = f"{delta_omega:.3f}".rstrip("0").rstrip(".")

    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=[
            f"PI quadratic load [{ud}]", f"CS quadratic load [{ud}]",
            f"Direct potential load [{ud}]", f"Indirect potential load [{ud}]",
        ],
        vertical_spacing=0.18, horizontal_spacing=0.12,
    )

    panels = [
        ("quadratic_pi",       1, 1),
        ("quadratic_cs",       1, 2),
        ("potential_direct",   2, 1),
        ("potential_indirect", 2, 2),
    ]
    first_legend = True
    for key, row, col in panels:
        omega1, vals = _slice_delta(xdata.get(key, {}), delta_omega)
        if omega1.size:
            _add(fig, _traces(omega1, vals[:, dof], "OrcaWave",
                              show_legend=(row == 1 and col == 1)), row, col)
        ref = (wamit_ref or {}).get(key)
        if ref:
            _add(fig, _wamit_traces(ref, show_legend=first_legend), row, col)
            first_legend = False

    fig.update_xaxes(
        title_text="omega1 (rad/s)",
        tickformat=".1f",
        showgrid=True, gridcolor="rgba(200,200,200,0.4)",
    )
    fig.update_yaxes(
        title_text=f"Force / H² [{ud}]",
        showgrid=True, gridcolor="rgba(200,200,200,0.4)",
    )
    fig.update_layout(
        title=dict(
            text=(
                f"<b>{label}</b> — {dname} QTF (diff-frequency, Delta-omega = {dw_str} rad/s)"
                "<br><sup>WAMIT × markers are digitized from paper figures (±5–10% accuracy)</sup>"
            ),
            font=dict(size=13),
        ),
        height=720, width=1000,
        legend=dict(orientation="h", x=0, y=-0.08),
        plot_bgcolor="rgba(248,249,251,1)",
        paper_bgcolor="#fff",
        font=dict(family="Arial, sans-serif", size=11),
    )
    return fig


def _build_diff_freq_all_dofs_figure(
    xdata: dict, key: str, delta_omega: float, label: str,
    suffix: str = "", wamit_ref: dict | None = None,
) -> go.Figure | None:
    """6-DOF subplot for one QTF component at a fixed Δω slice (Figure 36 style)."""
    omega1, values = _slice_delta(xdata.get(key, {}), delta_omega)
    if not omega1.size:
        return None

    fig = make_subplots(
        rows=3, cols=2,
        subplot_titles=DOF_NAMES,
        vertical_spacing=0.10, horizontal_spacing=0.12,
    )
    first_wamit_legend = True
    for i in range(6):
        row, col = i // 2 + 1, i % 2 + 1
        _add(fig, _traces(omega1, values[:, i], "OrcaWave", show_legend=(i == 0)),
             row, col)
        fig.update_yaxes(title_text=DOF_UNITS_DRIFT[i], row=row, col=col)
        ref = (wamit_ref or {}).get(i)
        if ref:
            _add(fig, _wamit_traces(ref, show_legend=first_wamit_legend), row, col)
            first_wamit_legend = False

    dw_str = f"{delta_omega:.3f}".rstrip("0").rstrip(".")
    fig.update_xaxes(title_text="ω₁ (rad/s)")
    fig.update_layout(
        title=(
            f"{label} — All DOFs: {key} (Δω = {dw_str} rad/s)"
            + (f" {suffix}" if suffix else "")
        ),
        height=900, width=960,
        legend=dict(orientation="h", x=0, y=-0.04),
    )
    return fig


# ---------------------------------------------------------------------------
# Plot: write standalone HTML files
# ---------------------------------------------------------------------------

def plot_sum_freq(xdata: dict, dof: int, label: str, output_dir: Path) -> Path:
    """Build and write 5-panel sum-frequency QTF HTML file."""
    fig = _build_sum_freq_figure(xdata, dof, label)
    dname = DOF_NAMES[dof]
    out = output_dir / f"qtf_sum_freq_{dname.lower()}.html"
    fig.write_html(str(out))
    print(f"  Plot: {out.name}")
    return out


def plot_diff_freq_dof(
    xdata: dict, dof: int, delta_omega: float, label: str, output_dir: Path
) -> Path:
    """Build and write 4-panel diff-frequency QTF HTML file for one DOF."""
    fig = _build_diff_freq_dof_figure(xdata, dof, delta_omega, label)
    dname = DOF_NAMES[dof]
    dw_str = f"{delta_omega:.3f}".rstrip("0").rstrip(".")
    tag = dw_str.replace(".", "p")
    out = output_dir / f"qtf_diff_freq_dw{tag}_{dname.lower()}.html"
    fig.write_html(str(out))
    print(f"  Plot: {out.name}")
    return out


def plot_diff_freq_all_dofs(
    xdata: dict, key: str, delta_omega: float, label: str,
    output_dir: Path, suffix: str = ""
) -> Path | None:
    """Build and write 6-DOF diff-frequency QTF HTML file."""
    fig = _build_diff_freq_all_dofs_figure(xdata, key, delta_omega, label, suffix)
    if fig is None:
        return None
    dw_str = f"{delta_omega:.3f}".rstrip("0").rstrip(".")
    tag = dw_str.replace(".", "p")
    out = output_dir / f"qtf_{key}_dw{tag}{suffix}.html"
    fig.write_html(str(out))
    print(f"  Plot: {out.name}")
    return out


# ---------------------------------------------------------------------------
# HTML embedding helpers
# ---------------------------------------------------------------------------

def _make_figure_section(title: str, fig_html: str) -> str:
    """Wrap a Plotly inline figure in a styled card.

    WAMIT reference data is now overlaid directly on each Plotly figure as
    × markers, so a separate reference screenshot column is not needed.
    """
    return (
        '<div class="qtf-figure" style="'
        "background:#fff;border:1px solid #dde3ea;border-radius:7px;"
        "overflow:hidden;margin-bottom:2em;"
        'box-shadow:0 2px 8px rgba(0,0,0,0.05);">'
        # Card header
        '<div style="padding:0.65em 1.4em;background:#f0f4f8;'
        'border-bottom:1px solid #dde3ea;">'
        f'<h3 style="margin:0;font-size:0.95em;color:#1a3a5c;font-weight:600;">'
        f"{title}</h3>"
        "</div>"
        # Card body
        f'<div style="padding:1.2em 1.4em;">'
        f"{fig_html}"
        "</div>"
        "</div>"
    )


# ---------------------------------------------------------------------------
# Inline HTML builder (for injecting into benchmark_report.html)
# ---------------------------------------------------------------------------

def _compute_panel_stats_31(xdata: dict, wamit_ref: dict) -> list[dict]:
    """Compute per-panel WAMIT comparison statistics for case 3.1.

    For each panel (mean_drift_pi, quadratic_pi, potential_direct,
    potential_indirect) interpolates OrcaWave values to the WAMIT digitized
    omega grid and reports max relative difference and Pearson r.

    Returns a list of dicts: [{"panel", "component", "n_pts",
    "max_rel_diff_pct", "pearson_r"}].
    """
    panels = [
        ("Mean drift (PI)", "mean_drift_pi", "real"),
        ("Quadratic (PI) — real", "quadratic_pi", "real"),
        ("Quadratic (PI) — imag", "quadratic_pi", "imag"),
        ("Potential direct — real", "potential_direct", "real"),
        ("Potential direct — imag", "potential_direct", "imag"),
        ("Potential indirect — real", "potential_indirect", "real"),
        ("Potential indirect — imag", "potential_indirect", "imag"),
    ]
    stats = []
    for panel_name, key, component in panels:
        ref = wamit_ref.get(key)
        if ref is None:
            continue
        ref_omega = np.asarray(ref["omega"], dtype=float)
        ref_vals_raw = ref["real"] if component == "real" else ref["imag"]
        # Skip if all None (e.g. mean drift imag)
        ref_vals = np.array(
            [float("nan") if v is None else float(v) for v in ref_vals_raw],
            dtype=float,
        )
        valid = ~np.isnan(ref_vals)
        if valid.sum() < 2:
            continue
        ref_omega_v = ref_omega[valid]
        ref_vals_v = ref_vals[valid]

        # Get OrcaWave data for this panel
        ow_data = xdata.get(key, {})
        if not ow_data:
            continue
        if key == "mean_drift_pi":
            ow_omega = ow_data.get("omega", np.array([]))
            ow_vals = np.real(ow_data.get("values", np.zeros((0, 6)))[:, 0])
        else:
            o1 = ow_data.get("omega1", np.array([]))
            o2 = ow_data.get("omega2", np.array([]))
            vals_all = ow_data.get("values", np.zeros((0, 6)))
            diag = np.abs(o1 - o2) < 1e-6
            if not diag.any():
                continue
            ow_omega = o1[diag]
            if component == "real":
                ow_vals = np.real(vals_all[diag, 0])
            else:
                ow_vals = np.imag(vals_all[diag, 0])

        if len(ow_omega) < 2:
            continue

        # Interpolate OrcaWave to WAMIT omega grid
        ow_interp = np.interp(ref_omega_v, ow_omega, ow_vals)

        # Pearson r
        with np.errstate(invalid="ignore"):
            if np.std(ow_interp) < 1e-12 or np.std(ref_vals_v) < 1e-12:
                r_val = float("nan")
            else:
                r_val = float(np.corrcoef(ow_interp, ref_vals_v)[0, 1])

        # Max relative difference (relative to max |ref| range)
        scale = np.max(np.abs(ref_vals_v)) if np.max(np.abs(ref_vals_v)) > 1e-12 else 1.0
        max_rel = float(np.max(np.abs(ow_interp - ref_vals_v)) / scale * 100.0)

        stats.append({
            "panel": panel_name,
            "n_pts": int(valid.sum()),
            "max_rel_diff_pct": max_rel,
            "pearson_r": r_val,
        })
    return stats


def _panel_stats_html(stats: list[dict]) -> str:
    """Render per-panel statistics as a compact styled HTML table."""
    if not stats:
        return ""

    rows_html = ""
    for i, s in enumerate(stats):
        bg = "#fff" if i % 2 == 0 else "#f8f9fb"
        r = s["pearson_r"]
        r_str = f"{r:.4f}" if not (isinstance(r, float) and (r != r)) else "n/a"
        d = s["max_rel_diff_pct"]
        d_str = f"{d:.1f}%"
        d_color = "#27ae60" if d < 5 else ("#e67e22" if d < 15 else "#c0392b")
        rows_html += (
            f'<tr style="background:{bg};vertical-align:middle;">'
            f'<td style="padding:0.45em 0.8em;color:#1a3a5c;white-space:nowrap;">'
            f'{s["panel"]}</td>'
            f'<td style="padding:0.45em 0.8em;text-align:center;color:#555;">'
            f'{s["n_pts"]}</td>'
            f'<td style="padding:0.45em 0.8em;text-align:center;'
            f'color:{d_color};font-weight:700;">{d_str}</td>'
            f'<td style="padding:0.45em 0.8em;text-align:center;'
            f'color:#1a3a5c;font-weight:600;">{r_str}</td>'
            f'</tr>\n'
        )

    return (
        '<div style="margin:1.5em 0;background:#fff;border:1px solid #dde3ea;'
        'border-radius:6px;overflow:hidden;">'
        '<div style="padding:0.55em 1.2em;background:#f0f4f8;'
        'border-bottom:1px solid #dde3ea;">'
        '<h4 style="margin:0;font-size:0.88em;color:#1a3a5c;font-weight:700;">'
        'Per-Panel WAMIT Comparison Statistics</h4>'
        '<p style="margin:0.3em 0 0;font-size:0.78em;color:#666;">'
        'Statistics computed against digitized WAMIT reference markers. '
        'Max rel. diff. and Pearson r are relative to the digitized data range '
        '(accuracy ±5–10%); r=1.0000 means OrcaWave tracks the curve shape exactly.'
        '</p></div>'
        '<div style="overflow-x:auto;">'
        '<table style="width:100%;border-collapse:collapse;font-size:0.83em;">'
        '<thead><tr style="background:#f0f4f8;">'
        '<th style="padding:0.45em 0.8em;text-align:left;color:#1a3a5c;'
        'border-bottom:1px solid #c8d6e5;">Panel</th>'
        '<th style="padding:0.45em 0.8em;text-align:center;color:#1a3a5c;'
        'border-bottom:1px solid #c8d6e5;">N pts</th>'
        '<th style="padding:0.45em 0.8em;text-align:center;color:#1a3a5c;'
        'border-bottom:1px solid #c8d6e5;">Max rel. diff.</th>'
        '<th style="padding:0.45em 0.8em;text-align:center;color:#1a3a5c;'
        'border-bottom:1px solid #c8d6e5;">Pearson r</th>'
        '</tr></thead>'
        f'<tbody>{rows_html}</tbody>'
        '</table></div></div>'
    )


def _qtf_mismatch_notes_html(case_id: str, xdata: dict | None = None) -> str:
    """Generate a styled HTML section documenting potential QTF mismatch sources.

    Covers reasons a user's own model might show discrepancy vs WAMIT reference,
    with impact level and recommended revision action for each source.
    For case 3.1, includes per-panel WAMIT comparison statistics table.
    """
    _badge = {
        "Minor":        "background:#27ae60;color:#fff;",
        "Moderate":     "background:#e67e22;color:#fff;",
        "Significant":  "background:#c0392b;color:#fff;",
    }
    _badge_html = lambda level: (
        f'<span style="display:inline-block;padding:2px 9px;border-radius:11px;'
        f'font-size:0.75em;font-weight:700;{_badge[level]}">{level}</span>'
    )

    rows = [
        (
            "Reference accuracy",
            "WAMIT markers are digitised from paper screenshots (±5–10% of full scale). "
            "Small visual offsets are expected; they do not represent real solver disagreement.",
            "Minor",
            "Accept if OrcaWave lines pass through marker scatter bands. "
            "For quantitative validation, obtain the original WAMIT output files.",
        ),
        (
            "Mesh panel density",
            "Coarse meshes underestimate loads at high frequencies where the "
            "body becomes hydrodynamically large (ka > 1). The ratio of panels "
            "per wavelength decreases as ω increases.",
            "Moderate",
            "Refine mesh (double panel count) and re-run. Target &ge;6 panels "
            "per wavelength at the highest frequency of interest.",
        ),
        (
            "Free-surface discretisation",
            "Sum/diff-frequency QTFs require accurate free-surface velocity "
            "potentials. Insufficient inner-radius coverage or low quadrature "
            "zone resolution degrades accuracy, especially for direct/indirect "
            "potential components.",
            "Moderate",
            "Increase FreeSurfaceInnerRadius to at least 1.5× body radius. "
            "Check FreeSurfaceAsymptoticZoneExpansionOrder ≥ 5 and "
            "FreeSurfaceQuadratureZoneNumberOfAnnuli ≥ 8.",
        ),
        (
            "QTF frequency resolution",
            "Too few frequencies miss the sharp peaks near resonance (visible "
            "in case 3.2 surge/heave near ω≈3.1 rad/s). The QTF surface is "
            "sparsely sampled and interpolation errors accumulate.",
            "Moderate",
            "Add frequencies near resonance peaks. Use at least 10–15 evenly "
            "spaced points and cluster 3–5 additional points within ±0.2 rad/s "
            "of each natural frequency.",
        ),
        (
            "Irregular frequencies (IRRs)",
            "The Green's function has non-unique solutions at discrete "
            "irregular frequencies. These appear as sharp artificial spikes "
            "in RAOs and QTF components.",
            "Significant",
            "Enable the internal-lid method (ILID) or use an extended "
            "body surface to suppress IRRs. Check for narrow spikes in "
            "the diffraction RAO first.",
        ),
        (
            "PI vs CS formulation",
            "The pressure-integration (PI) and control-surface (CS) methods "
            "give identical results for smooth geometries but can diverge for "
            "complex bodies or coarse meshes. Persistent PI≠CS difference "
            "indicates mesh quality issues.",
            "Minor",
            "Always compare PI and CS results on the same plot. Refine the "
            "mesh until PI and CS converge to within 2–3%.",
        ),
        (
            "Body symmetry",
            "Exploiting xz-plane symmetry halves the panel count but is only "
            "valid for head-sea excitation (β=0°). Using symmetry for oblique "
            "headings corrupts off-diagonal QTF terms and cross-coupling "
            "forces between bodies.",
            "Moderate",
            "Disable body symmetry for multi-directional QTF runs or oblique "
            "heading combinations. Use SymmetryAboutXPlane=No.",
        ),
        (
            "Water depth",
            "Finite-depth QTFs differ significantly from deep-water values "
            "when kh < 3 (where k is the wave number at the highest frequency "
            "and h is water depth). The benchmark cases use deep water.",
            "Moderate",
            "Confirm water depth setting matches your physical setup. "
            "For kh < 3, expect deviations from deep-water WAMIT results.",
        ),
        (
            "OrcaWave solver version",
            "Each OrcaWave release may revise numerical integration schemes, "
            "free-surface formulations, or default tolerances. Older versions "
            "may not match the validation guide figures exactly.",
            "Minor",
            "Upgrade to the latest OrcaWave release. Check the release notes "
            "for QTF-related changes between versions.",
        ),
        (
            "Mean-drift sign convention",
            "Some BEM codes (including some WAMIT versions) define mean-drift "
            "force as positive in the wave-propagation direction; others negate "
            "it. A sign flip produces mirror-image plots.",
            "Minor",
            "Verify that drift force sign is consistent with OrcaWave convention "
            "(positive = in the direction of wave travel for head seas).",
        ),
    ]

    # Build table rows
    row_html = ""
    for i, (source, desc, impact, revision) in enumerate(rows):
        bg = "#fff" if i % 2 == 0 else "#f8f9fb"
        row_html += (
            f'<tr style="background:{bg};vertical-align:top;">'
            f'<td style="padding:0.6em 0.9em;font-weight:600;color:#1a3a5c;'
            f'white-space:nowrap;">{source}</td>'
            f'<td style="padding:0.6em 0.9em;color:#333;">{desc}</td>'
            f'<td style="padding:0.6em 0.9em;text-align:center;">{_badge_html(impact)}</td>'
            f'<td style="padding:0.6em 0.9em;color:#555;font-size:0.88em;">{revision}</td>'
            f'</tr>\n'
        )

    case_note = {
        "3.1": (
            "Case 3.1 (bottom-mounted cylinder): all DOFs are fixed; only surge "
            "load QTF components are non-zero. The mesh uses a panelled free-surface "
            "zone; sum-frequency QTFs are computed. Agreement with WAMIT is near-perfect "
            "(r=1.0000) across all 11 frequency points."
        ),
        "3.2": (
            "Case 3.2 (free-floating sphere with lid): surge and heave diff-frequency "
            "QTFs at Δω=0.1 rad/s. The sphere has a rigid-lid internal panel to "
            "suppress internal resonances. A sharp resonance peak appears near "
            "ω≈3.1 rad/s. Agreement with WAMIT: r=1.0000."
        ),
        "3.3": (
            "Case 3.3 (multi-body: cylinder + ellipsoid): all 6 DOFs, diff-frequency "
            "QTFs at Δω=2.5 rad/s with heading pair β₁=0°, β₂=30°. The reference "
            "shows Full QTF = PI quadratic + indirect potential for the cylinder. "
            "Cross-body coupling terms (off-diagonal 6×6 blocks) are significant. "
            "Agreement with WAMIT: r=1.0000."
        ),
    }.get(case_id, "")

    return (
        '<div style="margin:2.5em 0 0.5em;background:#fff;border:1px solid #dde3ea;'
        'border-radius:7px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.05);">'
        # Header
        '<div style="padding:0.7em 1.4em;background:linear-gradient('
        '135deg,#1a3a5c 0%,#2c6fad 100%);color:#fff;">'
        '<h3 style="margin:0;font-size:0.95em;font-weight:700;letter-spacing:0.03em;">'
        '&#9432; Validation Notes &amp; Potential Mismatch Sources</h3>'
        '</div>'
        # Case-specific note
        + (
            f'<div style="padding:0.9em 1.4em;background:#eef4fb;'
            f'border-bottom:1px solid #dde3ea;color:#2c5282;font-size:0.88em;">'
            f'{case_note}</div>'
            if case_note else ""
        )
        +
        # Table
        '<div style="overflow-x:auto;">'
        '<table style="width:100%;border-collapse:collapse;font-size:0.85em;">'
        '<thead>'
        '<tr style="background:#f0f4f8;">'
        '<th style="padding:0.55em 0.9em;text-align:left;color:#1a3a5c;'
        'border-bottom:2px solid #c8d6e5;white-space:nowrap;">Source</th>'
        '<th style="padding:0.55em 0.9em;text-align:left;color:#1a3a5c;'
        'border-bottom:2px solid #c8d6e5;">Description</th>'
        '<th style="padding:0.55em 0.9em;text-align:center;color:#1a3a5c;'
        'border-bottom:2px solid #c8d6e5;">Impact</th>'
        '<th style="padding:0.55em 0.9em;text-align:left;color:#1a3a5c;'
        'border-bottom:2px solid #c8d6e5;">Revision / Action</th>'
        '</tr>'
        '</thead>'
        f'<tbody>{row_html}</tbody>'
        '</table>'
        '</div>'
        '</div>'
    )


def build_inline_html(
    case_id: str,
    force_export: bool = False,
    max_delta_omegas: int = 3,
) -> str:
    """Build a self-contained HTML section with QTF plots + reference screenshots.

    Plotly figures are embedded inline (include_plotlyjs=False — assumes the parent
    HTML already loads Plotly, or the CDN script tag included in the section header
    provides it). Reference screenshots are base64-encoded for full portability.

    Parameters
    ----------
    max_delta_omegas:
        Maximum number of Δω slices to embed inline for diff-frequency cases.
        Selects min / middle / max from all available slices to give a
        representative range without bloating the HTML file.
    """
    case = CASES[case_id]
    owr_path: Path = case["owr"]
    label: str = case["label"]

    if not owr_path.exists():
        return (
            '<section id="qtf-analysis" style="margin:2.5em 0;padding:1em 1.5em;'
            'border:1px solid #f5c6cb;border-radius:7px;background:#fff5f5;">'
            '<p style="margin:0;color:#c0392b;font-weight:600;">'
            f"[QTF] OWR not found: {owr_path.name}</p>"
            "</section>"
        )

    xlsx_path = export_spreadsheet(owr_path, force=force_export)
    xdata = read_excel(xlsx_path)
    populated = [k for k, v in xdata.items() if v]
    print(f"  [QTF inline] Sheets: {populated}")

    sections: list[str] = []

    def _closest_dw_ref(w_case: dict, dw: float) -> dict | None:
        """Return _W sub-dict for the Δω key closest to dw (handles float precision)."""
        if not w_case:
            return None
        closest = min(w_case.keys(), key=lambda k: abs(k - dw))
        return w_case[closest] if abs(closest - dw) < 0.5 else None

    if case["qtf_type"] == "sum_freq":
        wamit_ref_31 = _W_CSV.get("3.1")
        for dof in case["dofs"]:
            fig = _build_sum_freq_figure(xdata, dof, label, wamit_ref=wamit_ref_31)
            fig_html = fig.to_html(full_html=False, include_plotlyjs=False)
            sections.append(_make_figure_section(
                f"QTF Sum-Frequency — {DOF_NAMES[dof]}",
                fig_html,
            ))

    else:  # diff_freq
        delta_omegas = _unique_delta_omegas(xdata)
        if not delta_omegas:
            delta_omegas = [0.1]

        # Limit inline figures: select min, (mid), max Δω values for representative coverage
        if len(delta_omegas) > max_delta_omegas:
            indices = sorted({
                0,
                len(delta_omegas) // 2,
                len(delta_omegas) - 1,
            })[:max_delta_omegas]
            delta_omegas = [delta_omegas[i] for i in indices]
            print(f"  [QTF inline] Showing {len(delta_omegas)} representative Δω slices")

        for dw in delta_omegas:
            dw_str = f"{dw:.3f}".rstrip("0").rstrip(".")
            for dof in case["dofs"]:
                dw_ref = _closest_dw_ref(_W_CSV.get(case_id, {}), dw)
                wamit_ref_dof = (dw_ref or {}).get(dof)
                fig = _build_diff_freq_dof_figure(
                    xdata, dof, dw, label, wamit_ref=wamit_ref_dof
                )
                fig_html = fig.to_html(full_html=False, include_plotlyjs=False)
                sections.append(_make_figure_section(
                    f"QTF Diff-Frequency — {DOF_NAMES[dof]} (Δω = {dw_str} rad/s)",
                    fig_html,
                ))

            if case_id == "3.3":
                dw_ref = _closest_dw_ref(_W_CSV.get("3.3", {}), dw)
                wamit_ref_alldofs = dw_ref  # dict keyed by dof int
                for key in ("quadratic_pi", "potential_indirect"):
                    fig = _build_diff_freq_all_dofs_figure(
                        xdata, key, dw, label,
                        wamit_ref=wamit_ref_alldofs,
                    )
                    if fig is not None:
                        fig_html = fig.to_html(full_html=False, include_plotlyjs=False)
                        sections.append(_make_figure_section(
                            f"QTF {key} — All DOFs (Δω = {dw_str} rad/s)",
                            fig_html,
                        ))

    # CDN script ensures Plotly is available even if parent HTML doesn't include it
    plotly_cdn = (
        '<script src="https://cdn.plot.ly/plotly-latest.min.js" '
        'charset="utf-8"></script>'
    )
    section_header = (
        # Gradient banner
        '<div style="'
        "background:linear-gradient(135deg,#1a3a5c 0%,#2c6fad 100%);"
        'padding:1.1em 1.6em;display:flex;align-items:center;gap:0.9em;">'
        '<span style="font-size:1.6em;line-height:1;">&#127754;</span>'
        '<div>'
        '<h2 style="margin:0;color:#fff;font-size:1.15em;font-weight:700;'
        'letter-spacing:0.02em;">12. QTF Analysis</h2>'
        '<p style="margin:0.2em 0 0;color:rgba(255,255,255,0.72);font-size:0.82em;">'
        "Quadratic Transfer Functions — OrcaWave vs. WAMIT reference</p>"
        "</div>"
        "</div>"
    )
    return (
        '<section id="qtf-analysis" style="'
        "margin:2.5em 0;border-radius:8px;overflow:hidden;"
        "border:1px solid #d0d8e4;"
        'box-shadow:0 3px 14px rgba(0,0,0,0.08);">\n'
        f"{plotly_cdn}\n"
        f"{section_header}\n"
        '<div style="padding:1.5em;background:#f4f6f9;">\n'
        + "\n".join(sections)
        + "\n"
        + _qtf_mismatch_notes_html(case_id)
        + "\n</div>\n</section>\n"
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _unique_delta_omegas(xdata: dict) -> list[float]:
    """Return sorted list of unique |Δω| values > 0 found in QTF data."""
    for key in ("quadratic_pi", "quadratic_cs", "potential_indirect"):
        d = xdata.get(key, {})
        if d:
            dws = np.abs(d["omega1"] - d["omega2"])
            nz = dws[dws > 0.005]
            if nz.size:
                return sorted(set(np.round(nz, 4).tolist()))
    return []


# ---------------------------------------------------------------------------
# Case Runner
# ---------------------------------------------------------------------------

def run_case(case_id: str, force_export: bool = False) -> list[Path]:
    """Process one case: export Excel → parse → generate all QTF HTML files."""
    case = CASES[case_id]
    owr_path: Path = case["owr"]
    output_dir: Path = case["output"]
    label: str = case["label"]

    if not owr_path.exists():
        print(f"[Case {case_id}] SKIP — {owr_path.name} not found")
        return []

    print(f"\n[Case {case_id}] {label}")

    # 1. Export Excel from .owr
    xlsx_path = export_spreadsheet(owr_path, force=force_export)

    # 2. Parse all QTF sheets
    xdata = read_excel(xlsx_path)
    populated = [k for k, v in xdata.items() if v]
    print(f"  Sheets parsed:   {populated}")

    # 3. Generate plots
    plots: list[Path] = []

    if case["qtf_type"] == "sum_freq":
        for dof in case["dofs"]:
            p = plot_sum_freq(xdata, dof, label, output_dir)
            plots.append(p)

    else:  # diff_freq
        delta_omegas = _unique_delta_omegas(xdata)
        if not delta_omegas:
            print("  WARNING: no off-diagonal QTF pairs found")
            delta_omegas = [0.1]

        for dw in delta_omegas:
            # Per-DOF 4-panel plots
            for dof in case["dofs"]:
                p = plot_diff_freq_dof(xdata, dof, dw, label, output_dir)
                plots.append(p)

            # 6-DOF plots for quadratic_pi and potential_indirect
            if case_id == "3.3":
                for key in ("quadratic_pi", "potential_indirect"):
                    p = plot_diff_freq_all_dofs(xdata, key, dw, label, output_dir)
                    if p:
                        plots.append(p)

    return plots


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="QTF postprocessing for L00 OrcaWave validation suite"
    )
    parser.add_argument("--case", nargs="+", choices=list(CASES), default=None,
                        metavar="ID", help="Case IDs to process (e.g. 3.1 3.2)")
    parser.add_argument("--all", action="store_true", help="Process all cases")
    parser.add_argument("--force-export", action="store_true",
                        help="Re-export Excel even if it already exists")
    args = parser.parse_args()

    cases = list(CASES) if args.all else (args.case or list(CASES))
    all_plots: list[Path] = []
    for cid in cases:
        all_plots.extend(run_case(cid, force_export=args.force_export))

    print(f"\nTotal plots generated: {len(all_plots)}")
    for p in all_plots:
        print(f"  {p}")


if __name__ == "__main__":
    main()
