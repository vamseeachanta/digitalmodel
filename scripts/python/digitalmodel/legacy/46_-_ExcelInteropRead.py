import OrcFxAPI
import openpyxl

book = openpyxl.load_workbook("scratch/LoadCaseData.xlsx")
sheet = book["Data"]
model = OrcFxAPI.Model()

row = 2
while sheet.cell(row=row, column=1).value:
    H = sheet.cell(row=row, column=1).value
    T = sheet.cell(row=row, column=2).value
    model.environment.WaveHeight = H
    model.environment.WavePeriod = T

    id = row - 1
    model.SaveData(f"scratch/case{id:03d},H={H:.1f},T={T:.1f}.dat")

    row += 1
