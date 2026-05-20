"""Build interactive OCIMF coefficient explorer HTML.

Reads /mnt/ace/acma-codes/OCIMF/OCIMF Coef.xlsx and produces a single-file
interactive HTML with Plotly charts + usage descriptions + parametric
comparisons. Output: /mnt/local-analysis/digitalmodel/docs/domains/charts/
phase2/ocimf/ocimf_coefficient_explorer.html
"""
from __future__ import annotations
import html
from pathlib import Path
import openpyxl
import pandas as pd
import plotly.graph_objects as go
import plotly.io as pio

XLSX = Path("/mnt/ace/acma-codes/OCIMF/OCIMF Coef.xlsx")
OUT = Path("/mnt/local-analysis/digitalmodel/docs/domains/charts/phase2/ocimf/ocimf_coefficient_explorer.html")


def _read_block(ws, col_start, col_end, header_row=2, data_start=3, data_end=None):
    """Read a rectangular block as list of dicts using R2 headers."""
    if data_end is None:
        data_end = ws.max_row
    headers = [ws.cell(header_row, c).value for c in range(col_start, col_end + 1)]
    rows = []
    for r in range(data_start, data_end + 1):
        rec = {headers[i]: ws.cell(r, col_start + i).value for i in range(len(headers))}
        # skip rows where the first (angle/%) column is None or non-numeric
        first = rec.get(headers[0])
        if first is None:
            continue
        if not isinstance(first, (int, float)):
            continue
        rows.append(rec)
    return rows


def parse_5a_9a(ws):
    """5 figure-blocks of 4 cols + spacer. Each block has Angle/Cylind/Angle/Conv."""
    blocks = [
        ("A5", 1.1, 1, 4),
        ("A6", 1.2, 6, 9),
        ("A7", 1.5, 11, 14),
        ("A8", 3.0, 16, 19),
        ("A9", "4.4+", 21, 24),
    ]
    long = []
    for fig, wdt, c_lo, c_hi in blocks:
        recs = _read_block(ws, c_lo, c_hi)
        for r in recs:
            angle = r.get("Angle")
            if angle is None:
                continue
            for shape_label, key in [("Cylindrical", "Cylind."), ("Conventional", "Conv.")]:
                val = r.get(key)
                if val is None:
                    continue
                long.append(dict(
                    figure=fig, coefficient="Cxc", vessel_class="Tanker",
                    loading_state="Loaded", hull_shape=shape_label,
                    wdt_ratio=str(wdt), x_axis="heading_deg", x=float(angle), y=float(val),
                ))
    return long


def parse_10a_14a(ws):
    long = []
    # A10 Cyc Loaded — cols 1-7: Angle | 1.05 | 1.1 | 1.2 | 1.5 | 3 | >6.00
    for r in _read_block(ws, 1, 7):
        angle = r.get("Angle")
        for wdt_key in [1.05, 1.1, 1.2, 1.5, 3, "       >6.00 "]:
            val = r.get(wdt_key)
            if val is None:
                continue
            wdt_label = ">6" if isinstance(wdt_key, str) else str(wdt_key)
            long.append(dict(
                figure="A10", coefficient="Cyc", vessel_class="Tanker",
                loading_state="Loaded", hull_shape="Any",
                wdt_ratio=wdt_label, x_axis="heading_deg", x=float(angle), y=float(val),
            ))
    # A11 Cxyc Loaded — cols 9-15
    for r in _read_block(ws, 9, 15):
        angle = r.get("Angle")
        for wdt_key in [1.05, 1.1, 1.2, 1.5, 3, "       >6.00 "]:
            val = r.get(wdt_key)
            if val is None:
                continue
            wdt_label = ">6" if isinstance(wdt_key, str) else str(wdt_key)
            long.append(dict(
                figure="A11", coefficient="Cxyc", vessel_class="Tanker",
                loading_state="Loaded", hull_shape="Any",
                wdt_ratio=wdt_label, x_axis="heading_deg", x=float(angle), y=float(val),
            ))
    # A12 Cxc Ballast — cols 17-21: Angle | max | mean | min | 4.4 Cylindrical
    for r in _read_block(ws, 17, 21):
        angle = r.get("Angle")
        for stat_key in ["max", "mean", "min", "4.4 Cylindrical"]:
            val = r.get(stat_key)
            if val is None:
                continue
            long.append(dict(
                figure="A12", coefficient="Cxc", vessel_class="Tanker",
                loading_state="Ballast 40%T", hull_shape=stat_key,
                wdt_ratio="envelope" if stat_key != "4.4 Cylindrical" else "4.4",
                x_axis="heading_deg", x=float(angle), y=float(val),
            ))
    # A13 Cyc Ballast — cols 23-28: Angle | 1.1 Conv | 1.2 Conv | 1.1 Cyl | 1.5 Cyl | 4.4 Cyl
    for r in _read_block(ws, 23, 28):
        angle = r.get("Angle")
        for combo_key in ["1.1 Conv", "1.2 Conv", "1.1 Cyl", "1.5 Cyl", "4.4 Cyl"]:
            val = r.get(combo_key)
            if val is None:
                continue
            wdt, hull = combo_key.split(" ")
            hull_full = "Cylindrical" if hull.startswith("Cyl") else "Conventional"
            long.append(dict(
                figure="A13", coefficient="Cyc", vessel_class="Tanker",
                loading_state="Ballast 40%T", hull_shape=hull_full,
                wdt_ratio=wdt, x_axis="heading_deg", x=float(angle), y=float(val),
            ))
    # A14 Cxyc Ballast — cols 30-35: same schema
    for r in _read_block(ws, 30, 35):
        angle = r.get("Angle")
        for combo_key in ["1.1 Conv", "1.2 Conv", "1.1 Cyl", "1.5 Cyl", "4.4 Cyl"]:
            val = r.get(combo_key)
            if val is None:
                continue
            wdt, hull = combo_key.split(" ")
            hull_full = "Cylindrical" if hull.startswith("Cyl") else "Conventional"
            long.append(dict(
                figure="A14", coefficient="Cxyc", vessel_class="Tanker",
                loading_state="Ballast 40%T", hull_shape=hull_full,
                wdt_ratio=wdt, x_axis="heading_deg", x=float(angle), y=float(val),
            ))
    return long


def parse_16a_19a(ws):
    long = []
    # A16 K factor — cols 1-6: % | 1.05 | 1.1 | 1.2 | 1.5 | 3
    for r in _read_block(ws, 1, 6):
        pct = r.get("%")
        for wdt_key in [1.05, 1.1, 1.2, 1.5, 3]:
            val = r.get(wdt_key)
            if val is None:
                continue
            long.append(dict(
                figure="A16", coefficient="K", vessel_class="Any",
                loading_state="Any", hull_shape="Any",
                wdt_ratio=str(wdt_key), x_axis="percent_of_full_load",
                x=float(pct), y=float(val),
            ))
    # A17 Cxw Gas — cols 8-10: Angle | Prismatic | Spherical
    for r in _read_block(ws, 8, 10):
        angle = r.get("Angle")
        for tank_key in ["Prismatic", "Spherical"]:
            val = r.get(tank_key)
            if val is None:
                continue
            long.append(dict(
                figure="A17", coefficient="Cxw", vessel_class="Gas Carrier",
                loading_state="N/A", hull_shape=f"{tank_key} Tank",
                wdt_ratio="N/A", x_axis="heading_deg", x=float(angle), y=float(val),
            ))
    # A18 Cyw Gas — cols 12-14
    for r in _read_block(ws, 12, 14):
        angle = r.get("Angle")
        for tank_key in ["Prismatic", "Spherical"]:
            val = r.get(tank_key)
            if val is None:
                continue
            long.append(dict(
                figure="A18", coefficient="Cyw", vessel_class="Gas Carrier",
                loading_state="N/A", hull_shape=f"{tank_key} Tank",
                wdt_ratio="N/A", x_axis="heading_deg", x=float(angle), y=float(val),
            ))
    # A19 Cxyw Gas — cols 16-18
    for r in _read_block(ws, 16, 18):
        angle = r.get("Angle")
        for tank_key in ["Prismatic", "Spherical"]:
            val = r.get(tank_key)
            if val is None:
                continue
            long.append(dict(
                figure="A19", coefficient="Cxyw", vessel_class="Gas Carrier",
                loading_state="N/A", hull_shape=f"{tank_key} Tank",
                wdt_ratio="N/A", x_axis="heading_deg", x=float(angle), y=float(val),
            ))
    return long


def load_all():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=False)
    rows = []
    rows += parse_5a_9a(wb["Data 5a-9a"])
    rows += parse_10a_14a(wb["Data 10a-14a"])
    rows += parse_16a_19a(wb["Data 16a-19a"])
    df = pd.DataFrame(rows)
    return df


# ---------- Figure builders ----------

FIG_META = {
    "A5_A9": {
        "title": "Figures A5–A9 — Longitudinal Current Drag Coefficient (Cxc), Loaded Tanker",
        "subtitle": "By WD/T ratio (1.1, 1.2, 1.5, 3.0, >4.4) × hull bow shape (Cylindrical / Conventional)",
        "x": "Heading angle (deg)", "y": "Cxc",
        "filter": lambda df: df[df.figure.isin(["A5", "A6", "A7", "A8", "A9"])],
    },
    "A10": {"title": "Figure A10 — Lateral Current Drag Coefficient (Cyc), Loaded Tanker",
            "subtitle": "By WD/T ratio (1.05, 1.1, 1.2, 1.5, 3, >6)",
            "x": "Heading angle (deg)", "y": "Cyc",
            "filter": lambda df: df[df.figure == "A10"]},
    "A11": {"title": "Figure A11 — Current Yaw Moment Coefficient (Cxyc), Loaded Tanker",
            "subtitle": "By WD/T ratio", "x": "Heading angle (deg)", "y": "Cxyc",
            "filter": lambda df: df[df.figure == "A11"]},
    "A12": {"title": "Figure A12 — Longitudinal Current Drag (Cxc), Ballasted Tanker (40% T)",
            "subtitle": "Statistical envelope (max / mean / min) + reference 4.4 Cylindrical curve",
            "x": "Heading angle (deg)", "y": "Cxc",
            "filter": lambda df: df[df.figure == "A12"]},
    "A13": {"title": "Figure A13 — Lateral Current Drag (Cyc), Ballasted Tanker (40% T)",
            "subtitle": "By WD/T × hull bow shape (Conventional / Cylindrical)",
            "x": "Heading angle (deg)", "y": "Cyc",
            "filter": lambda df: df[df.figure == "A13"]},
    "A14": {"title": "Figure A14 — Current Yaw Moment (Cxyc), Ballasted Tanker (40% T, Midships)",
            "subtitle": "By WD/T × hull bow shape",
            "x": "Heading angle (deg)", "y": "Cxyc",
            "filter": lambda df: df[df.figure == "A14"]},
    "A16": {"title": "Figure A16 — Current Velocity Correction Factor (K)",
            "subtitle": "K vs. percent-of-load by WD/T ratio (apply to free-stream current velocity)",
            "x": "Percent of full load (%)", "y": "K",
            "filter": lambda df: df[df.figure == "A16"]},
    "A17": {"title": "Figure A17 — Longitudinal Wind Drag Coefficient (Cxw), Gas Carrier",
            "subtitle": "By tank type (Prismatic / Spherical/Moss)",
            "x": "Heading angle (deg)", "y": "Cxw",
            "filter": lambda df: df[df.figure == "A17"]},
    "A18": {"title": "Figure A18 — Lateral Wind Drag Coefficient (Cyw), Gas Carrier",
            "subtitle": "By tank type", "x": "Heading angle (deg)", "y": "Cyw",
            "filter": lambda df: df[df.figure == "A18"]},
    "A19": {"title": "Figure A19 — Wind Yaw Moment Coefficient (Cxyw), Gas Carrier",
            "subtitle": "By tank type", "x": "Heading angle (deg)", "y": "Cxyw",
            "filter": lambda df: df[df.figure == "A19"]},
}


def color_palette(n):
    base = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b", "#e377c2", "#17becf"]
    return [base[i % len(base)] for i in range(n)]


def make_lineplot(df_block, group_cols, x_col, y_col, title, subtitle, x_label, y_label):
    fig = go.Figure()
    df_block = df_block.sort_values(x_col)
    groups = df_block[group_cols].drop_duplicates().to_records(index=False)
    colors = color_palette(len(groups))
    for i, g in enumerate(groups):
        mask = pd.Series([True] * len(df_block), index=df_block.index)
        for col, val in zip(group_cols, g):
            mask &= df_block[col] == val
        sub = df_block[mask].sort_values(x_col)
        if sub.empty:
            continue
        label = " | ".join(f"{c}={v}" for c, v in zip(group_cols, g))
        fig.add_trace(go.Scatter(
            x=sub[x_col], y=sub[y_col], mode="lines+markers",
            name=label, line=dict(color=colors[i], width=2),
            marker=dict(size=5), hovertemplate=f"{x_label}: %{{x}}<br>{y_label}: %{{y:.4f}}<extra>{label}</extra>",
        ))
    fig.update_layout(
        title=dict(text=f"<b>{title}</b><br><span style='font-size:13px;color:#666'>{subtitle}</span>",
                   x=0.02, xanchor="left"),
        xaxis_title=x_label, yaxis_title=y_label,
        legend=dict(orientation="v", x=1.02, y=1, font=dict(size=10)),
        margin=dict(l=70, r=200, t=80, b=60),
        height=480, hovermode="closest",
        plot_bgcolor="#fafafa", paper_bgcolor="white",
        xaxis=dict(gridcolor="#e0e0e0", zerolinecolor="#888"),
        yaxis=dict(gridcolor="#e0e0e0", zerolinecolor="#888"),
    )
    return fig


def fig_a5_a9(df):
    block = df[df.figure.isin(["A5", "A6", "A7", "A8", "A9"])].copy()
    block["wdt_label"] = block["wdt_ratio"]
    return make_lineplot(
        block, ["wdt_label", "hull_shape"], "x", "y",
        FIG_META["A5_A9"]["title"], FIG_META["A5_A9"]["subtitle"],
        FIG_META["A5_A9"]["x"], FIG_META["A5_A9"]["y"],
    )


def fig_simple(df, key, group_cols):
    meta = FIG_META[key]
    block = meta["filter"](df).copy()
    return make_lineplot(
        block, group_cols, "x", "y",
        meta["title"], meta["subtitle"], meta["x"], meta["y"],
    )


# ---------- Parametric comparisons ----------

def make_comparison_loaded_vs_ballast_cxc(df):
    """Cxc loaded (A5-A9, Cylindrical) vs Ballast (A12)."""
    loaded = df[(df.figure.isin(["A5", "A6", "A7", "A8", "A9"])) & (df.hull_shape == "Cylindrical")].copy()
    ballast = df[df.figure == "A12"].copy()
    fig = go.Figure()
    colors = {"1.1": "#1f77b4", "1.2": "#ff7f0e", "1.5": "#2ca02c", "3.0": "#d62728", "4.4+": "#9467bd"}
    for wdt, col in colors.items():
        sub = loaded[loaded.wdt_ratio == wdt].sort_values("x")
        if not sub.empty:
            fig.add_trace(go.Scatter(x=sub.x, y=sub.y, mode="lines+markers",
                                     name=f"Loaded Cyl WD/T={wdt}", line=dict(color=col, dash="solid", width=1.5),
                                     marker=dict(size=4)))
    # ballast: 4.4 Cylindrical reference
    sub_ref = ballast[ballast.hull_shape == "4.4 Cylindrical"].sort_values("x")
    if not sub_ref.empty:
        fig.add_trace(go.Scatter(x=sub_ref.x, y=sub_ref.y, mode="lines+markers",
                                 name="Ballast 40%T (4.4 Cyl ref)", line=dict(color="black", dash="dot", width=2),
                                 marker=dict(size=5)))
    # ballast envelope
    for stat in ["max", "mean", "min"]:
        sub = ballast[ballast.hull_shape == stat].sort_values("x")
        if not sub.empty:
            fig.add_trace(go.Scatter(x=sub.x, y=sub.y, mode="lines",
                                     name=f"Ballast envelope {stat}", line=dict(color="grey", dash="dash", width=1)))
    fig.update_layout(
        title=dict(text="<b>Parametric: Cxc Loaded (A5–A9 Cylindrical) vs Ballasted (A12)</b><br>"
                        "<span style='font-size:13px;color:#666'>Effect of loading state on longitudinal current drag, all WD/T overlaid</span>",
                   x=0.02, xanchor="left"),
        xaxis_title="Heading angle (deg)", yaxis_title="Cxc",
        legend=dict(orientation="v", x=1.02, y=1, font=dict(size=10)),
        height=520, margin=dict(l=70, r=240, t=80, b=60),
        plot_bgcolor="#fafafa", paper_bgcolor="white",
        xaxis=dict(gridcolor="#e0e0e0"), yaxis=dict(gridcolor="#e0e0e0"),
    )
    return fig


def make_comparison_prismatic_vs_spherical(df):
    block = df[df.figure.isin(["A17", "A18", "A19"])].copy()
    fig = go.Figure()
    style = {"Prismatic Tank": dict(dash="solid"), "Spherical Tank": dict(dash="dash")}
    coef_color = {"Cxw": "#1f77b4", "Cyw": "#ff7f0e", "Cxyw": "#2ca02c"}
    for coef, c_col in coef_color.items():
        for shape, line_style in style.items():
            sub = block[(block.coefficient == coef) & (block.hull_shape == shape)].sort_values("x")
            if sub.empty:
                continue
            fig.add_trace(go.Scatter(x=sub.x, y=sub.y, mode="lines+markers",
                                     name=f"{coef} — {shape}",
                                     line=dict(color=c_col, dash=line_style["dash"], width=2),
                                     marker=dict(size=4)))
    fig.update_layout(
        title=dict(text="<b>Parametric: Gas Carrier Wind Coefficients — Prismatic vs Spherical Tank</b><br>"
                        "<span style='font-size:13px;color:#666'>All three wind coefficients overlaid (A17 Cxw, A18 Cyw, A19 Cxyw)</span>",
                   x=0.02, xanchor="left"),
        xaxis_title="Heading angle (deg)", yaxis_title="Coefficient value",
        legend=dict(orientation="v", x=1.02, y=1, font=dict(size=10)),
        height=520, margin=dict(l=70, r=220, t=80, b=60),
        plot_bgcolor="#fafafa", paper_bgcolor="white",
        xaxis=dict(gridcolor="#e0e0e0"), yaxis=dict(gridcolor="#e0e0e0", zerolinecolor="#888"),
    )
    return fig


def make_comparison_cyl_vs_conv_ballast(df):
    block = df[df.figure == "A13"].copy()
    fig = go.Figure()
    colors = {"1.1 Cyl": ("#1f77b4", "solid"), "1.5 Cyl": ("#1f77b4", "dot"),
              "4.4 Cyl": ("#1f77b4", "dash"),
              "1.1 Conv": ("#d62728", "solid"), "1.2 Conv": ("#d62728", "dash")}
    for combo, (c, dash) in colors.items():
        wdt, hull = combo.split(" ")
        hull_full = "Cylindrical" if hull.startswith("Cyl") else "Conventional"
        sub = block[(block.wdt_ratio == wdt) & (block.hull_shape == hull_full)].sort_values("x")
        if sub.empty:
            continue
        fig.add_trace(go.Scatter(x=sub.x, y=sub.y, mode="lines+markers",
                                 name=combo, line=dict(color=c, dash=dash, width=2), marker=dict(size=4)))
    fig.update_layout(
        title=dict(text="<b>Parametric: Hull Bow Shape Effect (Ballast Cyc — A13)</b><br>"
                        "<span style='font-size:13px;color:#666'>Cylindrical vs Conventional bow at multiple WD/T</span>",
                   x=0.02, xanchor="left"),
        xaxis_title="Heading angle (deg)", yaxis_title="Cyc",
        legend=dict(orientation="v", x=1.02, y=1, font=dict(size=10)),
        height=480, margin=dict(l=70, r=200, t=80, b=60),
        plot_bgcolor="#fafafa", paper_bgcolor="white",
        xaxis=dict(gridcolor="#e0e0e0"), yaxis=dict(gridcolor="#e0e0e0"),
    )
    return fig


def make_polar_overlay(df, figure_key, coef, group_col):
    """Polar plot preserving sign of the coefficient.

    REFACTORED for #616 (Phase 5): delegates to
    digitalmodel.marine_ops.marine_engineering.visualization.polar_force_overlay,
    which adds a transparent vessel silhouette in the middle + on-body force
    arrows at sampled headings. The data-trace shape (named pos/neg traces per
    group) is preserved for regression-test compatibility per the trace-signature
    fixture captured pre-refactor.

    Citation gate: OCIMF MEG3 §A1 / MEG4 §A2 convention is bound by
    OCIMF_CONVENTION_AUTHORITY in the new module's _convention.py.
    """
    from digitalmodel.hydrodynamics.hull_library.profile_schema import (
        HullProfile, HullStation, HullType,
    )
    from digitalmodel.marine_ops.marine_engineering.visualization import (
        ForceArrowKind, VesselSilhouetteSpec, polar_force_overlay,
    )

    # Convert build-script DataFrame to long-format expected by the new module
    block = df[df.figure == figure_key].copy()
    block = block.rename(columns={"x": "theta_deg", "y": "value"})
    block["component"] = coef
    # theta in workbook may be in [0, 360]; clamp/validate
    block = block[(block["theta_deg"] >= 0) & (block["theta_deg"] < 360)]

    # Default silhouette per figure-key convention: tanker for Cxc/Cyc/Cxyc,
    # gas_carrier for Cxw/Cyw/Cxyw. Coef string discriminates.
    silhouette_kind = "gas_carrier" if coef.endswith("w") else "tanker"
    # Representative VLCC dimensions for tanker / standard for gas carrier.
    if silhouette_kind == "tanker":
        length_bp, beam, draft, depth = 320.0, 58.0, 22.0, 30.0
    else:
        length_bp, beam, draft, depth = 280.0, 44.0, 11.5, 24.0
    half_beam = beam / 2.0
    hull = HullProfile(
        name=f"reference-{silhouette_kind}",
        hull_type=HullType.TANKER if silhouette_kind == "tanker" else HullType.LNGC,
        length_bp=length_bp, beam=beam, draft=draft, depth=depth,
        source=f"reference silhouette for OCIMF {coef} explorer (no engineering use)",
        stations=[
            HullStation(x_position=0.0,
                        waterline_offsets=[(0.0, 0.0), (depth * 0.7, half_beam * 0.5)]),
            HullStation(x_position=length_bp / 2.0,
                        waterline_offsets=[(0.0, 0.0), (depth * 0.7, half_beam)]),
            HullStation(x_position=length_bp,
                        waterline_offsets=[(0.0, 0.0), (depth * 0.7, half_beam * 0.5)]),
        ],
    )
    silhouette = VesselSilhouetteSpec(hull_profile=hull, silhouette_kind=silhouette_kind)

    title_text = (f"<b>Polar diagram — {figure_key} {coef} (signed)</b><br>"
                  f"<span style='font-size:13px;color:#666'>r=|{coef}|, θ=incidence heading; "
                  f"transparent silhouette = vessel (bow up); on-body arrows indicate "
                  f"force direction per OCIMF MEG3 §A1 / MEG4 §A2 convention. "
                  f"Solid/circle = positive Cy (force toward +Y_body = port per OCIMF).</span>")

    fig = polar_force_overlay(
        data=block,
        silhouette=silhouette,
        force_arrow_kind=ForceArrowKind.LATERAL_ONLY,
        title=title_text,
        group_col=group_col,
        arrow_sample_step_deg=30.0,
    )
    # Update radial-axis title to match original ("|{coef}|") for visual continuity.
    fig.update_layout(polar=dict(radialaxis=dict(
        gridcolor="#e0e0e0", title=dict(text=f"|{coef}|", font=dict(size=11)),
    )))
    return fig


# ---------- HTML composer ----------

def fig_to_html(fig, include_plotlyjs=False):
    return pio.to_html(fig, include_plotlyjs=include_plotlyjs, full_html=False, div_id=None,
                       config={"displaylogo": False, "responsive": True})


CSS = """
:root { --accent:#1f4e79; --soft:#f4f6f8; --rule:#e1e5ea; --note:#666; }
* { box-sizing: border-box; }
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
       max-width: 1400px; margin: 0 auto; padding: 24px; color: #222; background: #fff; line-height: 1.55; }
h1 { color: var(--accent); border-bottom: 3px solid var(--accent); padding-bottom: 8px; font-size: 28px; }
h2 { color: var(--accent); border-bottom: 1px solid var(--rule); padding-bottom: 4px; font-size: 22px; margin-top: 40px; }
h3 { color: #333; font-size: 17px; margin-top: 28px; }
section { margin-bottom: 36px; }
.callout { background: var(--soft); border-left: 4px solid var(--accent); padding: 12px 16px; margin: 16px 0; border-radius: 0 4px 4px 0; }
.warn { background: #fff8e0; border-left-color: #e8a800; }
.danger { background: #fde8e8; border-left-color: #c0392b; }
.formula { font-family: "Courier New", monospace; background: var(--soft); padding: 10px 14px; border-radius: 4px;
           display: inline-block; font-size: 14px; }
table.spec { border-collapse: collapse; width: 100%; margin: 12px 0; font-size: 14px; }
table.spec th, table.spec td { border: 1px solid var(--rule); padding: 6px 10px; text-align: left; vertical-align: top; }
table.spec th { background: var(--soft); font-weight: 600; }
.fig-wrap { margin: 20px 0; padding: 10px; background: #fff; border: 1px solid var(--rule); border-radius: 6px; }
.fig-note { font-size: 13px; color: var(--note); margin-top: 8px; padding: 6px 12px; border-left: 3px solid var(--rule); }
.meta { font-size: 12px; color: var(--note); }
.toc { background: var(--soft); border: 1px solid var(--rule); border-radius: 6px; padding: 14px 20px; }
.toc ul { margin: 0; padding-left: 20px; }
.toc a { color: var(--accent); text-decoration: none; }
.toc a:hover { text-decoration: underline; }
code { background: var(--soft); padding: 1px 5px; border-radius: 3px; font-size: 13px; }
.path { font-family: "Courier New", monospace; font-size: 12px; color: #555; word-break: break-all; }
"""

USAGE_HTML = """
<h2 id="how-to-use">2. How to use these coefficients</h2>

<h3>2.1 Wind force formulation (OCIMF MEG3 §3.2 / MEG4 Ch.3)</h3>
<div class="callout">
<p>Wind forces and yaw moment are computed from coefficient lookups on a vessel-specific reference area and projected dimensions:</p>
<p class="formula">F<sub>xw</sub> = (½) · ρ<sub>a</sub> · V<sub>w</sub>² · A<sub>T</sub> · C<sub>xw</sub>(θ<sub>w</sub>)</p>
<p class="formula">F<sub>yw</sub> = (½) · ρ<sub>a</sub> · V<sub>w</sub>² · A<sub>L</sub> · C<sub>yw</sub>(θ<sub>w</sub>)</p>
<p class="formula">M<sub>xyw</sub> = (½) · ρ<sub>a</sub> · V<sub>w</sub>² · A<sub>L</sub> · L<sub>BP</sub> · C<sub>xyw</sub>(θ<sub>w</sub>)</p>
<p>where ρ<sub>a</sub> ≈ 1.226 kg/m³, V<sub>w</sub> is wind speed (10 m reference), A<sub>T</sub> is transverse projected area, A<sub>L</sub> is lateral projected area, L<sub>BP</sub> is length-between-perpendiculars.</p>
</div>

<h3>2.2 Current force formulation</h3>
<div class="callout">
<p class="formula">F<sub>xc</sub> = (½) · ρ<sub>w</sub> · V<sub>c</sub>² · L<sub>BP</sub> · T · C<sub>xc</sub>(θ<sub>c</sub>, WD/T)</p>
<p class="formula">F<sub>yc</sub> = (½) · ρ<sub>w</sub> · V<sub>c</sub>² · L<sub>BP</sub> · T · C<sub>yc</sub>(θ<sub>c</sub>, WD/T)</p>
<p class="formula">M<sub>xyc</sub> = (½) · ρ<sub>w</sub> · V<sub>c</sub>² · L<sub>BP</sub>² · T · C<sub>xyc</sub>(θ<sub>c</sub>, WD/T)</p>
<p>where ρ<sub>w</sub> ≈ 1025 kg/m³, V<sub>c</sub> is current velocity (apply K-factor correction from Figure A16 if free-stream velocity is referenced), L<sub>BP</sub> is length-between-perpendiculars, T is draft.</p>
<p><b>Key parameter:</b> WD/T = water depth / draft ratio. Coefficient magnitude is highly sensitive to shallow water (low WD/T) and lateral coefficient especially.</p>
</div>

<h3>2.3 Step-by-step usage</h3>
<ol>
<li><b>Pick the right figure</b> by vessel class and condition:
  <ul>
    <li><i>Tanker, loaded</i> → A5–A9 (Cxc), A10 (Cyc), A11 (Cxyc)</li>
    <li><i>Tanker, ballasted at 40% T</i> → A12 (Cxc), A13 (Cyc), A14 (Cxyc)</li>
    <li><i>Gas carrier (LNG/LPG)</i> → A17 (Cxw), A18 (Cyw), A19 (Cxyw) — wind only; current coefs from tanker tables with appropriate area</li>
    <li><i>Current velocity correction</i> → A16 (K factor)</li>
  </ul>
</li>
<li><b>Determine secondary axis value</b>:
  <ul>
    <li>For current coefficients: compute WD/T = water depth / draft. Pick the figure curve at the closest WD/T (or interpolate between).</li>
    <li>For wind coefficients (gas carrier): pick tank type — prismatic membrane or spherical (Moss).</li>
    <li>For loaded-tanker Cxc (A5–A9): pick hull bow shape — cylindrical or conventional.</li>
  </ul>
</li>
<li><b>Look up coefficient</b> at the heading angle of interest (0° = bow-on, 90° = beam, 180° = stern-on).</li>
<li><b>Compute force/moment</b> using formulas above and vessel-specific reference areas.</li>
<li><b>Sum wind + current contributions</b> in the vessel-fixed reference frame.</li>
</ol>

<h3>2.4 Sign conventions and angle reference</h3>
<div class="callout warn">
<p>OCIMF uses a vessel-fixed coordinate system. Heading angle θ is measured from the bow (0°) clockwise toward the starboard beam (90°), stern (180°), port beam (270°). Cyw symmetric vessels have C<sub>x</sub> antisymmetric and C<sub>y</sub> symmetric about θ=180°, so the published tables span 0°–180° only.</p>
<p>Coefficient signs reflect force direction in the vessel frame:</p>
<ul>
<li>Positive C<sub>xw</sub>/C<sub>xc</sub>: force pointing toward stern (head wind/current pushes stern-ward when ahead of beam)</li>
<li>Positive C<sub>yw</sub>/C<sub>yc</sub>: force pointing toward starboard (transverse wind from port produces starboard force)</li>
<li>Positive C<sub>xyw</sub>/C<sub>xyc</sub>: bow-toward-starboard yaw moment</li>
</ul>
</div>

<h3>2.5 K-factor correction (Figure A16)</h3>
<div class="callout">
<p>When the available current measurement is a free-stream / undisturbed-flow velocity (e.g., from a metocean buoy upstream of the vessel), the actual velocity at vessel midships is reduced due to blockage. The K-factor scales the design current velocity:</p>
<p class="formula">V<sub>c, design</sub> = K(% load, WD/T) · V<sub>c, free-stream</sub></p>
<p>K depends on % full-load condition (0–100%) and WD/T. For deep water (WD/T ≥ 3), K is close to 1.0 across all loading conditions. For shallow water (WD/T ≤ 1.2), K drops to ~0.9 at light load.</p>
</div>
"""

PROVENANCE_HTML = """
<h2 id="provenance">5. Data provenance and citation</h2>
<table class="spec">
<tr><th>Artifact</th><th>Path</th><th>Role</th></tr>
<tr><td>Full MEG3 (2008) PDF</td><td class="path">/mnt/ace/acma-codes/OCIMF/OCIMF - 2008 - Mooring Equipment Guidelines.pdf</td><td>Canonical published standard, 3rd edition</td></tr>
<tr><td>Full MEG4 (2018) PDF</td><td class="path">/mnt/ace/acma-codes/OCIMF (MEG 4)/Mooring Equipment Guidelines - MEG4.pdf</td><td>Canonical published standard, 4th edition</td></tr>
<tr><td><b>OCIMF Coef.xlsx (this explorer's data source)</b></td><td class="path">/mnt/ace/acma-codes/OCIMF/OCIMF Coef.xlsx</td><td>Digitized Annex A coefficient lookup tables (Figures A5–A19)</td></tr>
<tr><td>WebPlotDigitizer alt extraction (A5–A9 only)</td><td class="path">/mnt/ace/acma-codes/OCIMF 3rd ed/Digitizer/</td><td>Independent re-digitization for cross-validation</td></tr>
<tr><td>Figure PDFs (15 files)</td><td class="path">/mnt/ace/acma-codes/OCIMF/Figures/</td><td>Per-figure extract PDFs</td></tr>
<tr><td>Project-bundled MEG4 extracts</td><td class="path">/mnt/ace/acma-projects/B1522/ctr-7/_data/ocimf/</td><td>Project-specific bundle (Fig. A13–A21 + worked examples)</td></tr>
</table>
<div class="callout">
<p><b>Citation contract:</b> Per <code>.claude/rules/calc-citation-contract.md</code>, any digitalmodel calc module emitting a coefficient from this data must produce a <code>Citation</code> sidecar binding to <code>wiki/standards/ocimf-meg3.md</code> or <code>wiki/standards/ocimf-meg4.md</code>. Those wiki pages are tracked at <a href="https://github.com/vamseeachanta/workspace-hub/issues/2284">workspace-hub#2284</a>. Precedent: DNV-OS-E301 pilot at <code>digitalmodel/src/digitalmodel/orcaflex/mooring_design.py::check_mbl_with_safety_factor</code>.</p>
</div>
<div class="callout warn">
<p><b>Coverage gaps:</b></p>
<ul>
<li><b>Figure A15 absent</b> from the digitization (sheet jumps Data 10a-14a → Data 16a-19a). Cross-check against the standard PDFs if A15 is needed.</li>
<li>Gas carrier <b>current</b> coefficients not in this workbook (only wind). Use tanker-class current tables with gas-carrier-specific reference area.</li>
<li>Tanker <b>wind</b> coefficients not in this workbook (only current). Use OCIMF 1994 VLCC wind prediction tables or vessel-specific wind-tunnel data.</li>
</ul>
</div>
"""

FOOTER_HTML = """
<h2 id="related">6. Related references and follow-up</h2>
<ul>
<li><a href="https://github.com/vamseeachanta/workspace-hub/issues/2284">workspace-hub#2284</a> — wiki promotion of MEG3/MEG4 to mooring wiki domain</li>
<li><a href="https://github.com/vamseeachanta/workspace-hub/issues/2278">workspace-hub#2278</a> — OCIMF MEG fragments reconcile (resolved 2026-05-20)</li>
<li><a href="https://github.com/vamseeachanta/digitalmodel/issues/556">digitalmodel#556</a> — OCIMF coefficient interpolation fix (real-data ingestion path)</li>
<li><a href="https://github.com/vamseeachanta/digitalmodel/issues/563">digitalmodel#563</a> — OCIMF database performance (adapter scope)</li>
<li><a href="https://github.com/vamseeachanta/workspace-hub/issues/2685">workspace-hub#2685</a> — DNV-OS-E301 citation pilot (precedent for OCIMF citation pattern)</li>
</ul>
<p class="meta">Generated 2026-05-20 from <code>/mnt/ace/acma-codes/OCIMF/OCIMF Coef.xlsx</code> · explorer build script <code>/tmp/build_ocimf_explorer.py</code></p>
"""


def render_html(figs_meta):
    parts = []
    parts.append("<!DOCTYPE html><html lang='en'><head><meta charset='utf-8'>")
    parts.append("<title>OCIMF MEG3/MEG4 Coefficient Explorer</title>")
    parts.append(f"<style>{CSS}</style>")
    # plotly.js loaded once via CDN
    parts.append('<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>')
    parts.append("</head><body>")
    parts.append("<h1>OCIMF MEG3/MEG4 Coefficient Explorer</h1>")
    parts.append(
        "<p class='meta'>Interactive visualization of digitized Annex A wind and current coefficients "
        "for tankers and gas carriers — sourced from <code>/mnt/ace/acma-codes/OCIMF/OCIMF Coef.xlsx</code>, "
        "covering OCIMF Mooring Equipment Guidelines Annex A Figures A5–A19. "
        "Generated 2026-05-20.</p>"
    )

    # TOC
    parts.append("<section class='toc'><h3 style='margin-top:0'>Contents</h3><ul>")
    parts.append("<li><a href='#overview'>1. Overview</a></li>")
    parts.append("<li><a href='#how-to-use'>2. How to use these coefficients</a></li>")
    parts.append("<li><a href='#figures'>3. Coefficient plots — per figure</a></li>")
    parts.append("<ul>")
    for key, meta in FIG_META.items():
        parts.append(f"<li><a href='#fig-{key}'>{meta['title']}</a></li>")
    parts.append("</ul>")
    parts.append("<li><a href='#parametric'>4. Parametric comparisons</a></li>")
    parts.append("<li><a href='#provenance'>5. Data provenance and citation</a></li>")
    parts.append("<li><a href='#related'>6. Related references</a></li>")
    parts.append("</ul></section>")

    # Section 1
    parts.append("<section>")
    parts.append("<h2 id='overview'>1. Overview</h2>")
    parts.append("<p>The OCIMF Mooring Equipment Guidelines provide dimensionless force coefficients used to compute wind and current loads on moored vessels for mooring design (MBL sizing, fairlead arrangement, fender selection). This explorer renders all digitized Annex A figures and supports parametric comparisons across the principal variables.</p>")
    parts.append("<table class='spec'>")
    parts.append("<tr><th>Variable</th><th>Role</th><th>Range</th><th>Notes</th></tr>")
    parts.append("<tr><td>Heading angle θ</td><td>Primary axis on all coefficient lookups</td><td>0° (bow) – 180° (stern)</td><td>Symmetric vessels: 0–180 spans full physical range</td></tr>")
    parts.append("<tr><td>WD/T ratio</td><td>Water depth / draft (current only)</td><td>1.05 – >6</td><td>Shallow water amplifies lateral current loads dramatically</td></tr>")
    parts.append("<tr><td>Hull bow shape</td><td>Tanker Cxc loaded (A5–A9), Cyc/Cxyc ballast (A13/A14)</td><td>Cylindrical / Conventional</td><td>Conventional bows show stronger asymmetric Cxc near head-current angles</td></tr>")
    parts.append("<tr><td>Tank type</td><td>Gas carrier wind (A17–A19)</td><td>Prismatic / Spherical (Moss)</td><td>Spherical tanks have larger lateral wind area for same displacement</td></tr>")
    parts.append("<tr><td>Loading state</td><td>Tanker current (A5–A11 loaded, A12–A14 ballast 40% T)</td><td>Loaded / Ballast 40% T</td><td>Ballast condition typically gives smaller current loads at large heading</td></tr>")
    parts.append("</table>")
    parts.append("</section>")

    # Section 2
    parts.append(f"<section>{USAGE_HTML}</section>")

    # Section 3 — per-figure plots
    parts.append("<section><h2 id='figures'>3. Coefficient plots — per figure</h2>")
    for key, html_block, note in figs_meta["per_figure"]:
        parts.append(f"<div class='fig-wrap' id='fig-{key}'>")
        parts.append(html_block)
        if note:
            parts.append(f"<div class='fig-note'>{note}</div>")
        parts.append("</div>")
    parts.append("</section>")

    # Section 4 — parametric comparisons
    parts.append("<section><h2 id='parametric'>4. Parametric comparisons</h2>")
    parts.append("<p>These cross-cutting views surface the principal sensitivities engineers most often need to inspect when selecting design coefficients for a specific mooring analysis.</p>")
    for title, html_block, note in figs_meta["parametric"]:
        parts.append(f"<h3>{html.escape(title)}</h3>")
        parts.append(f"<div class='fig-wrap'>{html_block}")
        if note:
            parts.append(f"<div class='fig-note'>{note}</div>")
        parts.append("</div>")
    parts.append("</section>")

    # Section 5 + 6
    parts.append(f"<section>{PROVENANCE_HTML}</section>")
    parts.append(f"<section>{FOOTER_HTML}</section>")

    parts.append("</body></html>")
    return "\n".join(parts)


def main():
    print("Loading workbook...")
    df = load_all()
    print(f"Total long-form rows: {len(df)}")
    print(f"Figures: {sorted(df.figure.unique())}")
    print(f"Coefficients: {sorted(df.coefficient.unique())}")
    print(f"Vessel classes: {sorted(df.vessel_class.unique())}")

    figs_meta = {"per_figure": [], "parametric": []}

    # Per-figure plots
    print("Building per-figure plots...")
    figs_meta["per_figure"].append((
        "A5_A9",
        fig_to_html(fig_a5_a9(df)),
        "Curves shown for both Cylindrical and Conventional bow shapes; observe how WD/T strongly modulates Cxc magnitude at oblique angles near 90°.",
    ))
    figs_meta["per_figure"].append((
        "A10",
        fig_to_html(fig_simple(df, "A10", ["wdt_ratio"])),
        "Lateral current drag scales dramatically with reducing WD/T at beam (≈90°). At WD/T=1.05 the coefficient exceeds 3.0; deep water (>6) is below 0.3.",
    ))
    figs_meta["per_figure"].append((
        "A11",
        fig_to_html(fig_simple(df, "A11", ["wdt_ratio"])),
        "Yaw moment is signed — bow-from-current produces opposite sign to stern-from-current. Peak magnitude near ~30° and ~150° heading.",
    ))
    figs_meta["per_figure"].append((
        "A12",
        fig_to_html(fig_simple(df, "A12", ["hull_shape"])),
        "A12 reports a statistical envelope (max/mean/min) over multiple ballasted vessel configurations rather than per-WD/T lookup. The 4.4 Cylindrical reference line is provided for direct comparison with A5–A9.",
    ))
    figs_meta["per_figure"].append((
        "A13",
        fig_to_html(fig_simple(df, "A13", ["wdt_ratio", "hull_shape"])),
        "Bow shape effect is pronounced for ballasted tankers — conventional bows show different lateral drag distribution vs cylindrical at the same WD/T.",
    ))
    figs_meta["per_figure"].append((
        "A14",
        fig_to_html(fig_simple(df, "A14", ["wdt_ratio", "hull_shape"])),
        "Yaw moment for ballasted tankers (40% T) referenced to midships. Sign convention as in A11.",
    ))
    figs_meta["per_figure"].append((
        "A16",
        fig_to_html(fig_simple(df, "A16", ["wdt_ratio"])),
        "Apply K as a multiplier to free-stream current velocity before computing forces. Deep-water curves (WD/T=3, 1.5) sit near K≈1; shallow-water at light load drops to K≈0.9.",
    ))
    figs_meta["per_figure"].append((
        "A17",
        fig_to_html(fig_simple(df, "A17", ["hull_shape"])),
        "Longitudinal wind drag — prismatic-tank LNGCs show different above-water profile vs Moss-spherical tankers.",
    ))
    figs_meta["per_figure"].append((
        "A18",
        fig_to_html(fig_simple(df, "A18", ["hull_shape"])),
        "Lateral wind drag — typically the dominant gas-carrier wind component at beam wind.",
    ))
    figs_meta["per_figure"].append((
        "A19",
        fig_to_html(fig_simple(df, "A19", ["hull_shape"])),
        "Wind yaw moment — magnitude small, but sign and shape matter for fairlead arrangement design.",
    ))

    # Parametric comparisons
    print("Building parametric comparisons...")
    figs_meta["parametric"].append((
        "Cxc Loaded vs Ballasted Tanker (effect of loading state)",
        fig_to_html(make_comparison_loaded_vs_ballast_cxc(df)),
        "Loaded-tanker Cxc (A5–A9, cylindrical bow only) is overlaid against the A12 ballasted envelope. Useful for sizing for the conservative load case across vessel-condition uncertainty.",
    ))
    figs_meta["parametric"].append((
        "Gas Carrier wind: Prismatic vs Spherical tank (all three wind coefficients)",
        fig_to_html(make_comparison_prismatic_vs_spherical(df)),
        "All three wind coefficients (Cxw, Cyw, Cxyw) overlaid for both tank types. Spherical tanks typically show larger Cyw magnitudes at beam wind.",
    ))
    figs_meta["parametric"].append((
        "Hull bow shape effect — Ballast Cyc (A13)",
        fig_to_html(make_comparison_cyl_vs_conv_ballast(df)),
        "Direct comparison of how bow shape selection changes ballast lateral drag at a fixed WD/T. Conventional bows appear at the smaller WD/T values only.",
    ))
    figs_meta["parametric"].append((
        "Polar — Tanker Loaded Cyc by WD/T (A10)",
        fig_to_html(make_polar_overlay(df, "A10", "Cyc", "wdt_ratio")),
        "Polar projection emphasizes the directional pattern of lateral current drag. The radial extent at WD/T=1.05 vastly exceeds deep-water — visual confirmation that shallow water dominates lateral mooring loads. All Cyc values in A10 are positive in OCIMF's vessel-fixed convention, so all traces appear solid/circle (force points to +Y starboard for a starboard-incidence current heading).",
    ))
    figs_meta["parametric"].append((
        "Polar — Gas Carrier Cyw by tank type (A18)",
        fig_to_html(make_polar_overlay(df, "A18", "Cyw", "hull_shape")),
        "Gas carrier lateral wind drag polar — prismatic and spherical curves are similar at most headings; differences cluster near beam wind. All Cyw values are negative across 0-180° (all traces dashed/×) because OCIMF tabulates only the starboard-incidence half; the negative sign indicates lateral force pointing to port. For port-incidence wind (180-360°), the published curves are mirrored with sign flipped (Cyw becomes positive, force to starboard) — vessel port/starboard symmetry.",
    ))

    print("Rendering HTML...")
    html_out = render_html(figs_meta)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(html_out, encoding="utf-8")
    size_kb = OUT.stat().st_size / 1024
    print(f"Wrote {OUT} ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
