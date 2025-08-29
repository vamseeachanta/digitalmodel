#!/usr/bin/env python3
"""
Inspect the structure of the revised Excel file to understand the data layout.
"""

import openpyxl
from pathlib import Path

def inspect_excel():
    """Inspect Excel file structure."""
    # Set up paths
    script_dir = Path(__file__).parent
    data_dir = script_dir.parent / 'data'
    excel_file = data_dir / 'B1512 Gyradius Calcs_rev2.xlsx'
    
    print(f"Inspecting: {excel_file}")
    print("="*80)
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    # List all sheets
    print("\nAvailable sheets:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    # Check if 'Calcs' sheet exists
    if 'Calcs' in wb.sheetnames:
        ws = wb['Calcs']
        print(f"\n'Calcs' sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Check vessel dimensions
        print("\nVessel dimensions (rows 7-9):")
        for row in range(7, 10):
            row_data = []
            for col in range(1, min(6, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                row_data.append(f"{cell.value}")
            print(f"  Row {row}: {' | '.join(row_data)}")
        
        # Check weight configurations
        print("\nChecking weight configuration areas:")
        
        # Configuration 1 area (around row 21)
        print("\nConfiguration 1 area (rows 19-27):")
        for row in range(19, 28):
            row_data = []
            for col in range(1, min(9, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                row_data.append(f"{cell.value}")
            print(f"  Row {row}: {' | '.join(row_data[:5])}...")  # Just show first 5 columns
        
        # Configuration 2 area (around row 42)
        print("\nConfiguration 2 area (rows 40-48):")
        for row in range(40, 49):
            row_data = []
            for col in range(1, min(9, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                row_data.append(f"{cell.value}")
            print(f"  Row {row}: {' | '.join(row_data[:5])}...")
            
    # Check Summary sheet if it exists
    if 'Summary' in wb.sheetnames:
        ws_summary = wb['Summary']
        print(f"\n'Summary' sheet dimensions: {ws_summary.max_row} rows x {ws_summary.max_column} columns")
        
        print("\nFirst 10 rows of Summary sheet:")
        for row in range(1, min(11, ws_summary.max_row + 1)):
            row_data = []
            for col in range(1, min(10, ws_summary.max_column + 1)):
                cell = ws_summary.cell(row=row, column=col)
                row_data.append(f"{cell.value}")
            print(f"  Row {row}: {' | '.join(row_data[:5])}...")
    
    wb.close()
    print("\n" + "="*80)
    print("Inspection complete")

if __name__ == "__main__":
    inspect_excel()