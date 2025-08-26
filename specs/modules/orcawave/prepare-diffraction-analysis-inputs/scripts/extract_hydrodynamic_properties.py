#!/usr/bin/env python3
"""
Extract hydrodynamic properties from Excel file and generate hydrodynamic.yml

This script reads vessel mass properties from the B1512 Gyradius Calcs.xlsx file
and generates a standardized hydrodynamic.yml file for use with OrcaWave, AQWA,
and other hydrodynamic analysis programs.
"""

import os
import yaml
import openpyxl
from pathlib import Path
from typing import Dict, List, Any


def extract_vessel_properties(excel_path: str, draft_ft: float = None) -> List[Dict[str, Any]]:
    """
    Extract vessel properties from Excel file for all configurations.
    
    Args:
        excel_path: Path to the Excel file
        draft_ft: Vessel draft in feet (optional, will use estimated value if not provided)
        
    Returns:
        List of dictionaries containing vessel properties for each configuration
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Get vessel dimensions
    ws_calcs = wb['Calcs']
    lbp_ft = ws_calcs.cell(row=7, column=4).value  # 75 ft
    beam_ft = ws_calcs.cell(row=8, column=4).value  # 28 ft
    depth_ft = ws_calcs.cell(row=9, column=4).value  # 10 ft
    
    # If draft not provided, estimate as typical for supply vessel (60-70% of depth)
    if draft_ft is None:
        draft_ft = depth_ft * 0.65  # 6.5 ft estimated
        print(f"Note: Draft not specified, using estimated value of {draft_ft:.1f} ft based on vessel depth of {depth_ft} ft")
    
    # Configuration data locations in Calcs sheet
    config_data = [
        {
            'name': 'incident_draft_fuel_centered',
            'description': 'Incident Draft Fuel centered',
            'draft_ft': draft_ft,  # Incident draft condition
            'total_row': 21,
            'mass_row': 24,
            'ixx_row': 25,
            'iyy_row': 26,
            'izz_row': 27
        },
        {
            'name': 'incident_draft_fuel_centered_adjusted',
            'description': 'Incident Draft Fuel centered (adjusted)',
            'draft_ft': draft_ft,  # Same draft, fuel redistributed
            'total_row': 42,
            'mass_row': 45,
            'ixx_row': 46,
            'iyy_row': 47,
            'izz_row': 48
        },
        {
            'name': 'fo_to_port',
            'description': '4070 gallons FO to Port',
            'draft_ft': draft_ft + 0.2,  # Slightly deeper due to added fuel
            'total_row': 63,
            'mass_row': 66,
            'ixx_row': 67,
            'iyy_row': 68,
            'izz_row': 69
        },
        {
            'name': 'fo_to_port_with_ingress',
            'description': '4070 gallons FO to Port w/ 5% Eng Rm Ingress',
            'draft_ft': draft_ft + 0.4,  # Deeper due to fuel and water ingress
            'total_row': 84,
            'mass_row': 87,
            'ixx_row': 88,
            'iyy_row': 89,
            'izz_row': 90
        }
    ]
    
    # Gyradii from Summary sheet
    gyradii_data = [
        {'kxx': 8.656621927605684, 'kyy': 18.040519241571527, 'kzz': 19.11849260368657},
        {'kxx': 8.660674737013315, 'kyy': 18.038106251597068, 'kzz': 19.122604438246967},
        {'kxx': 8.697252887973765, 'kyy': 17.903586333230976, 'kzz': 18.976617866183517},
        {'kxx': 8.718207413396735, 'kyy': 17.76770102813071, 'kzz': 18.833970399464512}
    ]
    
    vessels = []
    
    for idx, config in enumerate(config_data):
        # Extract mass and CoG
        mass_lt = ws_calcs.cell(row=config['total_row'], column=5).value
        lcg_ft = ws_calcs.cell(row=config['total_row'], column=6).value
        tcg_ft = ws_calcs.cell(row=config['total_row'], column=7).value
        vcg_ft = ws_calcs.cell(row=config['total_row'], column=8).value
        
        # Extract moments of inertia (in LT-ft²)
        ixx_lt_ft2 = ws_calcs.cell(row=config['ixx_row'], column=3).value
        iyy_lt_ft2 = ws_calcs.cell(row=config['iyy_row'], column=3).value
        izz_lt_ft2 = ws_calcs.cell(row=config['izz_row'], column=3).value
        
        # Get gyradii from summary data
        gyradii = gyradii_data[idx]
        
        # Get draft for this configuration
        config_draft_ft = config['draft_ft']
        
        # Convert units to SI (metric)
        # 1 LT (long ton) = 1016.0469 kg
        # 1 ft = 0.3048 m
        LT_TO_KG = 1016.0469
        FT_TO_M = 0.3048
        
        mass_kg = mass_lt * LT_TO_KG
        lcg_m = lcg_ft * FT_TO_M
        tcg_m = tcg_ft * FT_TO_M
        vcg_m = vcg_ft * FT_TO_M
        draft_m = config_draft_ft * FT_TO_M
        
        # Convert moments of inertia from LT-ft² to kg-m²
        # 1 LT-ft² = 1016.0469 kg * (0.3048 m)² = 94.485 kg-m²
        LT_FT2_TO_KG_M2 = LT_TO_KG * (FT_TO_M ** 2)
        
        ixx_kg_m2 = ixx_lt_ft2 * LT_FT2_TO_KG_M2
        iyy_kg_m2 = iyy_lt_ft2 * LT_FT2_TO_KG_M2
        izz_kg_m2 = izz_lt_ft2 * LT_FT2_TO_KG_M2
        
        # Convert gyradii from ft to m
        kxx_m = gyradii['kxx'] * FT_TO_M
        kyy_m = gyradii['kyy'] * FT_TO_M
        kzz_m = gyradii['kzz'] * FT_TO_M
        
        vessel_data = {
            'name': config['name'],
            'description': config['description'],
            'draft': {
                'value': round(draft_m, 3),
                'unit': 'm',
                'original': {
                    'value': config_draft_ft,
                    'unit': 'ft'
                }
            },
            'mass_properties': {
                'mass': {
                    'value': round(mass_kg, 2),
                    'unit': 'kg',
                    'original': {
                        'value': mass_lt,
                        'unit': 'LT'
                    }
                },
                'center_of_gravity': {
                    'x': {
                        'value': round(lcg_m, 4),
                        'unit': 'm',
                        'description': 'Longitudinal (from Fr. 0, +Fwd)',
                        'original': {
                            'value': lcg_ft,
                            'unit': 'ft'
                        }
                    },
                    'y': {
                        'value': round(tcg_m, 4),
                        'unit': 'm',
                        'description': 'Transverse (from CL, +Port)',
                        'original': {
                            'value': tcg_ft,
                            'unit': 'ft'
                        }
                    },
                    'z': {
                        'value': round(vcg_m, 4),
                        'unit': 'm',
                        'description': 'Vertical (from ABL, +Up)',
                        'original': {
                            'value': vcg_ft,
                            'unit': 'ft'
                        }
                    }
                },
                'moments_of_inertia': {
                    'ixx': {
                        'value': round(ixx_kg_m2, 2),
                        'unit': 'kg.m^2',
                        'description': 'Roll moment of inertia',
                        'original': {
                            'value': ixx_lt_ft2,
                            'unit': 'LT.ft^2'
                        }
                    },
                    'iyy': {
                        'value': round(iyy_kg_m2, 2),
                        'unit': 'kg.m^2',
                        'description': 'Pitch moment of inertia',
                        'original': {
                            'value': iyy_lt_ft2,
                            'unit': 'LT.ft^2'
                        }
                    },
                    'izz': {
                        'value': round(izz_kg_m2, 2),
                        'unit': 'kg.m^2',
                        'description': 'Yaw moment of inertia',
                        'original': {
                            'value': izz_lt_ft2,
                            'unit': 'LT.ft^2'
                        }
                    }
                },
                'radii_of_gyration': {
                    'kxx': {
                        'value': round(kxx_m, 4),
                        'unit': 'm',
                        'description': 'Roll radius of gyration',
                        'original': {
                            'value': gyradii['kxx'],
                            'unit': 'ft'
                        }
                    },
                    'kyy': {
                        'value': round(kyy_m, 4),
                        'unit': 'm',
                        'description': 'Pitch radius of gyration',
                        'original': {
                            'value': gyradii['kyy'],
                            'unit': 'ft'
                        }
                    },
                    'kzz': {
                        'value': round(kzz_m, 4),
                        'unit': 'm',
                        'description': 'Yaw radius of gyration',
                        'original': {
                            'value': gyradii['kzz'],
                            'unit': 'ft'
                        }
                    }
                }
            }
        }
        
        vessels.append(vessel_data)
    
    wb.close()
    return vessels


def create_simplified_format(vessels: List[Dict[str, Any]], vessel_dims: Dict = None) -> Dict[str, Any]:
    """
    Create a simplified format suitable for direct use in hydrodynamic programs.
    
    Args:
        vessels: List of vessel configurations
        vessel_dims: Dictionary with vessel dimensions (LBP, beam, depth)
        
    Returns:
        Simplified dictionary format
    """
    simplified = {
        'vessel': {
            'name': 'Sea Cypress',
            'type': 'Supply Vessel',
            'configurations': {}
        }
    }
    
    if vessel_dims:
        simplified['vessel']['dimensions'] = vessel_dims
    
    for vessel in vessels:
        config_name = vessel['name']
        
        # Extract just the values in SI units
        mass_props = vessel['mass_properties']
        
        simplified['vessel']['configurations'][config_name] = {
            'description': vessel['description'],
            'draft': vessel['draft']['value'],  # m
            'mass': mass_props['mass']['value'],  # kg
            'cog': [
                mass_props['center_of_gravity']['x']['value'],  # m
                mass_props['center_of_gravity']['y']['value'],  # m
                mass_props['center_of_gravity']['z']['value']   # m
            ],
            'inertia_matrix': [
                [mass_props['moments_of_inertia']['ixx']['value'], 0.0, 0.0],
                [0.0, mass_props['moments_of_inertia']['iyy']['value'], 0.0],
                [0.0, 0.0, mass_props['moments_of_inertia']['izz']['value']]
            ],
            'gyradii': [
                mass_props['radii_of_gyration']['kxx']['value'],  # m
                mass_props['radii_of_gyration']['kyy']['value'],  # m
                mass_props['radii_of_gyration']['kzz']['value']   # m
            ],
            'units': {
                'mass': 'kg',
                'length': 'm',
                'inertia': 'kg.m^2'
            }
        }
    
    return simplified


def main():
    """Main execution function."""
    # Set up paths
    script_dir = Path(__file__).parent
    data_dir = script_dir.parent / 'data'
    excel_file = data_dir / 'B1512 Gyradius Calcs.xlsx'
    
    # Output directory
    output_dir = script_dir.parent / 'outputs'
    output_dir.mkdir(exist_ok=True)
    
    print(f"Reading Excel file: {excel_file}")
    
    # Extract vessel properties (draft will be estimated if not provided)
    # You can specify a draft value here if known, e.g., draft_ft=6.5
    vessels = extract_vessel_properties(str(excel_file), draft_ft=None)
    
    # Get vessel dimensions from first configuration (they're the same for all)
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb['Calcs']
    vessel_dims = {
        'LBP': {
            'value': round(ws.cell(row=7, column=4).value * 0.3048, 2),
            'unit': 'm',
            'original': {'value': ws.cell(row=7, column=4).value, 'unit': 'ft'}
        },
        'beam': {
            'value': round(ws.cell(row=8, column=4).value * 0.3048, 2),
            'unit': 'm',
            'original': {'value': ws.cell(row=8, column=4).value, 'unit': 'ft'}
        },
        'depth': {
            'value': round(ws.cell(row=9, column=4).value * 0.3048, 2),
            'unit': 'm',
            'original': {'value': ws.cell(row=9, column=4).value, 'unit': 'ft'}
        }
    }
    wb.close()
    
    # Create detailed output
    detailed_output = {
        'hydrodynamic_properties': {
            'source': 'B1512 Gyradius Calcs.xlsx',
            'project': 'B1512 Nork & Mendez - Sea Cypress Station Keeping Analysis',
            'vessel_dimensions': vessel_dims,
            'vessel_configurations': vessels
        }
    }
    
    # Save detailed version
    detailed_file = output_dir / 'hydrodynamic_detailed.yml'
    with open(detailed_file, 'w') as f:
        yaml.dump(detailed_output, f, default_flow_style=False, sort_keys=False)
    print(f"Created detailed file: {detailed_file}")
    
    # Create and save simplified version
    simplified_output = create_simplified_format(vessels, vessel_dims)
    simplified_file = output_dir / 'hydrodynamic.yml'
    with open(simplified_file, 'w') as f:
        yaml.dump(simplified_output, f, default_flow_style=False, sort_keys=False)
    print(f"Created simplified file: {simplified_file}")
    
    # Print summary
    print("\n=== Summary of Configurations ===")
    for vessel in vessels:
        mass = vessel['mass_properties']['mass']['value']
        draft = vessel['draft']['value']
        print(f"\n{vessel['description']}:")
        print(f"  Draft: {draft:.2f} m ({vessel['draft']['original']['value']:.1f} ft)")
        print(f"  Mass: {mass:.2f} kg")
        print(f"  CoG: ({vessel['mass_properties']['center_of_gravity']['x']['value']:.3f}, "
              f"{vessel['mass_properties']['center_of_gravity']['y']['value']:.3f}, "
              f"{vessel['mass_properties']['center_of_gravity']['z']['value']:.3f}) m")


if __name__ == "__main__":
    main()