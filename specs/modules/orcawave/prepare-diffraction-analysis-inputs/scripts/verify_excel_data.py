#!/usr/bin/env python3
"""
Thoroughly verify the Excel data to understand the exact configurations.
"""

import openpyxl
from pathlib import Path

def verify_excel_data():
    """Verify all data from Excel file."""
    excel_file = Path('data/B1512 Gyradius Calcs_rev2.xlsx')
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb['Calcs']
    
    print("="*80)
    print("COMPLETE EXCEL DATA VERIFICATION")
    print("="*80)
    
    # Vessel dimensions
    print("\nVESSEL DIMENSIONS:")
    print(f"  LBP: {ws.cell(row=7, column=4).value} {ws.cell(row=7, column=5).value}")
    print(f"  Beam: {ws.cell(row=8, column=4).value} {ws.cell(row=8, column=5).value}")
    print(f"  Depth: {ws.cell(row=9, column=4).value} {ws.cell(row=9, column=5).value}")
    
    # Configuration 1 - Rows 12-28
    print("\n" + "="*80)
    print("CONFIGURATION 1 (Rows 12-28):")
    print("-"*80)
    
    # Check title/description
    print("Description/Title:")
    for row in range(12, 15):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col).value
            if cell and str(cell).strip():
                print(f"  Row {row}, Col {col}: {cell}")
    
    # Weight items
    print("\nWeight Items:")
    for row in range(15, 21):
        col1 = ws.cell(row=row, column=1).value
        col2 = ws.cell(row=row, column=2).value
        col3 = ws.cell(row=row, column=3).value
        col4 = ws.cell(row=row, column=4).value
        col5 = ws.cell(row=row, column=5).value
        if col2 or col3:
            print(f"  Row {row}: {col2} | {col3} | Weight: {col5}")
    
    # Total and CoG
    print("\nTotal and CoG (Row 21):")
    print(f"  Label: {ws.cell(row=21, column=4).value}")
    print(f"  Mass: {ws.cell(row=21, column=5).value} LT")
    print(f"  LCG: {ws.cell(row=21, column=6).value} ft")
    print(f"  TCG: {ws.cell(row=21, column=7).value} ft")
    print(f"  VCG: {ws.cell(row=21, column=8).value} ft")
    
    # Moments of inertia
    print("\nMoments of Inertia:")
    print(f"  Ixx: {ws.cell(row=26, column=3).value} LT-ft²")
    print(f"  Iyy: {ws.cell(row=27, column=3).value} LT-ft²")
    print(f"  Izz: {ws.cell(row=28, column=3).value} LT-ft²")
    
    # Configuration 2 - Rows 30-50
    print("\n" + "="*80)
    print("CONFIGURATION 2 (Rows 30-50):")
    print("-"*80)
    
    # Check title/description
    print("Description/Title:")
    for row in range(30, 34):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col).value
            if cell and str(cell).strip():
                print(f"  Row {row}, Col {col}: {cell}")
    
    # Weight items
    print("\nWeight Items:")
    for row in range(35, 43):
        col1 = ws.cell(row=row, column=1).value
        col2 = ws.cell(row=row, column=2).value
        col3 = ws.cell(row=row, column=3).value
        col4 = ws.cell(row=row, column=4).value
        col5 = ws.cell(row=row, column=5).value
        if col2 or col3:
            print(f"  Row {row}: {col2} | {col3} | Weight: {col5}")
    
    # Total and CoG
    print("\nTotal and CoG (Row 43):")
    print(f"  Label: {ws.cell(row=43, column=4).value}")
    print(f"  Mass: {ws.cell(row=43, column=5).value} LT")
    print(f"  LCG: {ws.cell(row=43, column=6).value} ft")
    print(f"  TCG: {ws.cell(row=43, column=7).value} ft")
    print(f"  VCG: {ws.cell(row=43, column=8).value} ft")
    
    # Moments of inertia
    print("\nMoments of Inertia:")
    print(f"  Ixx: {ws.cell(row=48, column=3).value} LT-ft²")
    print(f"  Iyy: {ws.cell(row=49, column=3).value} LT-ft²")
    print(f"  Izz: {ws.cell(row=50, column=3).value} LT-ft²")
    
    # Compare the two configurations
    print("\n" + "="*80)
    print("COMPARISON BETWEEN CONFIGURATIONS:")
    print("-"*80)
    
    mass1 = ws.cell(row=21, column=5).value
    mass2 = ws.cell(row=43, column=5).value
    tcg1 = ws.cell(row=21, column=7).value
    tcg2 = ws.cell(row=43, column=7).value
    
    print(f"Mass difference: {mass2 - mass1:.4f} LT")
    print(f"TCG shift: {tcg2 - tcg1:.4f} ft (Config 2 shifted to port)")
    
    if abs(tcg1) < 0.01:
        print("Config 1: Appears to be UPRIGHT (TCG near centerline)")
    if tcg2 > 0.1:
        print("Config 2: Appears to be FUEL OIL TO PORT (TCG shifted to port)")
    
    wb.close()
    
    print("\n" + "="*80)
    print("VERIFICATION COMPLETE")
    print("="*80)

if __name__ == "__main__":
    verify_excel_data()