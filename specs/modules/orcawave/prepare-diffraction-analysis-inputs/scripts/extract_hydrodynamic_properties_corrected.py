#!/usr/bin/env python3
"""
Extract hydrodynamic properties from Excel file - CORRECTED VERSION
Uses the "OrcaWave with FS Correction" rows for CoG values

This script reads vessel mass properties from the B1512 Gyradius Calcs_rev2.xlsx file
and generates a standardized hydrodynamic.yml file for use with OrcaWave, AQWA,
and other hydrodynamic analysis programs.
"""

import os
import yaml
import openpyxl
from pathlib import Path
from typing import Dict, List, Any
import math


def extract_vessel_properties(excel_path: str, draft_ft: float = None) -> List[Dict[str, Any]]:
    """
    Extract vessel properties from Excel file for all configurations.
    
    Args:
        excel_path: Path to the Excel file
        draft_ft: Vessel draft in feet (optional, will use estimated value if not provided)
        
    Returns:
        List of dictionaries containing vessel properties for each configuration
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Get vessel dimensions
    ws_calcs = wb['Calcs']
    lbp_ft = ws_calcs.cell(row=7, column=4).value  # 75 ft
    beam_ft = ws_calcs.cell(row=8, column=4).value  # 28 ft
    depth_ft = ws_calcs.cell(row=9, column=4).value  # 10 ft
    
    # If draft not provided, estimate as typical for supply vessel (60-70% of depth)
    if draft_ft is None:
        draft_ft = depth_ft * 0.65  # 6.5 ft estimated
        print(f"Note: Draft not specified, using estimated value of {draft_ft:.1f} ft based on vessel depth of {depth_ft} ft")
    
    # Configuration data for rev2 structure
    # IMPORTANT: Using row 22 and 44 for "OrcaWave with FS Correction" data
    config_data = [
        {
            'name': 'upright',
            'description': 'Upright Condition',
            'draft_ft': draft_ft,  # Base draft condition
            'orcawave_row': 22,  # OrcaWave with FS Correction row
            'mass_row': 25,
            'ixx_row': 26,
            'iyy_row': 27,
            'izz_row': 28
        },
        {
            'name': 'fuel_oil_to_port',
            'description': 'Fuel Oil to Port (4070 Gallons)',
            'draft_ft': draft_ft + 0.1,  # Slightly deeper due to fuel shift
            'orcawave_row': 44,  # OrcaWave with FS Correction row
            'mass_row': 47,
            'ixx_row': 48,
            'iyy_row': 49,
            'izz_row': 50
        }
    ]
    
    vessels = []
    
    for idx, config in enumerate(config_data):
        # Extract mass and CoG from "OrcaWave with FS Correction" row
        mass_lt = ws_calcs.cell(row=config['orcawave_row'], column=5).value
        lcg_ft = ws_calcs.cell(row=config['orcawave_row'], column=6).value
        tcg_ft = ws_calcs.cell(row=config['orcawave_row'], column=7).value
        vcg_ft = ws_calcs.cell(row=config['orcawave_row'], column=8).value
        
        # Extract moments of inertia (in LT-ft²) from calculation summary rows
        ixx_lt_ft2 = ws_calcs.cell(row=config['ixx_row'], column=3).value
        iyy_lt_ft2 = ws_calcs.cell(row=config['iyy_row'], column=3).value
        izz_lt_ft2 = ws_calcs.cell(row=config['izz_row'], column=3).value
        
        # Get draft for this configuration
        config_draft_ft = config['draft_ft']
        
        # Convert units to SI (metric)
        # 1 LT (long ton) = 1016.0469 kg
        # 1 ft = 0.3048 m
        LT_TO_KG = 1016.0469
        FT_TO_M = 0.3048
        
        mass_kg = mass_lt * LT_TO_KG
        lcg_m = lcg_ft * FT_TO_M
        tcg_m = tcg_ft * FT_TO_M
        vcg_m = vcg_ft * FT_TO_M
        draft_m = config_draft_ft * FT_TO_M
        
        # Convert moments of inertia from LT-ft² to kg-m²
        # 1 LT-ft² = 1016.0469 kg * (0.3048 m)² = 94.485 kg-m²
        LT_FT2_TO_KG_M2 = LT_TO_KG * (FT_TO_M ** 2)
        
        ixx_kg_m2 = ixx_lt_ft2 * LT_FT2_TO_KG_M2
        iyy_kg_m2 = iyy_lt_ft2 * LT_FT2_TO_KG_M2
        izz_kg_m2 = izz_lt_ft2 * LT_FT2_TO_KG_M2
        
        # Calculate gyradii from inertia and mass
        # k = sqrt(I/m)
        kxx_m = math.sqrt(ixx_kg_m2 / mass_kg)
        kyy_m = math.sqrt(iyy_kg_m2 / mass_kg)
        kzz_m = math.sqrt(izz_kg_m2 / mass_kg)
        
        print(f"\n{config['description']}:")
        print(f"  Using row {config['orcawave_row']} (OrcaWave with FS Correction)")
        print(f"  Mass: {mass_lt:.4f} LT")
        print(f"  CoG: ({lcg_ft:.4f}, {tcg_ft:.4f}, {vcg_ft:.4f}) ft")
        print(f"  Inertia: Ixx={ixx_lt_ft2:.2f}, Iyy={iyy_lt_ft2:.2f}, Izz={izz_lt_ft2:.2f} LT-ft²")
        
        # Round all values to 4 decimals for clean input files
        vessel_data = {
            'name': config['name'],
            'description': config['description'],
            'draft': {
                'value': round(draft_m, 4),
                'unit': 'm',
                'original': {
                    'value': round(config_draft_ft, 4),
                    'unit': 'ft'
                }
            },
            'mass_properties': {
                'mass': {
                    'value': round(mass_kg, 4),
                    'unit': 'kg',
                    'original': {
                        'value': round(mass_lt, 4),
                        'unit': 'LT'
                    }
                },
                'center_of_gravity': {
                    'x': {
                        'value': round(lcg_m, 4),
                        'unit': 'm',
                        'description': 'Longitudinal (from Fr. 0, +Fwd)',
                        'original': {
                            'value': round(lcg_ft, 4),
                            'unit': 'ft'
                        }
                    },
                    'y': {
                        'value': round(tcg_m, 4),
                        'unit': 'm',
                        'description': 'Transverse (from CL, +Port)',
                        'original': {
                            'value': round(tcg_ft, 4),
                            'unit': 'ft'
                        }
                    },
                    'z': {
                        'value': round(vcg_m, 4),
                        'unit': 'm',
                        'description': 'Vertical (from ABL, +Up)',
                        'original': {
                            'value': round(vcg_ft, 4),
                            'unit': 'ft'
                        }
                    }
                },
                'moments_of_inertia': {
                    'ixx': {
                        'value': round(ixx_kg_m2, 4),
                        'unit': 'kg.m^2',
                        'description': 'Roll moment of inertia',
                        'original': {
                            'value': round(ixx_lt_ft2, 4),
                            'unit': 'LT.ft^2'
                        }
                    },
                    'iyy': {
                        'value': round(iyy_kg_m2, 4),
                        'unit': 'kg.m^2',
                        'description': 'Pitch moment of inertia',
                        'original': {
                            'value': round(iyy_lt_ft2, 4),
                            'unit': 'LT.ft^2'
                        }
                    },
                    'izz': {
                        'value': round(izz_kg_m2, 4),
                        'unit': 'kg.m^2',
                        'description': 'Yaw moment of inertia',
                        'original': {
                            'value': round(izz_lt_ft2, 4),
                            'unit': 'LT.ft^2'
                        }
                    }
                },
                'radii_of_gyration': {
                    'kxx': {
                        'value': round(kxx_m, 4),
                        'unit': 'm',
                        'description': 'Roll radius of gyration',
                        'original': {
                            'value': round(kxx_m / FT_TO_M, 4),
                            'unit': 'ft'
                        }
                    },
                    'kyy': {
                        'value': round(kyy_m, 4),
                        'unit': 'm',
                        'description': 'Pitch radius of gyration',
                        'original': {
                            'value': round(kyy_m / FT_TO_M, 4),
                            'unit': 'ft'
                        }
                    },
                    'kzz': {
                        'value': round(kzz_m, 4),
                        'unit': 'm',
                        'description': 'Yaw radius of gyration',
                        'original': {
                            'value': round(kzz_m / FT_TO_M, 4),
                            'unit': 'ft'
                        }
                    }
                }
            }
        }
        
        vessels.append(vessel_data)
    
    wb.close()
    return vessels


def create_simplified_format(vessels: List[Dict[str, Any]], vessel_dims: Dict = None) -> Dict[str, Any]:
    """
    Create a simplified format suitable for direct use in hydrodynamic programs.
    
    Args:
        vessels: List of vessel configurations
        vessel_dims: Dictionary with vessel dimensions (LBP, beam, depth)
        
    Returns:
        Simplified dictionary format with 4-decimal rounding
    """
    simplified = {
        'vessel': {
            'name': 'Sea Cypress',
            'type': 'Supply Vessel',
            'configurations': {}
        }
    }
    
    if vessel_dims:
        simplified['vessel']['dimensions'] = vessel_dims
    
    for vessel in vessels:
        config_name = vessel['name']
        
        # Extract just the values in SI units
        mass_props = vessel['mass_properties']
        
        # Round all values to 4 decimals for clean formatting
        simplified['vessel']['configurations'][config_name] = {
            'description': vessel['description'],
            'draft': round(vessel['draft']['value'], 4),  # m
            'mass': round(mass_props['mass']['value'], 4),  # kg
            'cog': [
                round(mass_props['center_of_gravity']['x']['value'], 4),  # m
                round(mass_props['center_of_gravity']['y']['value'], 4),  # m
                round(mass_props['center_of_gravity']['z']['value'], 4)   # m
            ],
            'inertia_matrix': [
                [round(mass_props['moments_of_inertia']['ixx']['value'], 4), 0.0, 0.0],
                [0.0, round(mass_props['moments_of_inertia']['iyy']['value'], 4), 0.0],
                [0.0, 0.0, round(mass_props['moments_of_inertia']['izz']['value'], 4)]
            ],
            'gyradii': [
                round(mass_props['radii_of_gyration']['kxx']['value'], 4),  # m
                round(mass_props['radii_of_gyration']['kyy']['value'], 4),  # m
                round(mass_props['radii_of_gyration']['kzz']['value'], 4)   # m
            ],
            'units': {
                'mass': 'kg',
                'length': 'm',
                'inertia': 'kg.m^2'
            }
        }
    
    return simplified


def main():
    """Main execution function."""
    # Set up paths
    script_dir = Path(__file__).parent
    data_dir = script_dir.parent / 'data'
    excel_file = data_dir / 'B1512 Gyradius Calcs_rev2.xlsx'
    
    # Output directory
    output_dir = script_dir.parent / 'outputs'
    output_dir.mkdir(exist_ok=True)
    
    print(f"Reading Excel file: {excel_file}")
    print("="*60)
    
    # Extract vessel properties (draft will be estimated if not provided)
    vessels = extract_vessel_properties(str(excel_file), draft_ft=None)
    
    # Get vessel dimensions from first configuration (they're the same for all)
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb['Calcs']
    vessel_dims = {
        'LBP': {
            'value': round(ws.cell(row=7, column=4).value * 0.3048, 4),
            'unit': 'm',
            'original': {'value': round(ws.cell(row=7, column=4).value, 4), 'unit': 'ft'}
        },
        'beam': {
            'value': round(ws.cell(row=8, column=4).value * 0.3048, 4),
            'unit': 'm',
            'original': {'value': round(ws.cell(row=8, column=4).value, 4), 'unit': 'ft'}
        },
        'depth': {
            'value': round(ws.cell(row=9, column=4).value * 0.3048, 4),
            'unit': 'm',
            'original': {'value': round(ws.cell(row=9, column=4).value, 4), 'unit': 'ft'}
        }
    }
    wb.close()
    
    # Create detailed output
    detailed_output = {
        'hydrodynamic_properties': {
            'source': 'B1512 Gyradius Calcs_rev2.xlsx',
            'note': 'Using OrcaWave with FS Correction values',
            'project': 'B1512 Nork & Mendez - Sea Cypress Station Keeping Analysis',
            'vessel_dimensions': vessel_dims,
            'vessel_configurations': vessels
        }
    }
    
    # Save detailed version
    detailed_file = output_dir / 'hydrodynamic_detailed.yml'
    with open(detailed_file, 'w') as f:
        yaml.dump(detailed_output, f, default_flow_style=False, sort_keys=False, width=200)
    print(f"\nCreated detailed file: {detailed_file}")
    
    # Create and save simplified version
    simplified_output = create_simplified_format(vessels, vessel_dims)
    simplified_file = output_dir / 'hydrodynamic.yml'
    with open(simplified_file, 'w') as f:
        yaml.dump(simplified_output, f, default_flow_style=False, sort_keys=False, width=200)
    print(f"Created simplified file: {simplified_file}")
    
    # Print summary
    print("\n" + "="*60)
    print("SUMMARY OF CONFIGURATIONS (WITH FS CORRECTION)")
    print("="*60)
    for vessel in vessels:
        mass = vessel['mass_properties']['mass']['value']
        mass_lt = vessel['mass_properties']['mass']['original']['value']
        draft = vessel['draft']['value']
        print(f"\n{vessel['description']}:")
        print(f"  Draft: {draft:.4f} m ({vessel['draft']['original']['value']:.4f} ft)")
        print(f"  Mass: {mass:.4f} kg ({mass_lt:.4f} LT)")
        print(f"  CoG (m): ({vessel['mass_properties']['center_of_gravity']['x']['value']:.4f}, "
              f"{vessel['mass_properties']['center_of_gravity']['y']['value']:.4f}, "
              f"{vessel['mass_properties']['center_of_gravity']['z']['value']:.4f})")
        print(f"  CoG (ft): ({vessel['mass_properties']['center_of_gravity']['x']['original']['value']:.4f}, "
              f"{vessel['mass_properties']['center_of_gravity']['y']['original']['value']:.4f}, "
              f"{vessel['mass_properties']['center_of_gravity']['z']['original']['value']:.4f})")
        print(f"  Gyradii (m): ({vessel['mass_properties']['radii_of_gyration']['kxx']['value']:.4f}, "
              f"{vessel['mass_properties']['radii_of_gyration']['kyy']['value']:.4f}, "
              f"{vessel['mass_properties']['radii_of_gyration']['kzz']['value']:.4f})")


if __name__ == "__main__":
    main()