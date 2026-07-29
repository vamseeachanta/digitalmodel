import csv
import importlib.util
import os
from pathlib import Path
import subprocess
import sys

import pytest
from openpyxl import Workbook
import yaml


SCRIPT = (
    Path(__file__).parents[3]
    / "scripts"
    / "artificial_lift"
    / "extract_reference_catalogs.py"
)


def _load_extractor():
    spec = importlib.util.spec_from_file_location("extract_reference_catalogs", SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def _save_book(path, sheets):
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        worksheet = workbook.create_sheet(name)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)


def _source_tree(tmp_path, description="A-C-16-50-30"):
    root = tmp_path / "source"
    _save_book(
        root / "data/2018/Sucker Rod Pump Database 02_23_18.xlsx",
        {
            "Surface Unit Catalog": [
                [
                    "P.Unit Manf.",
                    "P.Unit Geom.",
                    "Gear Box Rating",
                    "Beam Rating",
                    "Max. S.Len.",
                    "Pumping Unit Description / Information",
                    'Dimen. "A"',
                    'Dimen. "C"',
                    'Dimen. "I"',
                    'Dimen. "K"',
                    'Dimen. "P"',
                ],
                ["A", "C", 16, 50, 30, description, 45, 45, 45, 61, 60],
                ["B", "C", 16, 50, 30, None, 45, 45, 45, 61, 60],
            ],
            "Rods Catalog": [
                [".", "DESC", "TENSILE", "AREA", "MOE", "VELOCITY", "DENSITY", "ELASTORQ"],
                [None, "grade - size", None, None, None, None, None, None],
                [0, '"97 - 0.875"', 140000, 0.601, 30.5, 16, 2.22, 0],
                ["separator", "catalog section", "-", "-", "-", "-", "-", "-"],
            ],
            "Rods Guide Catalog": [
                ["Manufacturer", "Model/Type", "Material"],
                ["Norris", "Standard", "PPA"],
                [None, None, "PPS"],
            ],
        },
    )
    _save_book(
        root / "REF/Rod Detail Table.xlsx",
        {
            "Sheet1": [
                ["Rod Grade", "Rod Area", "Unit Weight", "Modulus of Elast", "Spd of Sound", "Tensil Strength"],
                ["97  - 0.875", 0.601, 2.22, 30.5, 16, 140000],
                ["97-0.875", 0.601, 2.22, 30.5, 16, 140000],
            ]
        },
    )
    _save_book(
        root / "REF/Rod Coupling/nexus_catalog_couplings.xlsx",
        {
            "Sheet1": [
                [
                    "Rod Diameter (in.)",
                    "Coupling Manufacturer",
                    "Coupling Size",
                    "Coupling Type",
                    "Coupling Diameter (in.)",
                    "Coupling Length (in.)",
                    "Coupling Tensile (psi)",
                    "Friction Coefficient",
                ],
                [0.875, "Generic", "Full Size", "Spray Metal", 1.8125, 4, 90000, 0.2],
                [0.875, "N/A", "N/A", "Continuous Rod", 0.875, "N/A", "N/A", "N/A"],
            ]
        },
    )
    _save_book(
        root / "data/Rodpump Pumping Unit (1).xlsx",
        {
            "Sheet1": [
                [
                    "Id",
                    "Description",
                    "PumpingUnitManufacturer",
                    "PumpingUnitGeometry",
                    "GearBoxRating",
                    "BeamRating",
                    "MaxStrokeLength",
                    "DimensionalA",
                    "DimensionalC",
                    "DimensionalI",
                    "DimensionalK",
                    "DimensionalP",
                ],
                [1, "A-C-16-50-30", "A", "C", 16, 50, 30, 45, 45, 45, 61, 60],
                [2, None, "B", "C", 16, 50, 30, 45, 45, 45, 61, 60],
            ]
        },
    )
    _save_book(
        root / "data/2018/UniqueRodODData.xlsx",
        {
            "Rod ODs": [
                ["Rod OD", "Rod Connection Size"],
                [0.875, None],
                [0.875, "1+13/16"],
                [1, "1+1/16"],
                [1.25, "invalid"],
            ],
            "Look-up": [
                [None, None, None],
                [None, "Rod OD", "Coupling OD"],
                [None, "7/8", "1+13/16"],
                [None, 1.5, "1+5/8"],
            ],
        },
    )
    return root


def _csv_header(path):
    with path.open(newline="", encoding="utf-8") as stream:
        rows = (line for line in stream if not line.startswith("#"))
        return next(csv.reader(rows))


def test_extracts_allowlisted_columns_and_records_counts(tmp_path):
    extractor = _load_extractor()
    source = _source_tree(tmp_path)
    output = tmp_path / "v1"

    manifest = extractor.extract_catalogs(source, output, "2026-07-29")

    assert _csv_header(output / "rodpump_units.csv")[-1] != "UpdatedBy"
    assert "CreatedBy" not in _csv_header(output / "rodpump_units.csv")
    assert manifest["sources"]["rod_details"]["source_rows"] == 2
    assert manifest["sources"]["rod_details"]["emitted_rows"] == 1
    assert manifest["sources"]["rod_details"]["duplicate_rows"] == 1
    assert manifest["sources"]["rods_catalog"]["source_rows"] == 2
    assert manifest["sources"]["rods_catalog"]["emitted_rows"] == 1
    assert manifest["sources"]["rods_catalog"]["rejected_rows"] == 1
    coupling_text = (output / "couplings.csv").read_text()
    assert "Continuous Rod,0.875,,," in coupling_text
    assert manifest["sources"]["tubing"]["availability"] == "unavailable"
    assert manifest["sources"]["rod_connection_lookup"]["source_rows"] == 2
    assert manifest["sources"]["rod_connection_lookup"]["emitted_rows"] == 1
    assert manifest["sources"]["rod_connection_lookup"]["verified_rows"] == 1
    assert manifest["sources"]["rod_connection_lookup"]["quarantined_rows"] == 1
    assert manifest["sources"]["rod_connections"]["source_rows"] == 4
    assert manifest["sources"]["rod_connections"]["emitted_rows"] == 2
    assert manifest["sources"]["rod_connections"]["quarantined_rows"] == 2
    connections_text = (output / "rod_connections.csv").read_text()
    assert "1+13/16,1.8125,verified" in connections_text
    assert "1+1/16,,quarantined" in connections_text
    assert "invalid,,quarantined" in connections_text
    surface_source = manifest["sources"]["surface_unit_catalog"]
    assert surface_source["source_rows"] == 2
    assert surface_source["lookup_eligible_rows"] == 1
    assert surface_source["quarantined_rows"] == 1
    pump_source = manifest["sources"]["rodpump_units"]
    assert pump_source["source_rows"] == 2
    assert pump_source["lookup_eligible_rows"] == 1
    assert pump_source["quarantined_rows"] == 1
    surface_text = (output / "surface_unit_catalog.csv").read_text()
    assert ",lookup_eligible," in surface_text
    assert ",lookup_excluded,blank model key" in surface_text
    assert "catalog_velocity_relative_residual" in _csv_header(
        output / "rod_details.csv"
    )
    assert "catalog_velocity_relative_residual" in _csv_header(
        output / "rods_catalog.csv"
    )
    assert not (output / "tubing.csv").exists()
    lookup_text = (output / "rod_connection_lookup.csv").read_text()
    assert "0.875,1.8125,verified" in lookup_text
    assert "1.5,1.625,quarantined" in lookup_text


def test_extractor_is_byte_deterministic(tmp_path):
    extractor = _load_extractor()
    source = _source_tree(tmp_path)
    first = tmp_path / "first"
    second = tmp_path / "second"

    extractor.extract_catalogs(source, first, "2026-07-29")
    extractor.extract_catalogs(source, second, "2026-07-29")

    assert {
        path.relative_to(first): path.read_bytes()
        for path in first.rglob("*")
        if path.is_file()
    } == {
        path.relative_to(second): path.read_bytes()
        for path in second.rglob("*")
        if path.is_file()
    }


def test_cli_loads_helpers_from_its_own_worktree(tmp_path):
    source = _source_tree(tmp_path)
    output = tmp_path / "v1"
    environment = {"PATH": os.environ["PATH"]}

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--source-root",
            str(source),
            "--output-dir",
            str(output),
            "--extraction-date",
            "2026-07-29",
        ],
        cwd=tmp_path,
        env=environment,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0, result.stderr
    assert (output / "manifest.yml").is_file()


def test_check_mode_detects_catalog_drift(tmp_path):
    source = _source_tree(tmp_path)
    output = tmp_path / "v1"
    base_command = [
        sys.executable,
        str(SCRIPT),
        "--source-root",
        str(source),
        "--output-dir",
        str(output),
        "--extraction-date",
        "2026-07-29",
    ]
    environment = {"PATH": os.environ["PATH"]}
    assert subprocess.run(base_command, env=environment).returncode == 0
    assert subprocess.run(base_command + ["--check"], env=environment).returncode == 0
    (output / "rod_guides.csv").write_text("drift", encoding="utf-8")

    assert subprocess.run(base_command + ["--check"], env=environment).returncode != 0


@pytest.mark.parametrize(
    "description",
    [
        "well 30015410620000",
        "well 42-123-45678",
        "Smith Well",
        "well 300-15-410620-00-00",
        "well 42-123-45678-00-00",
        "rate 123 bopd",
        "rate 2 MBOPD",
        "rate 100 barrels/day",
        "host dynacard01.internal",
        "host db01.intranet",
        "source /mnt/client/private",  # abs-path-allowed
        "source /home/client/private",  # abs-path-allowed
        "source /opt/client/private",  # abs-path-allowed
        "source /srv/client/private",  # abs-path-allowed
        r"source \\server\client\private",
    ],
)
def test_extractor_redacts_prohibited_surface_keys(tmp_path, description):
    extractor = _load_extractor()
    source = _source_tree(tmp_path, description=description)
    output = tmp_path / "v1"

    manifest = extractor.extract_catalogs(source, output, "2026-07-29")

    output_text = (output / "surface_unit_catalog.csv").read_text()
    assert description not in output_text
    assert "lookup_excluded,prohibited model key removed" in output_text
    assert manifest["sources"]["surface_unit_catalog"]["quarantined_rows"] == 2


def test_extractor_preserves_safe_generic_well_sentinel(tmp_path):
    extractor = _load_extractor()
    source = _source_tree(tmp_path, description="No Unit on Well")
    output = tmp_path / "v1"

    extractor.extract_catalogs(source, output, "2026-07-29")

    output_text = (output / "surface_unit_catalog.csv").read_text()
    sentinel_line = next(
        line for line in output_text.splitlines() if "No Unit on Well" in line
    )
    assert "lookup_eligible" in sentinel_line
    assert "lookup_excluded" not in sentinel_line


def test_extractor_rejects_conflicting_normalized_rods(tmp_path):
    extractor = _load_extractor()
    source = _source_tree(tmp_path)
    workbook_path = source / "REF/Rod Detail Table.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Rod Grade", "Rod Area", "Unit Weight", "Modulus of Elast", "Spd of Sound", "Tensil Strength"])
    sheet.append(["97 - 0.875", 0.601, 2.22, 30.5, 16, 140000])
    sheet.append(["97-0.875", 0.601, 2.20, 30.5, 16, 140000])
    workbook.save(workbook_path)

    manifest = extractor.extract_catalogs(source, tmp_path / "v1", "2026-07-29")

    assert manifest["sources"]["rod_details"]["quarantined_rows"] == 2
    assert manifest["sources"]["rod_details"]["emitted_rows"] == 0
    quarantined = (tmp_path / "v1" / "rod_details_quarantine.csv").read_text()
    assert "conflicting normalized key" in quarantined
    assert "0.601,2.22,30.5,16,140000" in quarantined
    assert "0.601,2.2,30.5,16,140000" in quarantined


@pytest.mark.parametrize(
    "extraction_date",
    ["2026-7-29", "2026-07-29\n/mnt/client"],  # abs-path-allowed
)
def test_extractor_rejects_noncanonical_extraction_dates(
    tmp_path, extraction_date
):
    extractor = _load_extractor()

    with pytest.raises(ValueError, match="YYYY-MM-DD"):
        extractor.extract_catalogs(
            _source_tree(tmp_path), tmp_path / "v1", extraction_date
        )


def test_manifest_output_counts_and_hashes_match(tmp_path):
    extractor = _load_extractor()
    source = _source_tree(tmp_path)
    output = tmp_path / "v1"

    extractor.extract_catalogs(source, output, "2026-07-29")
    manifest = yaml.safe_load((output / "manifest.yml").read_text())

    for name, metadata in manifest["outputs"].items():
        path = output / name
        assert extractor.sha256_file(path) == metadata["sha256"]
        assert sum(
            1
            for line in path.read_text().splitlines()
            if line and not line.startswith("#")
        ) - 1 == metadata["row_count"]
