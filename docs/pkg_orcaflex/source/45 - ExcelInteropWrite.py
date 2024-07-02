import openpyxl

waveHeights = [0.5, 1.0, 2.0, 3.0, 4.0, 5.5]
wavePeriods = [6.0, 7.0, 8.0, 10.0]

book = openpyxl.Workbook()
sheet = book.active
sheet.title = "Data"
boldFont = openpyxl.styles.Font(bold=True)
centredAlignment = openpyxl.styles.Alignment(horizontal="center")
sheet["A1"] = "Height"
sheet["A1"].font = boldFont
sheet["A1"].alignment = centredAlignment
sheet["B1"] = "Period"
sheet["B1"].font = boldFont
sheet["B1"].alignment = centredAlignment

row = 2
for H in waveHeights:
    for T in wavePeriods:
        sheet.cell(row=row, column=1).value = H
        sheet.cell(row=row, column=2).value = T
        row += 1

book.save("scratch/LoadCaseData.xlsx")
